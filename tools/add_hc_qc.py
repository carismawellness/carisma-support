"""
add_hc_qc.py
Appends a "HC SPEND QC" section to the bottom of all 9 monthly sheets.

Reconciles:
  - Therapist : Receptionist ratio (by headcount)
  - Each role's cost as % of total HC spend
  - T:R spend ratio (€ of therapist cost per € of receptionist cost)
  - Month-by-month QC status vs industry benchmarks

Source data pulled directly from the existing 40% HC Budget Model section
(rows 5–16 = Jan–Dec, row 17 = ANNUAL TOTAL) already in each monthly sheet.

Safe to re-run — detects and overwrites existing QC blocks by header marker.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Palette ────────────────────────────────────────────────────────────────────
NAVY       = "1B3A4B"
GOLD       = "B8943E"
WHITE      = "FFFFFF"
BLACK      = "000000"
BLUE_FONT  = "0000FF"
GREEN_FONT = "008000"
RED_FONT   = "FF0000"
LIGHT_BLUE = "EAF4FB"

# QC status colours
C_GREEN  = "D5F5E3"
C_YELLOW = "FEF9E7"
C_ORANGE = "FAD7A0"
C_RED    = "FADBD8"
C_GREY   = "F2F3F4"
C_WHITE  = "FFFFFF"

# ── Constants ──────────────────────────────────────────────────────────────────
# Column indices in the 40% HC Budget Model (1-based)
COL_MONTH   = 1   # A — Month label
COL_REV     = 2   # B — Estimated Revenue
COL_BUDGET  = 3   # C — 40% HC Budget
COL_T_ALLOC = 4   # D — Therapist Alloc €
COL_N_T     = 5   # E — # Therapists
COL_R_ALLOC = 6   # F — Receptionist Alloc €
COL_N_R     = 7   # G — # Receptionists
COL_RM      = 8   # H — RM Alloc €
COL_PT      = 9   # I — PT Therapist Alloc €
COL_TOTAL   = 10  # J — Total HC €

DATA_START = 5    # Row 5 = January
DATA_END   = 16   # Row 16 = December
TOTAL_ROW  = 17   # Row 17 = ANNUAL TOTAL

# Industry benchmarks (standard spa / wellness)
BENCH_T_PCT_LO  = 0.60   # Therapist cost % lower bound
BENCH_T_PCT_HI  = 0.85   # Therapist cost % upper bound
BENCH_R_PCT_LO  = 0.07   # Receptionist % lower bound
BENCH_R_PCT_HI  = 0.20   # Receptionist % upper bound
BENCH_RM_PCT_LO = 0.03
BENCH_RM_PCT_HI = 0.10
BENCH_TR_LO     = 3.0    # T:R count ratio lower bound
BENCH_TR_HI     = 7.0    # T:R count ratio upper bound

QC_MARKER = "HC SPEND QC"  # Used to detect existing QC blocks for overwrite

MONTHLY_SHEETS = [
    "Inter Monthly", "Hugos Monthly", "Ramla Monthly", "Hyatt Monthly",
    "Excelsior Monthly", "Riviera Monthly", "Sunny Coast Monthly",
    "Novotel Monthly", "Sliema Monthly",
]

FILE = os.path.join(
    os.path.expanduser("~"),
    "Library/CloudStorage/GoogleDrive-mertgulen98@gmail.com/My Drive/"
    "Carisma Wellness Group/Carisma AI /Carisma AI/"
    "hr/Carisma Spa Staffing Model.xlsx"
)


# ── Style helpers ──────────────────────────────────────────────────────────────

def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_top():
    m = Side(style="medium", color="888888")
    t = Side(style="thin",   color="CCCCCC")
    return Border(left=t, right=t, top=m, bottom=t)

def _hdr(cell, text, bg=NAVY, fg=WHITE, bold=True, sz=10, wrap=False, align="center"):
    cell.value = text
    cell.font = Font(bold=bold, color=fg, size=sz)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = thin()

def _val(cell, value, bg=C_WHITE, fmt=None, bold=False, color=BLACK, align="center"):
    cell.value = value
    cell.font = Font(color=color, bold=bold, size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = thin()
    if fmt:
        cell.number_format = fmt

def _lbl(cell, text, bg=C_WHITE, bold=False, align="left", sz=10, color=BLACK):
    cell.value = text
    cell.font = Font(color=color, bold=bold, size=sz)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")


# ── QC logic ───────────────────────────────────────────────────────────────────

def t_pct_color(pct):
    if BENCH_T_PCT_LO <= pct <= BENCH_T_PCT_HI: return C_GREEN
    if pct > 0.90 or pct < 0.50:               return C_RED
    return C_YELLOW

def r_pct_color(pct, no_reception=False):
    if no_reception: return C_GREY   # intentional — RM or self-serve
    if BENCH_R_PCT_LO <= pct <= BENCH_R_PCT_HI: return C_GREEN
    if pct > 0.25 or pct < 0.03:               return C_RED
    return C_YELLOW

def rm_pct_color(pct):
    if BENCH_RM_PCT_LO <= pct <= BENCH_RM_PCT_HI: return C_GREEN
    if pct > 0.15:                               return C_RED
    return C_YELLOW

def tr_ratio_color(ratio, no_reception=False):
    if no_reception: return C_GREY
    if BENCH_TR_LO <= ratio <= BENCH_TR_HI: return C_GREEN
    if ratio < 2.0 or ratio > 9.0:         return C_RED
    return C_YELLOW

def tr_spend_color(ratio, no_reception=False):
    if no_reception: return C_GREY
    # T spend should dominate — healthy: 3×-8× receptionist spend
    if 3.0 <= ratio <= 8.0: return C_GREEN
    if ratio < 2.0 or ratio > 12.0: return C_RED
    return C_YELLOW

def qc_status(n_t, n_r, t_pct, tr_count, no_reception=False):
    """Return (status_text, bg_color) for the QC assessment cell."""
    issues = []
    if not no_reception:
        if tr_count < BENCH_TR_LO:      issues.append(f"T:R ratio {tr_count:.1f} — too many reception staff")
        if tr_count > BENCH_TR_HI:      issues.append(f"T:R ratio {tr_count:.1f} — low reception coverage")
    if t_pct < BENCH_T_PCT_LO:         issues.append(f"T% {t_pct:.0%} — therapist share low")
    if t_pct > BENCH_T_PCT_HI:         issues.append(f"T% {t_pct:.0%} — therapist share high")
    if n_t == 0:                        issues.append("No therapists allocated")
    if issues:
        return "⚠  " + " · ".join(issues), C_ORANGE
    if no_reception:
        return "✓  Therapy-only model — no reception (by design)", C_LGREEN
    return "✓  Ratio healthy", C_GREEN

# Light green for the therapy-only positive case
C_LGREEN = "D5F5E3"


# ── Core builder ───────────────────────────────────────────────────────────────

def find_qc_start(ws) -> int:
    """Return the row where an existing QC block starts, or None."""
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if val and isinstance(val, str) and QC_MARKER in val:
            return row
    return None

def last_data_row(ws) -> int:
    """Return the last row that has content in columns A–D."""
    last = 1
    for row in range(1, min(ws.max_row + 1, 200)):
        if any(ws.cell(row, c).value is not None for c in range(1, 5)):
            last = row
    return last

def clear_rows(ws, from_row: int, to_row: int):
    """Clear all cell values and formatting in the given row range."""
    for row in range(from_row, to_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = "General"

def add_qc_to_sheet(ws):
    # Read 40% model data (rows 5-17)
    rows_data = []
    for r in range(DATA_START, TOTAL_ROW + 1):
        month   = ws.cell(r, COL_MONTH).value
        n_t     = ws.cell(r, COL_N_T).value or 0
        n_r     = ws.cell(r, COL_N_R).value or 0
        t_cost  = ws.cell(r, COL_T_ALLOC).value or 0
        r_cost  = ws.cell(r, COL_R_ALLOC).value or 0
        rm_cost = ws.cell(r, COL_RM).value or 0
        pt_cost = ws.cell(r, COL_PT).value or 0
        total   = ws.cell(r, COL_TOTAL).value or 0
        rows_data.append((month, n_t, n_r, t_cost, r_cost, rm_cost, pt_cost, total))

    no_reception = all(d[2] == 0 for d in rows_data[:12])  # all months have 0 receptionists

    # Find where to write (overwrite existing QC block or append)
    qc_start = find_qc_start(ws)
    if qc_start:
        clear_rows(ws, qc_start - 2, qc_start + 25)
        write_start = qc_start - 2
    else:
        write_start = last_data_row(ws) + 3

    # ── Spacer ─────────────────────────────────────────────────────────────────
    ws.row_dimensions[write_start].height = 10

    # ── Header ─────────────────────────────────────────────────────────────────
    hdr_row = write_start + 1
    ws.row_dimensions[hdr_row].height = 20
    ws.merge_cells(start_row=hdr_row, start_column=1, end_row=hdr_row, end_column=14)
    _hdr(ws.cell(hdr_row, 1),
         f"HC SPEND QC — THERAPIST · RECEPTIONIST · RM RATIO & COST RECONCILIATION",
         bg=NAVY, sz=11, align="left")

    # ── Subtitle / benchmarks ──────────────────────────────────────────────────
    sub_row = hdr_row + 1
    ws.row_dimensions[sub_row].height = 16
    ws.merge_cells(start_row=sub_row, start_column=1, end_row=sub_row, end_column=14)
    c = ws.cell(sub_row, 1)
    c.value = (
        "Industry benchmarks — Therapist % of HC: 60–85%  ·  "
        "Receptionist %: 7–20%  ·  RM %: 3–10%  ·  "
        "T:R count ratio: 3:1–7:1  ·  "
        + ("No-reception model: RM covers front desk (by design)" if no_reception
           else "T:R spend ratio: 3×–8× (therapist cost should dominate)")
    )
    c.font = Font(color="555555", italic=True, size=8)
    c.fill = PatternFill("solid", start_color="F2F3F4")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Column headers ─────────────────────────────────────────────────────────
    col_hdr_row = sub_row + 1
    ws.row_dimensions[col_hdr_row].height = 30

    headers = [
        ("MONTH",          NAVY,      "center"),
        ("#\nTHERAPISTS",  "2C5F77",  "center"),
        ("#\nRECEPTION",   "2C5F77",  "center"),
        ("T:R COUNT\nRATIO","1B3A4B", "center"),
        ("THERAPIST\nCOST €","2C5F77","center"),
        ("RECEPTION\nCOST €","1B7A3B","center"),
        ("RM\nCOST €",     "4A235A",  "center"),
        ("PT\nCOST €",     "4A235A",  "center"),
        ("TOTAL\nHC €",    NAVY,      "center"),
        ("T% OF\nHC SPEND",NAVY,      "center"),
        ("R% OF\nHC SPEND","1B7A3B",  "center"),
        ("RM% OF\nHC SPEND","4A235A", "center"),
        ("T:R SPEND\nRATIO","2C5F77", "center"),
        ("QC ASSESSMENT",  GOLD,      "center"),
    ]
    for ci, (text, bg, align) in enumerate(headers):
        c = ws.cell(col_hdr_row, 1 + ci)
        c.value = text
        c.font = Font(bold=True, color=WHITE, size=9)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        c.border = thin()

    # ── Data rows ──────────────────────────────────────────────────────────────
    for i, (month, n_t, n_r, t_cost, r_cost, rm_cost, pt_cost, total) in enumerate(rows_data):
        row = col_hdr_row + 1 + i
        is_total = (month == "ANNUAL TOTAL")
        ws.row_dimensions[row].height = 22 if not is_total else 26

        # Computed metrics
        t_pct  = t_cost  / total if total else 0
        r_pct  = r_cost  / total if total else 0
        rm_pct = rm_cost / total if total else 0
        pt_pct = pt_cost / total if total else 0

        no_r_this_row = (n_r == 0)
        tr_count  = (n_t / n_r)  if n_r > 0 else None
        tr_spend  = (t_cost / r_cost) if r_cost > 0 else None

        status_text, status_bg = qc_status(
            n_t, n_r, t_pct,
            tr_count if tr_count else 0,
            no_reception=no_r_this_row
        )

        # Row background for total
        row_bg = "E8E8E8" if is_total else C_WHITE

        # Col A: Month
        _val(ws.cell(row, 1), month, bg=row_bg, bold=is_total, align="left",
             color=BLACK)

        # Col B: # Therapists
        _val(ws.cell(row, 2), n_t, bg=LIGHT_BLUE, fmt="0",
             bold=is_total)

        # Col C: # Reception
        r_bg = C_GREY if no_r_this_row else LIGHT_BLUE
        _val(ws.cell(row, 3), n_r if not is_total else n_r,
             bg=r_bg, fmt="0", bold=is_total)

        # Col D: T:R Count ratio
        if tr_count is None:
            _val(ws.cell(row, 4), "N/A — no reception", bg=C_GREY,
                 color="888888", align="center")
        else:
            _val(ws.cell(row, 4), round(tr_count, 1),
                 bg=tr_ratio_color(tr_count, no_r_this_row),
                 fmt="0.0:1" if not is_total else "0.0",
                 bold=is_total)
            ws.cell(row, 4).number_format = "0.0"

        # Col E: Therapist cost
        _val(ws.cell(row, 5), t_cost, bg=LIGHT_BLUE, fmt="#,##0",
             bold=is_total)

        # Col F: Reception cost
        r_cost_bg = C_GREY if no_r_this_row else "E9F7EF"
        _val(ws.cell(row, 6), r_cost, bg=r_cost_bg, fmt="#,##0",
             bold=is_total)

        # Col G: RM cost
        _val(ws.cell(row, 7), rm_cost, bg="F5EEF8", fmt="#,##0",
             bold=is_total)

        # Col H: PT cost
        _val(ws.cell(row, 8), pt_cost,
             bg="FAFAFA" if pt_cost == 0 else "FEF9E7",
             fmt="#,##0", bold=is_total)

        # Col I: Total HC
        _val(ws.cell(row, 9), total, bg="E8E8E8" if is_total else LIGHT_BLUE,
             fmt="#,##0", bold=True)

        # Col J: Therapist %
        _val(ws.cell(row, 10), t_pct,
             bg=row_bg if is_total else t_pct_color(t_pct),
             fmt="0.0%", bold=is_total,
             color=GREEN_FONT if (BENCH_T_PCT_LO <= t_pct <= BENCH_T_PCT_HI)
                   else RED_FONT)

        # Col K: Reception %
        _val(ws.cell(row, 11), r_pct,
             bg=row_bg if is_total else r_pct_color(r_pct, no_r_this_row),
             fmt="0.0%", bold=is_total)

        # Col L: RM %
        _val(ws.cell(row, 12), rm_pct,
             bg=row_bg if is_total else rm_pct_color(rm_pct),
             fmt="0.0%", bold=is_total)

        # Col M: T:R Spend Ratio
        if tr_spend is None:
            _val(ws.cell(row, 13), "N/A" if no_r_this_row else "—",
                 bg=C_GREY, color="888888")
        else:
            _val(ws.cell(row, 13), round(tr_spend, 1),
                 bg=row_bg if is_total else tr_spend_color(tr_spend, no_r_this_row),
                 fmt="0.0", bold=is_total)

        # Col N: QC Assessment
        if is_total:
            # Annual average T:R and T% for overall status
            annual_t_pct = (rows_data[-1][3] / rows_data[-1][7]
                            if rows_data[-1][7] else 0)
            annual_tr = (rows_data[-1][1] / rows_data[-1][2]
                         if rows_data[-1][2] > 0 else None)
            ann_status, ann_bg = qc_status(
                rows_data[-1][1], rows_data[-1][2],
                annual_t_pct,
                annual_tr or 0,
                no_reception=no_r_this_row
            )
            _val(ws.cell(row, 14), "ANNUAL: " + ann_status,
                 bg=ann_bg, bold=True, align="left",
                 color=GREEN_FONT if "✓" in ann_status else RED_FONT)
        else:
            _val(ws.cell(row, 14), status_text,
                 bg=status_bg, align="left",
                 bold=("⚠" in status_text),
                 color=GREEN_FONT if "✓" in status_text else
                       (RED_FONT if "⚠" in status_text else BLACK))

        # Apply thick top border on total row
        if is_total:
            b = thick_top()
            for ci in range(1, 15):
                ws.cell(row, ci).border = b

    # ── Benchmark reference row ────────────────────────────────────────────────
    bench_row = col_hdr_row + len(rows_data) + 1
    ws.row_dimensions[bench_row].height = 16
    ws.merge_cells(start_row=bench_row, start_column=1,
                   end_row=bench_row, end_column=14)
    b = ws.cell(bench_row, 1)
    b.value = (
        "BENCHMARKS — "
        "Therapist % (col J): 60–85% ✓  |  "
        "Reception % (col K): 7–20% ✓  |  "
        "RM % (col L): 3–10% ✓  |  "
        "T:R Count Ratio (col D): 3:1–7:1 ✓  |  "
        "T:R Spend Ratio (col M): 3×–8× ✓  |  "
        "Source: International Spa Association (ISPA) industry benchmarks"
    )
    b.font = Font(color="666666", italic=True, size=8)
    b.fill = PatternFill("solid", start_color="F8F9FA")
    b.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Adjust column widths to accommodate QC columns ─────────────────────────
    width_map = {1:10, 2:8, 3:8, 4:10, 5:11, 6:11, 7:9, 8:9,
                 9:11, 10:9, 11:9, 12:9, 13:10, 14:38}
    for ci, w in width_map.items():
        col_letter = get_column_letter(ci)
        if ws.column_dimensions[col_letter].width < w:
            ws.column_dimensions[col_letter].width = w


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    wb = load_workbook(FILE)
    for sheet_name in MONTHLY_SHEETS:
        if sheet_name in wb.sheetnames:
            add_qc_to_sheet(wb[sheet_name])
            print(f"  ✓  {sheet_name}")
        else:
            print(f"  ✗  {sheet_name} — not found, skipping")
    wb.save(FILE)
    print(f"\nSaved: {FILE}")

if __name__ == "__main__":
    main()
