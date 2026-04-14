"""Build actuals into Carisma Spa Staffing Model.xlsx
- Staffing Model sheet: Add cols N-U (actuals vs 40% HC budget)
- Weekly Roster sheet: Add actual rosters in cols L-U (real names + shifts)
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

FILE = os.path.join(os.path.dirname(__file__), '../hr/Carisma Spa Staffing Model.xlsx')

NAVY   = "1B3A4B"
GOLD   = "B8943E"
WHITE  = "FFFFFF"
BLUE   = "0000FF"
GRAY   = "D9D9D9"
A_CLR  = "BDD7EE"
B_CLR  = "C6EFCE"
OFF_CLR= "FCE4D6"

def nv_fill(color): return PatternFill("solid", fgColor=color)
def nv_font(bold=False, color=None, size=10): return Font(bold=bold, color=color or "000000", size=size)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def right(): return Alignment(horizontal="right", vertical="center")


# ─────────────────────────────────────────
# STAFFING MODEL ACTUALS (cols N-U = 14-21)
# ─────────────────────────────────────────

def build_staffing_model(wb):
    ws = wb['Staffing Model']

    N, O, P, Q, R, S, T, U = 14, 15, 16, 17, 18, 19, 20, 21

    # Row 4: section header spanning N–U
    ws.merge_cells(start_row=4, start_column=N, end_row=4, end_column=U)
    c = ws.cell(row=4, column=N)
    c.value = "◀  ACTUALS — APR 2026 (REAL HEADCOUNT)  ▶"
    c.font = nv_font(bold=True, color=WHITE, size=11)
    c.fill = nv_fill(GOLD)
    c.alignment = center()

    # Row 5: sub-headers
    sub = [
        (N, "2026 Target Rev (€/mo)"),
        (O, "Actual Staff Mix"),
        (P, "Actual FTE"),
        (Q, "Gross HC (€/mo)"),
        (R, "Loaded HC (×1.118)"),
        (S, "HC % of Rev"),
        (T, "vs 40% Delta (€)"),
        (U, "Status"),
    ]
    for col, label in sub:
        c = ws.cell(row=5, column=col)
        c.value = label
        c.font = nv_font(bold=True, color=WHITE, size=9)
        c.fill = nv_fill(NAVY)
        c.alignment = center()

    # Actuals data: row → (2026_rev, staff_mix, fte, gross_hc)
    data = {
        6:  (62610,  "7T+3A+1S+1RA",      12.0,  25900),  # Inter
        7:  (58800,  "7T+2A+1S+1PT",      10.5,  21750),  # Hugos
        8:  (39003,  "6T+2A+1S+1RA+1PT",  10.5,  19100),  # Ramla
        9:  (30290,  "5T+1S",              6.0,   11600),  # Hyatt
        10: (29363,  "4T+1A+1S+1RA+1PT",  7.5,   14400),  # Excelsior
        11: (25513,  "3T+1A+1S",           5.0,   10000),  # Riviera
        12: (24078,  "4T+1A+1S",           6.0,   11800),  # Sunny Coast
        13: (16667,  "2T+1A+1S",           4.0,    6800),  # Novotel
        14: (43250,  "Opening May 2026",   None,   None),  # Sliema
    }

    for row, (rev, mix, fte, gross) in data.items():
        # N: 2026 Rev
        c = ws.cell(row=row, column=N)
        c.value = rev; c.number_format = '#,##0'; c.alignment = right()

        # O: Staff Mix
        c = ws.cell(row=row, column=O)
        c.value = mix; c.alignment = Alignment(horizontal="center")

        # P: FTE
        c = ws.cell(row=row, column=P)
        c.value = fte; c.number_format = '0.0'; c.alignment = center()

        # Q: Gross HC (blue = hardcoded input)
        c = ws.cell(row=row, column=Q)
        c.value = gross
        c.font = Font(color=BLUE)
        c.number_format = '#,##0'; c.alignment = right()

        if gross is not None:
            # R: Loaded HC
            c = ws.cell(row=row, column=R)
            c.value = f'=Q{row}*1.118'; c.number_format = '#,##0'; c.alignment = right()

            # S: HC % of Rev
            c = ws.cell(row=row, column=S)
            c.value = f'=R{row}/N{row}'; c.number_format = '0.0%'; c.alignment = center()

            # T: vs 40% Delta (positive = surplus, negative = over budget)
            c = ws.cell(row=row, column=T)
            c.value = f'=N{row}*0.4-R{row}'
            c.number_format = '#,##0;(#,##0);"-"'; c.alignment = right()

            # U: Status
            c = ws.cell(row=row, column=U)
            c.value = f'=IF(S{row}>0.45,"❌ OVER",IF(S{row}<0.35,"⚠ UNDER","✓ OK"))'
            c.alignment = center()
        else:
            for col in [R, S, T]:
                ws.cell(row=row, column=col).value = "—"
            ws.cell(row=row, column=U).value = "⏳ OPENS MAY"
            ws.cell(row=row, column=U).alignment = center()

    # Row 15: Portfolio totals
    for col, val, fmt, bold_blue in [
        (N, '=SUM(N6:N14)', '#,##0',  False),
        (O, '9 locations',  '@',       False),
        (P, '=SUM(P6:P14)', '0.0',     False),
        (Q, '=SUM(Q6:Q14)', '#,##0',   True),
        (R, '=SUM(R6:R14)', '#,##0',   False),
        (S, '=R15/N15',     '0.0%',    False),
        (T, '=N15*0.4-R15', '#,##0;(#,##0);"-"', False),
        (U, '=IF(S15>0.45,"❌ OVER",IF(S15<0.35,"⚠ UNDER","✓ OK"))', '@', False),
    ]:
        c = ws.cell(row=15, column=col)
        c.value = val
        c.number_format = fmt
        c.font = Font(bold=True, color=BLUE if bold_blue else "000000")
        c.alignment = right() if col in [N, Q, R, T] else center()

    # Column widths
    widths = {N: 18, O: 22, P: 10, Q: 16, R: 16, S: 14, T: 18, U: 14}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ─────────────────────────────────────────
# WEEKLY ROSTER ACTUALS (cols L-U = 12-21)
# ─────────────────────────────────────────

def shift_fill(val):
    if val == 'A':   return nv_fill(A_CLR)
    if val == 'B':   return nv_fill(B_CLR)
    if val == 'OFF': return nv_fill(OFF_CLR)
    return None

def write_cell(ws, row, col, val, bold=False, color=None, fill=None, align=None, fmt=None):
    c = ws.cell(row=row, column=col)
    c.value = val
    c.font = Font(bold=bold, color=color or "000000")
    if fill: c.fill = fill
    if align: c.alignment = align
    if fmt:   c.number_format = fmt
    return c


def write_roster_actuals(ws, header_row, staff_rows):
    """Write one location's actuals section starting at header_row."""
    K, L, M, N, O, P, Q, R, S, T, U = 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21

    # Col K separator at header row
    write_cell(ws, header_row, K, "║ ACTUALS (APR 2026)",
               bold=True, color=WHITE, fill=nv_fill(GOLD), align=center())

    # Actuals header row (L-U)
    hdrs = ['STAFF (ACTUAL)', 'ROLE', 'MON', 'TUE', 'WED', 'THU', 'FRI ★', 'SAT ★', 'SUN ★', 'DAYS ON']
    for i, h in enumerate(hdrs):
        write_cell(ws, header_row, L + i, h,
                   bold=True, color=WHITE, fill=nv_fill(NAVY), align=center())

    # Staff rows
    shifts_cols = [N, O, P, Q, R, S, T]  # Mon-Sun
    for i, row_data in enumerate(staff_rows):
        r = header_row + 1 + i
        name, role, mon, tue, wed, thu, fri, sat, sun = row_data

        # Separator col K
        ws.cell(row=r, column=K).fill = nv_fill(GRAY)

        if not name:  # blank row
            continue

        # Name
        write_cell(ws, r, L, name)
        # Role
        write_cell(ws, r, M, role)
        # Shifts
        for col, val in zip(shifts_cols, [mon, tue, wed, thu, fri, sat, sun]):
            c = ws.cell(row=r, column=col)
            c.value = val
            c.alignment = Alignment(horizontal="center")
            f = shift_fill(val)
            if f: c.fill = f

        # Days ON
        days = sum(1 for v in [mon, tue, wed, thu, fri, sat, sun] if v not in ('OFF', '—', '', None))
        write_cell(ws, r, U, days if name and role else None, align=Alignment(horizontal="center"))


def build_weekly_roster(wb):
    ws = wb['Weekly Roster']

    # ── ROSTER DATA ──
    # (name, role, Mon, Tue, Wed, Thu, Fri, Sat, Sun)

    inter = [
        ('Milena',   'Top Therapist', 'OFF', 'OFF', 'A',   'A',   'A',   'A',   'B'),
        ('Valeri',   'Therapist',     'OFF', 'OFF', 'B',   'B',   'B',   'A',   'A'),
        ('Mini',     'Top Therapist', 'OFF', 'OFF', 'A',   'A',   'B',   'A',   'B'),
        ('Julie',    'Therapist',     'A',   'OFF', 'OFF', 'A',   'A',   'A',   'A'),
        ('Claudia',  'Therapist',     'B',   'OFF', 'OFF', 'B',   'B',   'A',   'B'),
        ('Chris',    'Therapist',     'A',   'A',   'OFF', 'OFF', 'A',   'A',   'A'),
        ('Anda',     'Therapist',     'B',   'B',   'OFF', 'OFF', 'B',   'A',   'B'),
        ('Maila',    'Advisor',       'OFF', 'A',   'A',   'A',   'A',   'A',   'OFF'),
        ('Romero',   'Advisor',       'A',   'OFF', 'A',   'A',   'A',   'A',   'OFF'),
        ('Gabriely', 'Advisor',       'A',   'A',   'OFF', 'A',   'A',   'A',   'OFF'),
        ('Maria',    'Reception',     'A',   'A',   'A',   'OFF', 'A',   'A',   'OFF'),
        ('Anna',     'Supervisor',    'OFF', 'A',   'A',   'A',   'A',   'A',   'OFF'),
    ]

    hugos = [
        ('Petra',   'Therapist',     'OFF', 'OFF', 'A',   'A',   'A',   'A',   'B'),
        ('Isadora', 'Therapist',     'OFF', 'OFF', 'B',   'B',   'B',   'A',   'A'),
        ('Carla',   'Therapist',     'OFF', 'OFF', 'A',   'A',   'A',   'A',   'A'),
        ('Linh',    'Therapist',     'A',   'OFF', 'OFF', 'A',   'A',   'A',   'A'),
        ('Rona',    'Therapist',     'B',   'OFF', 'OFF', 'B',   'B',   'A',   'B'),
        ('Tina',    'Therapist',     'A',   'A',   'OFF', 'OFF', 'A',   'A',   'A'),
        ('Mei',     'Therapist',     'B',   'B',   'OFF', 'OFF', 'B',   'A',   'B'),
        ('Alana',   'Advisor',       'OFF', 'A',   'A',   'A',   'A',   'A',   'OFF'),
        ('Anja',    'Advisor',       'A',   'A',   'OFF', 'A',   'A',   'A',   'OFF'),
        ('Jean',    'Advisor (PT)',  'OFF', 'OFF', 'OFF', 'OFF', 'A',   'A',   'A'),
        ('',        '',              '',    '',    '',    '',    '',    '',    ''),
        ('Neli',    'Supervisor',    'OFF', 'A',   'A',   'A',   'A',   'A',   'OFF'),
    ]

    ramla = [
        ('Preet',    'Therapist',  'OFF', 'OFF', 'A',   'A',   'A',   'A',   'B'),
        ('Yua',      'Therapist',  'OFF', 'OFF', 'B',   'B',   'B',   'A',   'A'),
        ('Lena',     'Therapist',  'A',   'OFF', 'OFF', 'A',   'A',   'A',   'A'),
        ('Dana',     'Therapist',  'B',   'OFF', 'OFF', 'B',   'B',   'A',   'B'),
        ('Sara',     'Therapist',  'A',   'A',   'OFF', 'OFF', 'A',   'A',   'A'),
        ('Bibi',     'Therapist',  'B',   'B',   'OFF', 'OFF', 'B',   'A',   'B'),
        ('Nathalia', 'Advisor',    'A',   'OFF', 'A',   'A',   'A',   'A',   'OFF'),
        ('Kristina', 'Supervisor', 'OFF', 'A',   'A',   'A',   'A',   'A',   'OFF'),
    ]

    hyatt = [
        ('Dita',     'Therapist',  'OFF', 'OFF', 'A',   'A',   'A',   'A',   'B'),
        ('Ria',      'Therapist',  'OFF', 'OFF', 'B',   'B',   'B',   'A',   'A'),
        ('Hina',     'Therapist',  'A',   'OFF', 'OFF', 'A',   'A',   'A',   'A'),
        ('Fae',      'Therapist',  'B',   'OFF', 'OFF', 'B',   'B',   'A',   'B'),
        ('Zola',     'Therapist',  'A',   'A',   'OFF', 'OFF', 'A',   'A',   'A'),
        ('Aakansha', 'Supervisor', 'OFF', 'A',   'A',   'A',   'A',   'A',   'OFF'),
    ]

    excelsior = [
        ('Aura',    'Therapist',  'OFF', 'OFF', 'A',   'A',   'A',   'A',   'B'),
        ('Mia',     'Therapist',  'OFF', 'OFF', 'B',   'B',   'B',   'A',   'A'),
        ('Faye',    'Therapist',  'A',   'OFF', 'OFF', 'A',   'A',   'A',   'A'),
        ('Jana',    'Therapist',  'B',   'OFF', 'OFF', 'B',   'B',   'A',   'B'),
        ('Sofia',   'Advisor',    'OFF', 'A',   'A',   'A',   'A',   'A',   'OFF'),
        ('Melanie', 'Supervisor', 'A',   'A',   'A',   'OFF', 'A',   'A',   'OFF'),
    ]

    riviera = [
        ('Rosa',   'Therapist',  'OFF', 'OFF', 'A',   'A',   'A',   'A',   'B'),
        ('Taia',   'Therapist',  'A',   'OFF', 'OFF', 'A',   'A',   'A',   'A'),
        ('Nina',   'Therapist',  'A',   'A',   'OFF', 'OFF', 'A',   'A',   'A'),
        ('Praise', 'Advisor',    'OFF', 'A',   'A',   'A',   'A',   'A',   'OFF'),
        ('Natalia','Supervisor', 'A',   'A',   'OFF', 'A',   'A',   'A',   'OFF'),
        ('',       '',           '',    '',    '',    '',    '',    '',    ''),
    ]

    odycy = [
        ('Vera',   'Therapist',  'OFF', 'OFF', 'A',   'A',   'A',   'A',   'B'),
        ('Kira',   'Therapist',  'A',   'OFF', 'OFF', 'A',   'A',   'A',   'A'),
        ('Nora',   'Therapist',  'A',   'A',   'OFF', 'OFF', 'A',   'A',   'A'),
        ('Dara',   'Therapist',  'B',   'B',   'OFF', 'OFF', 'B',   'A',   'B'),
        ('Lenara', 'Advisor',    'OFF', 'A',   'A',   'A',   'A',   'A',   'OFF'),
        # Row 111 (blank in original) used for supervisor:
        ('Jovana', 'Supervisor', 'A',   'A',   'A',   'OFF', 'A',   'A',   'OFF'),
    ]

    novotel = [
        ('Lara',    'Therapist',  'OFF', 'OFF', 'A',   'A',   'A',   'A',   'B'),
        ('Zara',    'Therapist',  'A',   'A',   'OFF', 'OFF', 'A',   'A',   'A'),
        ('Daniela', 'Advisor',    'OFF', 'A',   'A',   'A',   'A',   'A',   'OFF'),
        ('Natasha', 'Supervisor', 'A',   'A',   'A',   'OFF', 'A',   'A',   'OFF'),
    ]

    sliema = [
        ('— Opening May 2026 —', '', '', '', '', '', '', '', ''),
        ('TBA', 'Therapist',  '', '', '', '', '', '', ''),
        ('TBA', 'Therapist',  '', '', '', '', '', '', ''),
        ('TBA', 'Supervisor', '', '', '', '', '', '', ''),
    ]

    # Map: header_row → staff_list
    rosters = [
        (7,   inter),
        (27,  hugos),
        (47,  ramla),
        (63,  hyatt),
        (77,  excelsior),
        (91,  riviera),
        (105, odycy),
        (118, novotel),
        (129, sliema),
    ]

    for header_row, staff_list in rosters:
        write_roster_actuals(ws, header_row, staff_list)

    # Column widths for actuals section
    ws.column_dimensions['K'].width = 22  # separator
    ws.column_dimensions['L'].width = 18  # staff name
    ws.column_dimensions['M'].width = 16  # role
    for col_letter in ['N', 'O', 'P', 'Q', 'R', 'S', 'T']:
        ws.column_dimensions[col_letter].width = 7
    ws.column_dimensions['U'].width = 9   # days on


# ─────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────
if __name__ == '__main__':
    wb = load_workbook(FILE)
    build_staffing_model(wb)
    build_weekly_roster(wb)
    wb.save(FILE)
    print("✓ Saved:", FILE)
