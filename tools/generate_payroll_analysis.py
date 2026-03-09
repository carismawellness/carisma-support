"""
Payroll Cost Analysis Generator
Compiles 12 months of salary data (March 2025 - February 2026) from the Salary Master Sheet
and creates an Excel workbook with summary, monthly breakdown, and role-based pivot tables.

Methodology notes:
- Sep 2025 - Feb 2026: "Payroll" column (Z) from new-format calculation tabs = full payroll cost
- Mar 2025 - Aug 2025: "Bank transfer" column (V/W) from old-format tabs = payroll bank transfer portion
  (may slightly understate total cost for employees with cash payment components)
- Best practice: use the 6-month annualized figure (Sep-Feb) for decision-making
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from collections import defaultdict
import os

# ── Employee Master Data (from Staff Master tab) ──
# Format: name -> (role, status)
STAFF_MASTER = {
    "Milena Lazorova": ("Trainer Therapist", "Active"),
    "Ni Made Sayuti": ("Therapist", "Past employee"),
    "Melinda Liachi": ("Management", "Past employee"),
    "Anna Maria Mirisola": ("Spa Manager", "Active"),
    "Ni Made Ety Diantari": ("Therapist", "Active"),
    "Ni Made Sudarmini": ("Therapist", "Active"),
    "Melanie Mitic Vella": ("Spa Manager", "Active"),
    "Svetlana Matviets": ("Spa Manager", "Past employee"),
    "SARA VISCONTI": ("Therapist", "Past employee"),
    "Neli Radeva": ("Spa Manager", "Active"),
    "Viktorija Shojlevska": ("Therapist", "Past employee"),
    "Martina Vella": ("Therapist", "Past employee"),
    "Cindy Lorena Varon Prieto": ("Therapist", "Active"),
    "Arpit Sharma": ("Therapist", "Past employee"),
    "Kristina Alisauskaite": ("Manager", "Active"),
    "NATASHA MARJANOVIC": ("Trainer Therapist", "Active"),
    "Benjawan Phereewong": ("Therapist", "Active"),
    "Leticia Bonassi": ("Aesthetics", "Active"),
    "Ajlin Macheva": ("Spa Advisor / Concierge", "Past employee"),
    "Helene Abrev": ("Therapist", "Past employee"),
    "Veronica Astudillo": ("Spa Advisor / Concierge", "Past employee"),
    "Mariali Freites": ("Spa Advisor / Concierge", "Past employee"),
    "Lovely Sison": ("Therapist", "Active"),
    "Ana Paula": ("Spa Advisor / Concierge", "Past employee"),
    "Natasha Naumcheska": ("Therapist", "Active"),
    "Diana Romero": ("Therapist", "Past employee"),
    "IGOR GOLUBOVIC": ("Therapist", "Past employee"),
    "JOVANA MARKOVIC": ("Ass. Manager", "Active"),
    "LILIANA MONROY LOPEZ": ("Spa Advisor / Concierge", "Past employee"),
    "Can Artar": ("Growth", "Active"),
    "Sinan Tefik": ("Management", "Active"),
    "CLAUDIA GARCIA": ("Therapist", "Active"),
    "Dania Castillo": ("Therapist", "Past employee"),
    "CLAUDIA EDITH MENDEZ": ("Therapist", "Past employee"),
    "Paula Aramendez": ("Therapist", "Past employee"),
    "Putri Puspitha": ("Growth", "Past employee"),
    "Patricia Silveira": ("Therapist", "Past employee"),  # using ASCII name
    "Sanchi Mishra": ("Marketing Manager", "Past employee"),
    "BLERINA PETANI": ("Ass. Manager", "Past employee"),
    "VIOLETTA BORCZI": ("Spa Advisor / Concierge", "Past employee"),
    "Lisa Mhlanga": ("Therapist", "Past employee"),
    "CAROLINA HERMES": ("Therapist Advisor", "Past employee"),
    "Cosmin Alesteu": ("Growth", "Past employee"),
    "Prithwiraj Mohanan": ("Therapist", "Past employee"),
    "Aline Campos": ("Therapist", "Past employee"),
    "Patricia Eva": ("Spa Advisor / Concierge", "Past employee"),
    "Lucia Boldisova": ("Management", "Past employee"),
    "Ihebeddin Slama": ("Therapist", "Past employee"),
    "Nirmala Chhetri": ("Therapist", "Past employee"),
    "Rowena Madas": ("Support", "Past employee"),
    "Hafizh": ("Growth", "Past employee"),
    "Ege": ("Receptionist", "Past employee"),
    "FLORA SANTANA": ("Ass. Manager", "Active"),
    "VITORIA GOMES": ("Receptionist", "Past employee"),
    "Laura Camila": ("Therapist", "Active"),
    "Muriel Chammougon": ("Therapist", "Past employee"),
    "Jella Maye Garibay Almario": ("Aesthetics", "Past employee"),
    "DANIJELA JOVICIC": ("Therapist", "Past employee"),
    "ANGELINE MORALES": ("Therapist", "Past employee"),
    "MARTINA PLACZYNSKA": ("Therapist", "Past employee"),
    "Donna Whisken": ("Management", "Past employee"),
    "PHACHA TONNA": ("Therapist", "Past employee"),
    "SEBASTIJAN LOMSEK": ("Therapist", "Active"),
    "Kath Concio": ("CRM", "Active"),
    "Tamara Videc": ("Therapist", "Active"),
    "Valeri Kiseev": ("Trainer Therapist", "Active"),
    "Lourdes M. De Leon": ("Therapist", "Active"),
    "Romelia Reyes": ("Spa Advisor / Concierge", "Past employee"),
    "laura Martinez": ("Spa Advisor / Concierge", "Past employee"),
    "Hugh Grech": ("CRM", "Past employee"),
    "Samina Retezanu": ("Therapist", "Past employee"),
    "Tristan Prins": ("Therapist", "Past employee"),
    "FLAVIA MACEDO": ("Therapist", "Past employee"),
    "LORENA RODRIGUEZ": ("Therapist", "Past employee"),
    "EVELIN PARRA": ("Therapist", "Past employee"),
    "DANIELA LOPEZ": ("Therapist", "Past employee"),
    "Shana Marie Mifsud": ("Aesthetics", "Past employee"),
}

# ── Monthly payroll data ──
# Old format tabs (Mar-Aug 25): Column V = Bank transfer (payroll portion)
# New format tabs (Sep 25-Feb 26): Column Z = Payroll (full cost)

MONTHS = [
    "Mar 25", "Apr 25", "May 25", "Jun 25", "Jul 25", "Aug 25",
    "Sep 25", "Oct 25", "Nov 25", "Dec 25", "Jan 26", "Feb 26"
]

# ── Old format data (March - August 2025, from column V / W) ──

mar_25 = {
    "Milena Lazorova": 1578.04,
    "Melinda Liachi": 1450.28,
    "Anna Maria Mirisola": 1641.50,
    "Ni Made Ety Diantari": 1355.20,
    "Ni Made Sudarmini": 1359.73,
    "Melanie Mitic Vella": 1606.05,
    "Neli Radeva": 1665.68,
    "Cindy Lorena Varon Prieto": 1323.33,
    "Kristina Alisauskaite": 1323.26,
    "NATASHA MARJANOVIC": 1254.93,
    "Leticia Bonassi": 1189.28,
    "Ajlin Macheva": 1206.16,
    "Mariali Freites": 1133.79,
    "Lovely Sison": 1320.54,
    "Natasha Naumcheska": 1505.47,
    "IGOR GOLUBOVIC": 1207.00,
    "JOVANA MARKOVIC": 1324.20,
    "Sinan Tefik": 1474.68,
    "BLERINA PETANI": 1450.21,
    "VIOLETTA BORCZI": 2060.98,
    "Ihebeddin Slama": 1132.31,
    "FLORA SANTANA": 1164.26,
    "Laura Camila": 1200.05,
    "PHACHA TONNA": 1269.61,
    "SEBASTIJAN LOMSEK": 1259.70,
}

apr_25 = {
    "Milena Lazorova": 1402.08,
    "Melinda Liachi": 1402.08,
    "Anna Maria Mirisola": 1599.41,
    "Ni Made Ety Diantari": 1301.92,
    "Ni Made Sudarmini": 1302.92,
    "Melanie Mitic Vella": 1506.00,
    "Neli Radeva": 1623.41,
    "Cindy Lorena Varon Prieto": 1269.51,
    "Kristina Alisauskaite": 1270.51,
    "NATASHA MARJANOVIC": 1196.00,
    "Leticia Bonassi": 1180.00,
    "Ajlin Macheva": 1137.00,
    "Mariali Freites": 1059.24,
    "Lovely Sison": 1270.51,
    "Natasha Naumcheska": 1402.08,
    "IGOR GOLUBOVIC": 1135.00,
    "JOVANA MARKOVIC": 1270.52,
    "Sinan Tefik": 1423.63,
    "BLERINA PETANI": 1402.07,
    "Ihebeddin Slama": 1132.31,
    "FLORA SANTANA": 1210.76,
    "Laura Camila": 1137.00,
    "PHACHA TONNA": 1203.76,
    "SEBASTIJAN LOMSEK": 1204.76,
}

may_25 = {
    "Milena Lazorova": 1401.08,
    "Melinda Liachi": 1992.77,
    "Anna Maria Mirisola": 1598.41,
    "Ni Made Ety Diantari": 1302.92,
    "Ni Made Sudarmini": 1302.92,
    "Melanie Mitic Vella": 1505.00,
    "Neli Radeva": 1623.41,
    "Cindy Lorena Varon Prieto": 1252.59,
    "Kristina Alisauskaite": 1269.51,
    "NATASHA MARJANOVIC": 1197.00,
    "Leticia Bonassi": 1522.84,
    "Ajlin Macheva": 1270.51,
    "Mariali Freites": 1293.94,
    "Lovely Sison": 1269.51,
    "Natasha Naumcheska": 1190.91,
    "IGOR GOLUBOVIC": 1988.70,
    "JOVANA MARKOVIC": 1269.52,
    "Sinan Tefik": 1422.63,
    "BLERINA PETANI": 1681.39,
    "Ihebeddin Slama": 1132.31,
    "FLORA SANTANA": 1210.76,
    "Laura Camila": 1135.00,
    "PHACHA TONNA": 1203.76,
    "SEBASTIJAN LOMSEK": 1269.51,
}

jun_25 = {
    "Milena Lazorova": 1462.95,
    "Anna Maria Mirisola": 1652.22,
    "Ni Made Ety Diantari": 1366.13,
    "Ni Made Sudarmini": 1368.25,
    "Melanie Mitic Vella": 1565.10,
    "Neli Radeva": 1677.24,
    "Cindy Lorena Varon Prieto": 1335.61,
    "Kristina Alisauskaite": 1337.00,
    "NATASHA MARJANOVIC": 1265.71,
    "Leticia Bonassi": 1566.65,
    "Ajlin Macheva": 1337.00,
    "Lovely Sison": 1336.99,
    "Natasha Naumcheska": 1459.16,
    "JOVANA MARKOVIC": 1335.99,
    "Sinan Tefik": 1485.62,
    "Ihebeddin Slama": 1132.31,
    "FLORA SANTANA": 1286.24,
    "Laura Camila": 1210.90,
    "PHACHA TONNA": 1273.55,
    "SEBASTIJAN LOMSEK": 1337.00,
}

jul_25 = {
    "Milena Lazorova": 1402.08,
    "Anna Maria Mirisola": 1598.41,
    "Ni Made Ety Diantari": 1302.92,
    "Ni Made Sudarmini": 1302.92,
    "Melanie Mitic Vella": 1505.00,
    "Neli Radeva": 1623.40,
    "Cindy Lorena Varon Prieto": 1269.52,
    "Kristina Alisauskaite": 1335.32,
    "NATASHA MARJANOVIC": 1197.00,
    "Benjawan Phereewong": 1180.00,
    "Leticia Bonassi": 1502.84,
    "Ajlin Macheva": 1269.51,
    "Lovely Sison": 1269.52,
    "Natasha Naumcheska": 1401.08,
    "JOVANA MARKOVIC": 1270.51,
    "Sinan Tefik": 1623.40,
    "CLAUDIA GARCIA": 1180.00,
    "Ihebeddin Slama": 1132.31,
    "FLORA SANTANA": 1210.76,
    "Laura Camila": 1203.77,
    "PHACHA TONNA": 1203.76,
    "SEBASTIJAN LOMSEK": 1269.52,
}

aug_25 = {
    "Milena Lazorova": 1401.08,
    "Anna Maria Mirisola": 1599.41,
    "Ni Made Ety Diantari": 1302.92,
    "Ni Made Sudarmini": 1301.92,
    "Melanie Mitic Vella": 1505.00,
    "Neli Radeva": 1623.41,
    "Cindy Lorena Varon Prieto": 1270.51,
    "Kristina Alisauskaite": 1335.32,
    "NATASHA MARJANOVIC": 1197.00,
    "Benjawan Phereewong": 1180.00,
    "Leticia Bonassi": 1466.84,
    "Ajlin Macheva": 1270.51,
    "Lovely Sison": 1269.52,
    "Natasha Naumcheska": 1402.08,
    "JOVANA MARKOVIC": 1361.52,
    "Sinan Tefik": 1624.41,
    "CLAUDIA GARCIA": 1180.00,
    "Ihebeddin Slama": 1132.31,
    "FLORA SANTANA": 1210.76,
    "Laura Camila": 1204.76,
    "PHACHA TONNA": 2279.76,
    "SEBASTIJAN LOMSEK": 1269.52,
}

# ── New format data (Sep 25 - Feb 26, from column Z = Payroll) ──

sep_25 = {
    "Milena Lazorova": 2852,
    "Anna Maria Mirisola": 3975,
    "Ni Made Ety Diantari": 2532,
    "Ni Made Sudarmini": 2559,
    "Melanie Mitic Vella": 3040,
    "Neli Radeva": 3737,
    "Cindy Lorena Varon Prieto": 1955,
    "Kristina Alisauskaite": 2745,
    "NATASHA MARJANOVIC": 1905,
    "Benjawan Phereewong": 1815,
    "Leticia Bonassi": 3366,
    "Ajlin Macheva": 2198,
    "Lovely Sison": 2029,
    "Natasha Naumcheska": 2376,
    "JOVANA MARKOVIC": 2170,
    "Can Artar": 2088,
    "Sinan Tefik": 2868,
    "CLAUDIA GARCIA": 1663,
    "FLORA SANTANA": 2467,
    "Laura Camila": 1971,
    "SEBASTIJAN LOMSEK": 2007,
}

oct_25 = {
    "Milena Lazorova": 2776,
    "Anna Maria Mirisola": 2772,
    "Ni Made Ety Diantari": 2405,
    "Ni Made Sudarmini": 2618,
    "Melanie Mitic Vella": 2177,
    "Neli Radeva": 2954,
    "Cindy Lorena Varon Prieto": 2109,
    "Kristina Alisauskaite": 2512,
    "NATASHA MARJANOVIC": 1751,
    "Benjawan Phereewong": 1749,
    "Leticia Bonassi": 3182,
    "Ajlin Macheva": 1919,
    "Lovely Sison": 2353,
    "Natasha Naumcheska": 1945,
    "JOVANA MARKOVIC": 1500,
    "Can Artar": 1680,
    "Sinan Tefik": 1623,
    "CLAUDIA GARCIA": 1582,
    "FLORA SANTANA": 1863,
    "Laura Camila": 2256,
    "SEBASTIJAN LOMSEK": 1667,
}

nov_25 = {
    "Milena Lazorova": 2488,
    "Anna Maria Mirisola": 2793,
    "Ni Made Ety Diantari": 2224,
    "Ni Made Sudarmini": 2006,
    "Melanie Mitic Vella": 2219,
    "Neli Radeva": 2769,
    "Cindy Lorena Varon Prieto": 2079,
    "Kristina Alisauskaite": 2061,
    "NATASHA MARJANOVIC": 1913,
    "Benjawan Phereewong": 1649,
    "Leticia Bonassi": 4627,
    "Ajlin Macheva": 1756,
    "Lovely Sison": 2032,
    "Natasha Naumcheska": 2078,
    "JOVANA MARKOVIC": 1500,
    "Can Artar": 1850,
    "Sinan Tefik": 1623,
    "CLAUDIA GARCIA": 1378,
    "FLORA SANTANA": 1833,
    "Laura Camila": 1735,
    "SEBASTIJAN LOMSEK": 1767,
}

dec_25 = {
    "Milena Lazorova": 2615,
    "Anna Maria Mirisola": 3401,
    "Ni Made Ety Diantari": 2234,
    "Ni Made Sudarmini": 1866,
    "Melanie Mitic Vella": 2839,
    "Neli Radeva": 3779,
    "Cindy Lorena Varon Prieto": 1802,
    "Kristina Alisauskaite": 2815,
    "NATASHA MARJANOVIC": 2119,
    "Benjawan Phereewong": 1513,
    "Leticia Bonassi": 4456,
    "Ajlin Macheva": 2115,
    "Lovely Sison": 1711,
    "Natasha Naumcheska": 2040,
    "JOVANA MARKOVIC": 1500,
    "Can Artar": 2224,
    "Sinan Tefik": 1668,
    "CLAUDIA GARCIA": 1588,
    "FLORA SANTANA": 2265,
    "Laura Camila": 1795,
    "SEBASTIJAN LOMSEK": 1735,
}

jan_26 = {
    "Milena Lazorova": 2587,
    "Anna Maria Mirisola": 2835,
    "Ni Made Ety Diantari": 2275,
    "Ni Made Sudarmini": 2154,
    "Melanie Mitic Vella": 2166,
    "Neli Radeva": 2896,
    "Cindy Lorena Varon Prieto": 1793,
    "Kristina Alisauskaite": 2200,
    "NATASHA MARJANOVIC": 1689,
    "Benjawan Phereewong": 1663,
    "Leticia Bonassi": 4558,
    "Ajlin Macheva": 1770,
    "Lovely Sison": 1980,
    "Natasha Naumcheska": 2141,
    "JOVANA MARKOVIC": 1743,
    "Can Artar": 1850,
    "Sinan Tefik": 3600,
    "CLAUDIA GARCIA": 1596,
    "FLORA SANTANA": 1841,
    "Laura Camila": 1734,
    "SEBASTIJAN LOMSEK": 1635,
}

feb_26 = {
    "Milena Lazorova": 2589,
    "Anna Maria Mirisola": 3012,
    "Ni Made Ety Diantari": 2119,
    "Ni Made Sudarmini": 2474,
    "Melanie Mitic Vella": 2293,
    "Neli Radeva": 3011,
    "Cindy Lorena Varon Prieto": 1978,
    "Kristina Alisauskaite": 2049,
    "NATASHA MARJANOVIC": 1994,
    "Benjawan Phereewong": 1703,
    "Leticia Bonassi": 3758,
    "Ajlin Macheva": 1593,
    "Lovely Sison": 1938,
    "Natasha Naumcheska": 2285,
    "JOVANA MARKOVIC": 1894,
    "Can Artar": 1850,
    "Sinan Tefik": 3600,
    "CLAUDIA GARCIA": 1605,
    "FLORA SANTANA": 1837,
    "Laura Camila": 1822,
    "SEBASTIJAN LOMSEK": 1749,
}

ALL_MONTHS_DATA = [
    mar_25, apr_25, may_25, jun_25, jul_25, aug_25,
    sep_25, oct_25, nov_25, dec_25, jan_26, feb_26,
]


def compile_payroll():
    """Compile all monthly data into per-employee totals."""
    employees = {}  # name -> {months: [...], total: x, role: ..., status: ...}

    # Collect all unique employee names across all months
    all_names = set()
    for month_data in ALL_MONTHS_DATA:
        all_names.update(month_data.keys())

    for name in sorted(all_names):
        role, status = STAFF_MASTER.get(name, ("Unknown", "Unknown"))
        monthly = []
        for month_data in ALL_MONTHS_DATA:
            monthly.append(month_data.get(name, 0))

        total_12m = sum(monthly)
        months_active = sum(1 for v in monthly if v > 0)
        avg_monthly = total_12m / months_active if months_active > 0 else 0

        # Recent 6 months (Sep 25 - Feb 26) - more reliable data
        recent_6m = monthly[6:]  # indices 6-11
        total_recent_6m = sum(recent_6m)
        recent_months_active = sum(1 for v in recent_6m if v > 0)
        avg_recent = total_recent_6m / recent_months_active if recent_months_active > 0 else 0
        annualized = avg_recent * 12

        employees[name] = {
            "role": role,
            "status": status,
            "monthly": monthly,
            "total_12m": total_12m,
            "months_active": months_active,
            "avg_monthly": avg_monthly,
            "total_recent_6m": total_recent_6m,
            "recent_months_active": recent_months_active,
            "avg_recent_monthly": avg_recent,
            "annualized_from_recent": annualized,
        }

    return employees


def create_excel(employees, output_path):
    """Create the Excel workbook with multiple analysis sheets."""
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    active_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    past_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    euro_format = '#,##0'
    euro_format_dec = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

    def style_range(ws, start_row, end_row, max_col):
        for row in range(start_row, end_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

    # ═══════════════════════════════════════════════════
    # SHEET 1: PAYROLL SUMMARY (Active employees only)
    # ═══════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Active Staff Summary"

    headers1 = [
        "Employee Name", "Role", "Months on Payroll",
        "12-Month Total (EUR)", "Monthly Average (EUR)",
        "Recent 6M Total (EUR)", "Recent 6M Avg (EUR)",
        "Annualized Cost (EUR)"
    ]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header(ws1, 1, len(headers1))

    # Sort: active first, then by annualized cost descending
    active_employees = {k: v for k, v in employees.items() if v["status"] == "Active" and v["total_12m"] > 0}
    sorted_active = sorted(active_employees.items(), key=lambda x: -x[1]["annualized_from_recent"])

    row = 2
    total_12m_sum = 0
    total_recent_sum = 0
    total_annualized_sum = 0

    for name, data in sorted_active:
        ws1.cell(row=row, column=1, value=name)
        ws1.cell(row=row, column=2, value=data["role"])
        ws1.cell(row=row, column=3, value=data["months_active"])
        ws1.cell(row=row, column=4, value=round(data["total_12m"])).number_format = euro_format
        ws1.cell(row=row, column=5, value=round(data["avg_monthly"])).number_format = euro_format
        ws1.cell(row=row, column=6, value=round(data["total_recent_6m"])).number_format = euro_format
        ws1.cell(row=row, column=7, value=round(data["avg_recent_monthly"])).number_format = euro_format
        ws1.cell(row=row, column=8, value=round(data["annualized_from_recent"])).number_format = euro_format

        for col in range(1, len(headers1) + 1):
            ws1.cell(row=row, column=col).fill = active_fill
            ws1.cell(row=row, column=col).border = thin_border

        total_12m_sum += data["total_12m"]
        total_recent_sum += data["total_recent_6m"]
        total_annualized_sum += data["annualized_from_recent"]
        row += 1

    # Totals row
    ws1.cell(row=row, column=1, value="TOTAL ACTIVE STAFF")
    ws1.cell(row=row, column=1).font = Font(bold=True)
    ws1.cell(row=row, column=3, value=len(sorted_active))
    ws1.cell(row=row, column=4, value=round(total_12m_sum)).number_format = euro_format
    ws1.cell(row=row, column=6, value=round(total_recent_sum)).number_format = euro_format
    ws1.cell(row=row, column=8, value=round(total_annualized_sum)).number_format = euro_format
    for col in range(1, len(headers1) + 1):
        ws1.cell(row=row, column=col).fill = total_fill
        ws1.cell(row=row, column=col).font = Font(bold=True)
        ws1.cell(row=row, column=col).border = thin_border

    # Column widths
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 22
    for c in range(3, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(c)].width = 18

    # ═══════════════════════════════════════════════════
    # SHEET 2: ALL EMPLOYEES (including past)
    # ═══════════════════════════════════════════════════
    ws2 = wb.create_sheet("All Employees")

    headers2 = [
        "Employee Name", "Role", "Status", "Months on Payroll",
        "12-Month Total (EUR)", "Monthly Average (EUR)",
        "Recent 6M Total (EUR)", "Annualized Cost (EUR)"
    ]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, 1, len(headers2))

    sorted_all = sorted(employees.items(), key=lambda x: (-1 if x[1]["status"] == "Active" else 0, -x[1]["total_12m"]))

    row = 2
    grand_total = 0
    for name, data in sorted_all:
        if data["total_12m"] == 0:
            continue
        ws2.cell(row=row, column=1, value=name)
        ws2.cell(row=row, column=2, value=data["role"])
        ws2.cell(row=row, column=3, value=data["status"])
        ws2.cell(row=row, column=4, value=data["months_active"])
        ws2.cell(row=row, column=5, value=round(data["total_12m"])).number_format = euro_format
        ws2.cell(row=row, column=6, value=round(data["avg_monthly"])).number_format = euro_format
        ws2.cell(row=row, column=7, value=round(data["total_recent_6m"])).number_format = euro_format
        ws2.cell(row=row, column=8, value=round(data["annualized_from_recent"])).number_format = euro_format

        fill = active_fill if data["status"] == "Active" else past_fill
        for col in range(1, len(headers2) + 1):
            ws2.cell(row=row, column=col).fill = fill
            ws2.cell(row=row, column=col).border = thin_border

        grand_total += data["total_12m"]
        row += 1

    ws2.cell(row=row, column=1, value="GRAND TOTAL")
    ws2.cell(row=row, column=5, value=round(grand_total)).number_format = euro_format
    for col in range(1, len(headers2) + 1):
        ws2.cell(row=row, column=col).fill = total_fill
        ws2.cell(row=row, column=col).font = Font(bold=True)
        ws2.cell(row=row, column=col).border = thin_border

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 15
    for c in range(4, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 18

    # ═══════════════════════════════════════════════════
    # SHEET 3: MONTHLY BREAKDOWN
    # ═══════════════════════════════════════════════════
    ws3 = wb.create_sheet("Monthly Breakdown")

    headers3 = ["Employee Name", "Role", "Status"] + MONTHS + ["12M Total"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    style_header(ws3, 1, len(headers3))

    row = 2
    for name, data in sorted_all:
        if data["total_12m"] == 0:
            continue
        ws3.cell(row=row, column=1, value=name)
        ws3.cell(row=row, column=2, value=data["role"])
        ws3.cell(row=row, column=3, value=data["status"])
        for i, val in enumerate(data["monthly"]):
            cell = ws3.cell(row=row, column=4 + i, value=round(val) if val > 0 else "")
            if val > 0:
                cell.number_format = euro_format
        ws3.cell(row=row, column=4 + len(MONTHS), value=round(data["total_12m"])).number_format = euro_format

        fill = active_fill if data["status"] == "Active" else past_fill
        for col in range(1, len(headers3) + 1):
            ws3.cell(row=row, column=col).fill = fill
            ws3.cell(row=row, column=col).border = thin_border
        row += 1

    # Monthly totals row
    ws3.cell(row=row, column=1, value="MONTHLY TOTAL")
    ws3.cell(row=row, column=1).font = Font(bold=True)
    for i in range(len(MONTHS)):
        month_total = sum(d["monthly"][i] for d in employees.values())
        ws3.cell(row=row, column=4 + i, value=round(month_total)).number_format = euro_format
    grand = sum(d["total_12m"] for d in employees.values())
    ws3.cell(row=row, column=4 + len(MONTHS), value=round(grand)).number_format = euro_format
    for col in range(1, len(headers3) + 1):
        ws3.cell(row=row, column=col).fill = total_fill
        ws3.cell(row=row, column=col).font = Font(bold=True)
        ws3.cell(row=row, column=col).border = thin_border

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 22
    ws3.column_dimensions['C'].width = 15
    for c in range(4, len(headers3) + 1):
        ws3.column_dimensions[get_column_letter(c)].width = 12

    # ═══════════════════════════════════════════════════
    # SHEET 4: ROLE PIVOT - COST BY ROLE (Active only)
    # ═══════════════════════════════════════════════════
    ws4 = wb.create_sheet("Cost by Role")

    # Aggregate by role (active employees only)
    role_data = defaultdict(lambda: {"count": 0, "total_12m": 0, "annualized": 0, "names": []})
    for name, data in employees.items():
        if data["status"] != "Active" or data["total_12m"] == 0:
            continue
        role = data["role"]
        role_data[role]["count"] += 1
        role_data[role]["total_12m"] += data["total_12m"]
        role_data[role]["annualized"] += data["annualized_from_recent"]
        role_data[role]["names"].append(name)

    headers4 = [
        "Role", "Headcount", "12-Month Total (EUR)",
        "Annualized Total (EUR)", "Avg Annual per Person (EUR)",
        "% of Total Payroll", "Employees"
    ]
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=col, value=h)
    style_header(ws4, 1, len(headers4))

    total_annualized_all = sum(r["annualized"] for r in role_data.values())
    sorted_roles = sorted(role_data.items(), key=lambda x: -x[1]["annualized"])

    row = 2
    for role, rdata in sorted_roles:
        avg_per_person = rdata["annualized"] / rdata["count"] if rdata["count"] > 0 else 0
        pct = (rdata["annualized"] / total_annualized_all * 100) if total_annualized_all > 0 else 0

        ws4.cell(row=row, column=1, value=role)
        ws4.cell(row=row, column=2, value=rdata["count"])
        ws4.cell(row=row, column=3, value=round(rdata["total_12m"])).number_format = euro_format
        ws4.cell(row=row, column=4, value=round(rdata["annualized"])).number_format = euro_format
        ws4.cell(row=row, column=5, value=round(avg_per_person)).number_format = euro_format
        ws4.cell(row=row, column=6, value=round(pct, 1)).number_format = '0.0"%"'
        ws4.cell(row=row, column=7, value=", ".join(rdata["names"]))

        for col in range(1, len(headers4) + 1):
            ws4.cell(row=row, column=col).fill = subheader_fill
            ws4.cell(row=row, column=col).border = thin_border
        row += 1

    # Totals
    ws4.cell(row=row, column=1, value="TOTAL")
    ws4.cell(row=row, column=2, value=sum(r["count"] for r in role_data.values()))
    ws4.cell(row=row, column=3, value=round(sum(r["total_12m"] for r in role_data.values()))).number_format = euro_format
    ws4.cell(row=row, column=4, value=round(total_annualized_all)).number_format = euro_format
    ws4.cell(row=row, column=6, value=100.0).number_format = '0.0"%"'
    for col in range(1, len(headers4) + 1):
        ws4.cell(row=row, column=col).fill = total_fill
        ws4.cell(row=row, column=col).font = Font(bold=True)
        ws4.cell(row=row, column=col).border = thin_border

    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 12
    for c in range(3, 7):
        ws4.column_dimensions[get_column_letter(c)].width = 22
    ws4.column_dimensions['G'].width = 60

    # ═══════════════════════════════════════════════════
    # SHEET 5: ROLE PIVOT - DETAILED BREAKDOWN
    # ═══════════════════════════════════════════════════
    ws5 = wb.create_sheet("Role Detail")

    headers5 = [
        "Role", "Employee Name", "Base Salary (from Master)",
        "Recent 6M Avg Monthly (EUR)", "Annualized Cost (EUR)",
        "vs Base Salary %"
    ]
    for col, h in enumerate(headers5, 1):
        ws5.cell(row=1, column=col, value=h)
    style_header(ws5, 1, len(headers5))

    # Base salary lookup (approximate from Staff master)
    BASE_SALARIES = {
        "Milena Lazorova": 1700, "Anna Maria Mirisola": 2000,
        "Ni Made Ety Diantari": 1550, "Ni Made Sudarmini": 1550,
        "Melanie Mitic Vella": 1820, "Neli Radeva": 2000,
        "Cindy Lorena Varon Prieto": 1500, "Kristina Alisauskaite": 1600,
        "NATASHA MARJANOVIC": 1500, "Benjawan Phereewong": 1300,
        "Leticia Bonassi": 1800, "Ajlin Macheva": 1500,
        "Lovely Sison": 1500, "Natasha Naumcheska": 1700,
        "JOVANA MARKOVIC": 1500, "Can Artar": 1850,
        "Sinan Tefik": 3600, "CLAUDIA GARCIA": 1300,
        "FLORA SANTANA": 1400, "Laura Camila": 1400,
        "SEBASTIJAN LOMSEK": 1500,
    }

    row = 2
    for role, rdata in sorted_roles:
        # Sort employees within role by annualized cost
        role_employees = [(n, employees[n]) for n in rdata["names"]]
        role_employees.sort(key=lambda x: -x[1]["annualized_from_recent"])

        for i, (name, data) in enumerate(role_employees):
            ws5.cell(row=row, column=1, value=role if i == 0 else "")
            ws5.cell(row=row, column=2, value=name)
            base = BASE_SALARIES.get(name, "")
            ws5.cell(row=row, column=3, value=base)
            if base:
                ws5.cell(row=row, column=3).number_format = euro_format
            ws5.cell(row=row, column=4, value=round(data["avg_recent_monthly"])).number_format = euro_format
            ws5.cell(row=row, column=5, value=round(data["annualized_from_recent"])).number_format = euro_format
            if base and base > 0 and data["avg_recent_monthly"] > 0:
                pct_vs_base = (data["avg_recent_monthly"] / base - 1) * 100
                ws5.cell(row=row, column=6, value=round(pct_vs_base, 1)).number_format = '0.0"%"'

            for col in range(1, len(headers5) + 1):
                ws5.cell(row=row, column=col).border = thin_border
            row += 1

        # Role subtotal
        role_total_annual = sum(employees[n]["annualized_from_recent"] for n in rdata["names"])
        ws5.cell(row=row, column=1, value=f"  Subtotal: {role}")
        ws5.cell(row=row, column=1).font = Font(bold=True, italic=True)
        ws5.cell(row=row, column=5, value=round(role_total_annual)).number_format = euro_format
        for col in range(1, len(headers5) + 1):
            ws5.cell(row=row, column=col).fill = subheader_fill
            ws5.cell(row=row, column=col).border = thin_border
        row += 1

    ws5.column_dimensions['A'].width = 25
    ws5.column_dimensions['B'].width = 30
    for c in range(3, len(headers5) + 1):
        ws5.column_dimensions[get_column_letter(c)].width = 22

    # ═══════════════════════════════════════════════════
    # SHEET 6: METHODOLOGY NOTES
    # ═══════════════════════════════════════════════════
    ws6 = wb.create_sheet("Methodology")
    notes = [
        ["Payroll Cost Analysis - Methodology Notes"],
        [""],
        ["Data Source:", "Salary Master Sheet - Calculation (C) tabs"],
        ["Period:", "March 2025 - February 2026 (12 months)"],
        ["Generated:", "March 2026"],
        [""],
        ["DATA METHODOLOGY"],
        [""],
        ["September 2025 - February 2026 (6 months):"],
        ["  Column Z ('Payroll') from the new-format calculation tabs."],
        ["  This represents the full payroll cost: Base + Commissions + Bonuses - Deductions."],
        ["  This is the most reliable data for cost analysis."],
        [""],
        ["March 2025 - August 2025 (6 months):"],
        ["  Column V ('Bank transfer') from the old-format calculation tabs."],
        ["  This represents the bank transfer portion of payroll."],
        ["  May understate total cost for employees with cash payment components."],
        ["  The old format did not have a unified 'Payroll' column."],
        [""],
        ["RECOMMENDATION FOR ORG DESIGN:"],
        ["  Use the 'Annualized Cost' column (based on recent 6 months) as the primary metric."],
        ["  This reflects the current compensation structure including commissions."],
        ["  The annualized figure = average monthly cost (Sep-Feb) x 12."],
        [""],
        ["KEY DEFINITIONS"],
        ["  Payroll Cost: Total cost to company per employee per month (gross, before NI/FSS)"],
        ["  Annualized: Recent 6-month average projected to 12 months"],
        ["  Base Salary: Fixed monthly amount per contract"],
        ["  Commission: Variable pay based on sales/services performance"],
        [""],
        ["NOTES"],
        ["  - Active employees with no payroll data in a month may have been on leave"],
        ["  - Past employees show payroll only for months they were employed"],
        ["  - Commissions cause significant month-to-month variation"],
        ["  - Aesthetics role (Leticia) has high variability due to commission structure"],
    ]
    for r, row_data in enumerate(notes, 1):
        for c, val in enumerate(row_data, 1):
            cell = ws6.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True, size=14)
            elif val and val.isupper() and len(val) > 5:
                cell.font = Font(bold=True, size=11)
    ws6.column_dimensions['A'].width = 80

    # Save
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    employees = compile_payroll()
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(os.path.dirname(output_dir), ".tmp", "Payroll_Cost_Analysis_Mar25_Feb26.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    create_excel(employees, output_path)
    print(f"Excel file created: {output_path}")

    # Print summary
    active_total = sum(d["annualized_from_recent"] for d in employees.values() if d["status"] == "Active" and d["total_12m"] > 0)
    active_count = sum(1 for d in employees.values() if d["status"] == "Active" and d["total_12m"] > 0)
    print(f"\nActive employees with payroll: {active_count}")
    print(f"Total annualized payroll cost: EUR {active_total:,.0f}")
    print(f"Average annual cost per active employee: EUR {active_total/active_count:,.0f}" if active_count else "")
