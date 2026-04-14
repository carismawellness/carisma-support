"""
build_workload_check.py
Adds a "Workload Check" sheet to hr/Carisma Spa Staffing Model.xlsx

Reads therapist headcount per month directly from the Dynamic Staffing Plan sheet,
calculates TREATMENTS PER THERAPIST PER DAY for each location × month, and
colour-codes each cell to show whether the implied workload is feasible.

Formula:
  Tx / Therapist / Day = (Monthly Base Rev × Seasonal Idx) / (Avg Tx Rev × Days × Therapists)
"""

import re
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette ────────────────────────────────────────────────────────────────────
NAVY       = "1B3A4B"
GOLD       = "B8943E"
WHITE      = "FFFFFF"
BLACK      = "000000"
BLUE_FONT  = "0000FF"
GREEN_FONT = "008000"
RED_FONT   = "FF0000"
LIGHT_BLUE = "EAF4FB"

# Workload colour bands — treatments per therapist per day
C_GREY   = "F2F3F4"  # < 2.0  — severely under-utilised
C_LGREEN = "D5F5E3"  # 2.0–3.0 — light/comfortable
C_GREEN  = "A9DFBF"  # 3.0–4.0 — healthy / ideal target range
C_YELLOW = "FEF9E7"  # 4.0–5.0 — busy (sustainable)
C_ORANGE = "FAD7A0"  # 5.0–6.0 — at capacity limit
C_RED    = "FADBD8"  # > 6.0   — physically impossible

# ── Constants ──────────────────────────────────────────────────────────────────
MONTHS       = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
SEASONAL_IDX = [0.80, 0.85, 0.90, 0.95, 1.00, 1.10, 1.20, 1.25, 1.10, 1.00, 0.90, 0.95]
DAYS_IN_MO   = [31,   28,   31,   30,   31,   30,   31,   31,   30,   31,   30,   31]

AVG_TX_REV     = 85
MAX_TX_PER_DAY = 6    # per therapist per shift
UTIL_TARGET    = 0.80

# Monthly base revenues (annual ÷ 12, ex-VAT) — must match Staffing Model
REVENUES = {
    "Intercontinental": 49_475,
    "Hugos":            47_546,
    "Ramla":            29_886,
    "Hyatt":            25_358,
    "Excelsior":        24_263,
    "Riviera":          19_936,
    "Sunny Coast":      18_204,
    "Novotel":          14_000,
    "Sliema (NEW)":     41_667,
}

FILE = os.path.join(
    os.path.expanduser("~"),
    "Library/CloudStorage/GoogleDrive-mertgulen98@gmail.com/My Drive/"
    "Carisma Wellness Group/Carisma AI /Carisma AI/"
    "hr/Carisma Spa Staffing Model.xlsx"
)


# ── Helpers ────────────────────────────────────────────────────────────────────

def parse_therapists(cell_val: str) -> int:
    """Parse '5T', '5+3S', '⚠ 7T' etc. → total therapist count."""
    if not cell_val:
        return 0
    s = str(cell_val).replace("⚠", "").replace("△", "").strip()
    nums = re.findall(r'\d+', s)
    return sum(int(n) for n in nums) if nums else 0

def workload_color(tx_per_therapist: float) -> str:
    if tx_per_therapist > 6.0: return C_RED
    if tx_per_therapist > 5.0: return C_ORANGE
    if tx_per_therapist > 4.0: return C_YELLOW
    if tx_per_therapist > 3.0: return C_GREEN
    if tx_per_therapist > 2.0: return C_LGREEN
    return C_GREY

def is_understaffed(cell_val: str) -> bool:
    return cell_val and ("⚠" in str(cell_val) or "△" in str(cell_val))

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def med_border():
    m = Side(style="medium", color="888888")
    t = Side(style="thin",   color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=m)

def _hdr(cell, text, bg=NAVY, fg=WHITE, bold=True, sz=10, wrap=False, align="center"):
    cell.value = text
    cell.font = Font(bold=bold, color=fg, size=sz)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = thin_border()

def _lbl(cell, text, bold=False, bg=WHITE, align="left", sz=10, color=BLACK):
    cell.value = text
    cell.font = Font(color=color, bold=bold, size=sz)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")

def _val(cell, value, bg, fmt="0.0", bold=False, color=BLACK, note=""):
    cell.value = value
    cell.font = Font(color=color, bold=bold, size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell.border = thin_border()
    cell.number_format = fmt
    if note:
        from openpyxl.comments import Comment
        cell.comment = Comment(note, "Workload Check")

def _inp(cell, value, fmt=None, bg=WHITE, align="center"):
    cell.value = value
    cell.font = Font(color=BLUE_FONT, size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = thin_border()
    if fmt:
        cell.number_format = fmt


# ── Read Dynamic Staffing Plan ─────────────────────────────────────────────────

def read_dynamic_plan(wb) -> dict:
    """
    Returns {location_name: [t_jan, t_feb, ..., t_dec]} where each value is
    the total therapist count for that month parsed from the Dynamic Staffing Plan.
    """
    ws = wb["Dynamic Staffing Plan"]
    plan = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        loc = row[0]
        # Stop before the "Seasonal Contract Calendar" section — those rows have
        # text values (e.g. "Feb 15", "Oct 31", "Apr–Oct") that parse_therapists()
        # would misread as therapist counts (15, 31, 0…), corrupting Grid 1 & 2.
        if loc == "PORTFOLIO TOTAL ANNUAL SAVING":
            break
        if not loc or loc in (
            "Location",
            "SEASONAL CONTRACT CALENDAR  —  Fixed-term contracts typically run Apr 1 → Oct 31.  Recruitment to begin by Feb 15 each year.  Review and renew by Oct 1."
        ):
            continue
        # Columns 1-12 (0-indexed) = Jan-Dec
        month_vals = row[1:13]
        if all(v is None for v in month_vals):
            continue
        therapist_counts = [parse_therapists(v) for v in month_vals]
        if any(t > 0 for t in therapist_counts):
            # Normalise location name to match REVENUES dict
            clean = str(loc).replace("⚠ ", "").replace("⚠", "").strip()
            # Match to REVENUES key
            for rev_key in REVENUES:
                if rev_key.lower() in clean.lower() or clean.lower() in rev_key.lower():
                    plan[rev_key] = {
                        "therapists": therapist_counts,
                        "raw":        list(month_vals),
                        "original":   clean,
                    }
                    break
    return plan


# ── Main builder ───────────────────────────────────────────────────────────────

def build_workload_check(wb):
    if "Workload Check" in wb.sheetnames:
        del wb["Workload Check"]

    # Insert after "Utilisation Check" if present, else at end
    ws = wb.create_sheet("Workload Check")
    ws.sheet_view.showGridLines = False

    plan = read_dynamic_plan(wb)

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    for i in range(12):
        ws.column_dimensions[get_column_letter(2 + i)].width = 8
    ws.column_dimensions[get_column_letter(14)].width = 9   # FY Avg
    ws.column_dimensions[get_column_letter(15)].width = 38  # Note

    total_cols = 15

    def merge_full(row_num, text, bg=NAVY, fg=WHITE, sz=10, align="left", bold=True):
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num,   end_column=total_cols)
        c = ws.cell(row_num, 1)
        c.value = text
        c.font = Font(bold=bold, color=fg, size=sz)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        c.border = thin_border()

    def col_headers(row_num, bg_month="2C5F77", bg_stat="4A235A", bg_note=NAVY, extra_col=True):
        ws.row_dimensions[row_num].height = 28
        _hdr(ws.cell(row_num, 1), "LOCATION", bg=NAVY, sz=9, wrap=True)
        for i, m in enumerate(MONTHS):
            _hdr(ws.cell(row_num, 2 + i), m, bg=bg_month, sz=9)
        _hdr(ws.cell(row_num, 14), "FY\nAVG", bg=bg_stat, sz=9, wrap=True)
        if extra_col:
            _hdr(ws.cell(row_num, 15), "NOTE", bg=bg_note, sz=9, align="left")

    # ── Row 1: Title ───────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    merge_full(1,
        "CARISMA SPA & WELLNESS — THERAPIST WORKLOAD CHECK  |  "
        "Treatments Per Therapist Per Day · Dynamic Staffing Plan",
        sz=13)

    # ── Row 2: Subtitle ────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 16
    merge_full(2,
        "Formula: Monthly Base Revenue × Seasonal Index ÷ (€85 × Days in Month × Therapists that month)  |  "
        "Therapist counts sourced directly from the Dynamic Staffing Plan tab  |  "
        "Re-run build_workload_check.py to refresh",
        bg="F2F3F4", fg="555555", sz=8, bold=False)

    # ── Row 3: Spacer ──────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 8

    # ── Rows 4–7: Assumptions ─────────────────────────────────────────────────
    ws.row_dimensions[4].height = 20
    merge_full(4, "KEY ASSUMPTIONS  (Blue = hardcoded input)", bg=GOLD, sz=10)

    for i, (label, value, fmt, note_txt) in enumerate([
        ("Avg treatment revenue (€)",         AVG_TX_REV,     '"€"#,##0', "Mid-point €70–€100 range."),
        ("Max treatments / therapist / day",   MAX_TX_PER_DAY, "0",        "9-hr shift ÷ 90 min per slot = 6 max."),
        ("Peak utilisation target",            UTIL_TARGET,    "0%",       "80% industry benchmark."),
    ]):
        row = 5 + i
        ws.row_dimensions[row].height = 18
        _lbl(ws.cell(row, 1), label, bold=True)
        _inp(ws.cell(row, 2), value, fmt=fmt, bg=LIGHT_BLUE)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=total_cols)
        c = ws.cell(row, 3)
        c.value = "← " + note_txt
        c.font = Font(color="888888", italic=True, size=8)
        c.fill = PatternFill("solid", start_color="FAFAFA")
        c.alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 8: Spacer ──────────────────────────────────────────────────────────
    ws.row_dimensions[8].height = 8

    # ─────────────────────────────────────────────────────────────────────────
    # GRID 1 — Therapist Headcount by Month (from Dynamic Staffing Plan)
    # ─────────────────────────────────────────────────────────────────────────
    G1_HDR  = 9
    G1_COLS = 10
    G1_DATA = 11

    ws.row_dimensions[G1_HDR].height = 20
    merge_full(G1_HDR,
        "GRID 1 — THERAPIST HEADCOUNT BY MONTH  "
        "(Sourced from Dynamic Staffing Plan · T = permanent, S = seasonal, ⚠ = understaffed flag)",
        bg=NAVY, sz=10)
    col_headers(G1_COLS, bg_month="2C5F77", bg_stat="1B3A4B", extra_col=True)
    ws.cell(G1_COLS, 15).value = "DYNAMIC STAFFING PLAN RAW VALUE"

    LOC_ORDER = [
        "Intercontinental", "Hugos", "Ramla", "Hyatt", "Excelsior",
        "Riviera", "Sunny Coast", "Novotel", "Sliema (NEW)",
    ]

    for li, loc in enumerate(LOC_ORDER):
        row = G1_DATA + li
        ws.row_dimensions[row].height = 20
        data = plan.get(loc, {})
        counts = data.get("therapists", [0] * 12)
        raws   = data.get("raw",        ["—"] * 12)

        c = ws.cell(row, 1)
        c.value = loc
        c.font = Font(bold=True, color=BLACK, size=10)
        c.fill = PatternFill("solid", start_color="F4F6F7")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()

        for mi, (count, raw) in enumerate(zip(counts, raws)):
            cell = ws.cell(row, 2 + mi)
            raw_str = str(raw) if raw else "—"
            warn = is_understaffed(raw_str)
            cell.value = count
            cell.font = Font(color=RED_FONT if warn else BLACK,
                             bold=warn, size=10)
            cell.fill = PatternFill("solid", start_color="FADBD8" if warn else LIGHT_BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border()
            cell.number_format = "0"

        fy_avg_t = round(sum(counts) / 12, 1)
        cell = ws.cell(row, 14)
        cell.value = fy_avg_t
        cell.font = Font(bold=True, color=BLACK, size=10)
        cell.fill = PatternFill("solid", start_color=LIGHT_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        cell.number_format = "0.0"

        # Raw values summary
        raw_summary = "  |  ".join(
            f"{m}: {str(r)}" for m, r in zip(MONTHS, raws)
            if r and str(r).strip() not in ("", "None")
        )
        c = ws.cell(row, 15)
        c.value = raw_summary
        c.font = Font(color="666666", italic=True, size=8)
        c.fill = PatternFill("solid", start_color="FAFAFA")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = thin_border()

    # ─────────────────────────────────────────────────────────────────────────
    # GRID 2 — Treatments Per Therapist Per Day
    # ─────────────────────────────────────────────────────────────────────────
    G2_SP   = G1_DATA + len(LOC_ORDER)
    G2_HDR  = G2_SP  + 1
    G2_COLS = G2_HDR + 1
    G2_DATA = G2_COLS + 1

    ws.row_dimensions[G2_SP].height = 10
    ws.row_dimensions[G2_HDR].height = 20
    merge_full(G2_HDR,
        "GRID 2 — TREATMENTS PER THERAPIST PER DAY  "
        "(Monthly Rev × Seasonal Index ÷ (€85 × Days × Therapists)  |  "
        "Colour = feasibility: Green = ideal · Yellow = busy · Orange = at limit · Red = impossible)",
        bg=NAVY, sz=10)
    col_headers(G2_COLS, bg_month="1B7A3B", bg_stat="4A235A", extra_col=True)
    ws.cell(G2_COLS, 15).value = "FEASIBILITY NOTE"

    # Thresholds for annotation
    NOTES = {
        "impossible": "Exceeds 6 tx/day max — physically impossible with this headcount.",
        "at_limit":   "At absolute capacity. Zero margin for no-shows, consultation time, or extended treatments.",
        "very_busy":  "Very busy day. Sustainable short-term only. No flex for cancellations or extras.",
        "healthy":    "Healthy target range. Good utilisation with capacity headroom.",
        "light":      "Light utilisation. Room to grow revenue or reduce seasonal headcount.",
        "low":        "Under-utilised. Headcount exceeds what current revenue requires.",
    }

    for li, loc in enumerate(LOC_ORDER):
        row = G2_DATA + li
        ws.row_dimensions[row].height = 22
        data    = plan.get(loc, {})
        counts  = data.get("therapists", [0] * 12)
        raws    = data.get("raw",        ["—"] * 12)
        rev     = REVENUES.get(loc, 0)

        c = ws.cell(row, 1)
        c.value = loc
        c.font = Font(bold=True, color=BLACK, size=10)
        c.fill = PatternFill("solid", start_color="F4F6F7")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()

        month_loads = []
        worst_month = ("", 0.0)
        best_month  = ("", 99.0)

        for mi, (therapists, raw) in enumerate(zip(counts, raws)):
            month_rev = rev * SEASONAL_IDX[mi]
            days      = DAYS_IN_MO[mi]
            if therapists > 0:
                tx_per_therapist = month_rev / (AVG_TX_REV * days * therapists)
            else:
                tx_per_therapist = 99.0  # flag as error

            month_loads.append(tx_per_therapist)
            warn = is_understaffed(str(raw) if raw else "")

            if tx_per_therapist > worst_month[1]:
                worst_month = (MONTHS[mi], tx_per_therapist)
            if tx_per_therapist < best_month[1]:
                best_month = (MONTHS[mi], tx_per_therapist)

            bg   = workload_color(tx_per_therapist)
            bold = tx_per_therapist > 5.0 or warn
            fclr = RED_FONT if tx_per_therapist > 6.0 else BLACK

            cell = ws.cell(row, 2 + mi)
            cell.value = round(tx_per_therapist, 1)
            cell.font = Font(color=fclr, bold=bold, size=10)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border()
            cell.number_format = "0.0"

        # FY Average
        fy_avg = sum(month_loads) / 12
        bg_fy  = workload_color(fy_avg)
        cell   = ws.cell(row, 14)
        cell.value = round(fy_avg, 1)
        cell.font = Font(color=BLACK, bold=True, size=10)
        cell.fill = PatternFill("solid", start_color=bg_fy)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        cell.number_format = "0.0"

        # Feasibility note
        if fy_avg > 5.0:
            note_txt = f"IMPOSSIBLE — {worst_month[0]} peaks at {worst_month[1]:.1f} tx/therapist/day. Must add headcount."
            note_color = RED_FONT
        elif fy_avg > 4.0:
            note_txt = (f"At capacity — {worst_month[0]} at {worst_month[1]:.1f} tx/day. "
                        f"Add seasonal cover for peak months.")
            note_color = "C0392B"
        elif fy_avg > 3.0:
            note_txt = (f"Busy but sustainable — peaks {worst_month[0]} {worst_month[1]:.1f} tx/day. "
                        f"Light months ({best_month[0]}: {best_month[1]:.1f} tx/day) have headroom.")
            note_color = "7D6608"
        elif fy_avg > 2.0:
            note_txt = (f"Healthy utilisation — {worst_month[0]} {worst_month[1]:.1f} tx/day. "
                        f"Comfortable pacing year-round.")
            note_color = GREEN_FONT
        else:
            note_txt = (f"Under-utilised — avg {fy_avg:.1f} tx/therapist/day. "
                        f"Revenue and/or headcount may be misaligned.")
            note_color = "2E86C1"

        c = ws.cell(row, 15)
        c.value = note_txt
        c.font = Font(color=note_color, bold=(fy_avg > 4.0), size=9)
        c.fill = PatternFill("solid", start_color="FAFAFA")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = thin_border()

    # ── Legend ─────────────────────────────────────────────────────────────────
    LG_SP  = G2_DATA + len(LOC_ORDER)
    LG_HDR = LG_SP  + 1
    LG_R   = LG_HDR + 1

    ws.row_dimensions[LG_SP].height = 10
    ws.row_dimensions[LG_HDR].height = 20
    merge_full(LG_HDR, "COLOUR LEGEND — Treatments Per Therapist Per Day", bg=GOLD, sz=10)

    legend = [
        (C_GREY,   "< 2.0 tx / therapist / day",
         "Severely under-utilised. Headcount significantly exceeds demand for this location/month."),
        (C_LGREEN, "2.0 – 3.0 tx / therapist / day",
         "Light utilisation. Comfortable for staff; room to grow revenue or reduce seasonal headcount."),
        (C_GREEN,  "3.0 – 4.0 tx / therapist / day",
         "Ideal / target range. Good pacing — therapists are productive without being overloaded. "
         "Matches 80% utilisation benchmark at 75-min average treatment."),
        (C_YELLOW, "4.0 – 5.0 tx / therapist / day",
         "Busy. Manageable but leaves little margin for longer treatments, "
         "consultations, or same-day cancellations."),
        (C_ORANGE, "5.0 – 6.0 tx / therapist / day",
         "At absolute capacity. Every slot must be filled perfectly. "
         "Not sustainable beyond short peak periods without quality impact."),
        (C_RED,    "> 6.0 tx / therapist / day",
         "IMPOSSIBLE. Exceeds the physical maximum (6 × 90-min slots in a 9-hr shift). "
         "Revenue target cannot be met with current headcount — increase therapist allocation."),
    ]
    for i, (bg, band, desc) in enumerate(legend):
        row = LG_R + i
        ws.row_dimensions[row].height = 20
        c = ws.cell(row, 1)
        c.value = band
        c.font = Font(bold=True, color=RED_FONT if bg == C_RED else BLACK, size=9)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=total_cols)
        d = ws.cell(row, 2)
        d.value = desc
        d.font = Font(color=RED_FONT if bg == C_RED else BLACK,
                      bold=(bg == C_RED), size=9)
        d.fill = PatternFill("solid", start_color=bg)
        d.alignment = Alignment(horizontal="left", vertical="center")
        d.border = thin_border()

    ws.freeze_panes = "B11"


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    wb = load_workbook(FILE)
    build_workload_check(wb)
    wb.save(FILE)
    print(f"Saved: {FILE}")
    print(f"Sheets: {wb.sheetnames}")

if __name__ == "__main__":
    main()
