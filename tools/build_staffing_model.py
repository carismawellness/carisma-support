"""
build_staffing_model.py
Generates: HR/Carisma Spa Staffing Model.xlsx
Three sheets:
  1. Staffing Model   — Revenue tiers, 35% vs 40% HC team composition
  2. Weekly Roster    — 7-day roster for Tier 1 (Inter, 40%) and Tier 2 (Ramla, 35%)
  3. Regional Rotation — Cluster A/B floating model and cost savings
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os

# ── Palette ──────────────────────────────────────────────────────────────────
NAVY      = "1B3A4B"   # Carisma Deep Navy — headers
GOLD      = "B8943E"   # Carisma Muted Gold — accent / section titles
LIGHT_BLUE= "D6E4F0"   # Zebra row A
WHITE     = "FFFFFF"
YELLOW_BG = "FFFF00"   # Key assumptions
BLUE_FONT = "0000FF"   # Hardcoded inputs
BLACK     = "000000"
GREEN_FONT= "008000"   # Cross-sheet links
RED_FONT  = "FF0000"   # Negative / over-budget

# ── Style helpers ─────────────────────────────────────────────────────────────
def hdr(cell, text, bg=NAVY, fg=WHITE, bold=True, wrap=False, sz=10):
    cell.value = text
    cell.font = Font(bold=bold, color=fg, size=sz)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=wrap)

def sub_hdr(cell, text, bg=GOLD, fg=WHITE, bold=True, sz=10):
    cell.value = text
    cell.font = Font(bold=bold, color=fg, size=sz)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def inp(cell, value, bold=False, align="right", fmt=None):
    """Blue = hardcoded input."""
    cell.value = value
    cell.font = Font(color=BLUE_FONT, bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt

def calc(cell, formula, align="right", fmt=None):
    """Black = formula / calculated."""
    cell.value = formula
    cell.font = Font(color=BLACK, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt

def lbl(cell, text, bold=False, align="left"):
    cell.value = text
    cell.font = Font(color=BLACK, bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")

def note(cell, text, color=BLACK, italic=True, sz=9):
    cell.value = text
    cell.font = Font(color=color, italic=italic, size=sz)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_border(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = thin_border()

def zebra(ws, row, min_col, max_col, even):
    bg = LIGHT_BLUE if even else WHITE
    for c in range(min_col, max_col + 1):
        ws.cell(row, c).fill = PatternFill("solid", start_color=bg)

# ── Seasonal revenue index ────────────────────────────────────────────────────
# Malta spa seasonal pattern: peak in Jul/Aug (tourism), secondary in Dec (holidays)
# Indices sum to 12.0 so monthly_rev = annual_rev × idx / 12
MONTHS       = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
SEASONAL_IDX = [0.80, 0.85, 0.90, 0.95, 1.00, 1.10, 1.20, 1.25, 1.10, 1.00, 0.90, 0.95]
# sum = 12.00 ✓

# ── SHIFT COLOUR MAP (Sheet 2) ────────────────────────────────────────────────
SHIFT_COLORS = {
    "A":   ("2196A6", WHITE),   # Teal-blue  — Early shift
    "B":   ("5C6BC0", WHITE),   # Indigo     — Late shift
    "OFF": ("E8E8E8", "888888"),# Light grey — Day off
    "—":   ("D0D0D0", "888888"),# Grey       — RM away at another spa
    "M":   ("B8943E", WHITE),   # Gold       — RM management visit
}

def shift_cell(cell, code):
    bg, fg = SHIFT_COLORS.get(code, ("FFFFFF", "000000"))
    cell.value = code
    cell.font = Font(bold=True, color=fg, size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

# =============================================================================
# SHEET 1 — STAFFING MODEL
# =============================================================================
def build_staffing_model(wb):
    ws = wb.active
    ws.title = "Staffing Model"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    # ── Column widths ──
    widths = {"A": 22, "B": 14, "C": 14, "D": 8, "E": 8, "F": 8,
              "G": 14, "H": 14, "I": 8, "J": 8, "K": 8, "L": 14, "M": 26}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 14
    ws.row_dimensions[4].height = 22

    # ── Row 1 — Main title ──
    ws.merge_cells("A1:M1")
    hdr(ws["A1"], "CARISMA SPA & WELLNESS — OPTIMAL STAFFING MODEL", sz=14)

    ws.merge_cells("A2:M2")
    sub_hdr(ws["A2"], "HC Budget = Revenue × 35% (Lean) or 40% (Revenue-Max)  |  "
            "Operating Hours: 09:00–20:00, 7 Days/Week  |  Shift A: 09:00–18:00  |  Shift B: 11:00–20:00", sz=9)

    # ── Row 3 — blank spacer ──
    ws.row_dimensions[3].height = 8

    # ── Row 4 — Section headers ──
    ws.merge_cells("A4:A4"); hdr(ws["A4"], "LOCATION", bg=NAVY)
    ws.merge_cells("B4:B4"); hdr(ws["B4"], "MONTHLY REV (€)")
    ws.merge_cells("C4:C4"); hdr(ws["C4"], "TIER")

    # 35% block
    ws.merge_cells("D4:G4"); hdr(ws["D4"], "◀  35% HC BUDGET — FULLY LOADED COST  ▶", bg="2C5F77")
    # 40% block
    ws.merge_cells("H4:K4"); hdr(ws["H4"], "◀  40% HC BUDGET — FULLY LOADED COST  ▶", bg="1B7A3B")

    ws.merge_cells("L4:L4"); hdr(ws["L4"], "DELTA COST", bg="4A235A")
    ws.merge_cells("M4:M4"); hdr(ws["M4"], "NOTES", bg=NAVY)

    # ── Row 5 — Sub-headers ──
    ws.row_dimensions[5].height = 30
    for col, txt in [("A", ""), ("B", ""), ("C", ""),
                     ("D", "Budget (€)"), ("E", "Staff Mix"), ("F", "FTE"), ("G", "Cost (€)"),
                     ("H", "Budget (€)"), ("I", "Staff Mix"), ("J", "FTE"), ("K", "Cost (€)"),
                     ("L", "40% – 35% Cost"), ("M", "")]:
        c = ws[f"{col}5"]
        bg = "2C5F77" if col in ("D","E","F","G") else "1B7A3B" if col in ("H","I","J","K") else "4A235A" if col == "L" else NAVY
        hdr(c, txt, bg=bg, sz=9, wrap=True)

    # ── Cost model constants ─────────────────────────────────────────────────────
    # All costs FULLY LOADED: base salary + 11.8% employer SSC + all variable commissions
    # Regional Manager = what was previously called "supervisor". RM rotates across cluster spas.
    T_BASE          = 1500      # Therapist base salary / month (€)
    R_BASE          = 1400      # Receptionist base salary / month (€)
    RM_BASE         = 2500      # Regional Manager base salary / month (€)
    EMPLOYER_NI     = 0.118     # Malta Class 1 employer SSC — employer pays 11.8% of gross
    T_LOADED        = round(T_BASE  * (1 + EMPLOYER_NI))   # €1,677 / mo per therapist
    R_LOADED        = round(R_BASE  * (1 + EMPLOYER_NI))   # €1,565 / mo per receptionist
    RM_LOADED       = round(RM_BASE * (1 + EMPLOYER_NI))   # €2,795 / mo per RM (fixed, SSC incl.)
    TREATMENT_COMM  = 0.045     # Treatment commission — 4.5% blend (4.5% standard / 6% high performers)
    RETAIL_PCT      = 0.055     # Retail revenue as % of total spa revenue (5–6% midpoint = 5.5%)
    RETAIL_COMM     = 0.15      # Therapist retail commission — 15% on all retail sales
    RM_REV_COMM     = 0.005     # RM earns 0.5% of total cluster revenue (charged per location by FTE fraction)

    def _fc(rev, t, r, rm, pt=0):
        """Fully-loaded monthly employer cost for one location.
        Includes: loaded base salaries (NI incl.) + treatment commission + retail commission
        + RM revenue commission (1% of location revenue if RM is assigned).
        pt = additional monthly cost for part-time/weekend staff."""
        fixed    = t * T_LOADED + r * R_LOADED + round(rm * RM_LOADED)
        variable = round(
            rev * (1 - RETAIL_PCT) * TREATMENT_COMM +        # treatment commissions paid out
            rev * RETAIL_PCT * RETAIL_COMM +                  # retail commissions paid out
            (rev * RM_REV_COMM if rm > 0 else 0)             # RM earns 1% of this location's revenue
        )
        return fixed + variable + pt

    # ── Data rows ──
    # Format: (location, revenue, tier, t35, r35, rm35, cost35, t40, r40, rm40, cost40, note_text)
    # RM clusters (each RM covers exactly 3 spas at ⅓ FTE each):
    #   RM1: Inter + Hugos + Hyatt
    #   RM2: Ramla + Riviera + Sunny Coast
    #   RM3: Excelsior + Novotel + Sliema
    # Both 35% and 40% models use ⅓ RM per location (same allocation, same RM)
    # Revenues validated against Carisma Analytics "Revenue Improved" sheet, Apr 2026
    rows = [
        ("Inter",              49475, "Tier 1", 9, 2, 0.33, _fc(49475,9,2,0.33), 9, 2, 0.33, _fc(49475,9,2,0.33),
         "2025 actual: €593,700/yr÷12 (ex-VAT). RM1 cluster: Inter+Hugos+Hyatt. ⅓ FTE each. "
         "Fully loaded = SSC-loaded base + treatment 4.5% + retail 15% + RM 0.5% cluster rev commission."),
        ("Hugos",              47546, "Tier 1", 9, 2, 0.33, _fc(47546,9,2,0.33), 9, 2, 0.33, _fc(47546,9,2,0.33),
         "2025 actual: €570,552/yr÷12 (ex-VAT). RM1 cluster: Inter+Hugos+Hyatt. ⅓ FTE. "
         "Sister site to Inter — identical schedule structure."),
        ("Ramla",              29886, "Tier 2", 5, 2, 0.33, _fc(29886,5,2,0.33), 5, 2, 0.33, _fc(29886,5,2,0.33),
         "2025 actual: €358,632/yr÷12 (ex-VAT). RM2 cluster: Ramla+Riviera+Sunny Coast. ⅓ FTE each."),
        ("Hyatt",              25358, "Tier 2", 4, 1, 0.33, _fc(25358,4,1,0.33), 4, 1, 0.33, _fc(25358,4,1,0.33),
         "2025 actual: €304,296/yr÷12 (ex-VAT). RM1 cluster: Inter+Hugos+Hyatt. ⅓ FTE each."),
        ("Excelsior",          24263, "Tier 2", 4, 1, 0.33, _fc(24263,4,1,0.33), 4, 1, 0.33, _fc(24263,4,1,0.33),
         "2025 actual: €145,578÷6mo (ex-VAT, NEW). RM3 cluster: Excelsior+Novotel+Sliema. ⅓ FTE each."),
        ("Riviera",            19936, "Tier 3", 4, 1, 0.33, _fc(19936,4,1,0.33), 4, 1, 0.33, _fc(19936,4,1,0.33),
         "2025 actual: €239,232/yr÷12 (ex-VAT). RM2 cluster: Ramla+Riviera+Sunny Coast. ⅓ FTE each."),
        ("Sunny Coast (Odycy)", 18204, "Tier 3", 3, 1, 0.33, _fc(18204,3,1,0.33), 3, 1, 0.33, _fc(18204,3,1,0.33),
         "2025 actual: €218,448/yr÷12 (ex-VAT). RM2 cluster: Ramla+Riviera+Sunny Coast. ⅓ FTE each."),
        ("Novotel",            14000, "Tier 4", 2, 0, 0.33, _fc(14000,2,0,0.33,1092), 2, 0, 0.33, _fc(14000,2,0,0.33,1092),
         "2026 projected: €168,000/yr (ex-VAT). RM3 cluster: Excelsior+Novotel+Sliema. "
         "2 FT + PT Fri–Sun (€1,092/mo incl.). ⅓ FTE RM."),
        ("Sliema (NEW)",       41667, "Tier 3", 5, 1, 0.33, _fc(41667,5,1,0.33),  5, 1, 0.33, _fc(41667,5,1,0.33),
         "INA 2025: €426,018/yr ex-VAT. Projected 2026: €500,000/yr. "
         "RM3 cluster: Excelsior+Novotel+Sliema. ⅓ FTE. "
         "5 therapists confirmed on Fresha (utilisation check: ~75% FY avg, peak Aug ~93%). "
         "35% budget: €14,583 vs modelled cost ~€13,200 — within budget."),
    ]

    TIER_COLORS = {"Tier 1": "D4EFDF", "Tier 2": "D6EAF8", "Tier 3": "FEF9E7", "Tier 4": "FDEDEC"}

    for i, (loc, rev, tier, t35, r35, s35, cost35, t40, r40, s40, cost40, note_txt) in enumerate(rows):
        row = 6 + i
        ws.row_dimensions[row].height = 36
        tier_bg = TIER_COLORS.get(tier, WHITE)

        # Location
        c = ws.cell(row, 1)
        lbl(c, loc, bold=True)
        c.fill = PatternFill("solid", start_color=tier_bg)

        # Revenue
        c = ws.cell(row, 2)
        inp(c, rev, fmt='#,##0')
        c.fill = PatternFill("solid", start_color=tier_bg)

        # Tier
        c = ws.cell(row, 3)
        lbl(c, tier, bold=True, align="center")
        c.fill = PatternFill("solid", start_color=tier_bg)

        # 35% budget
        c = ws.cell(row, 4)
        calc(c, f"=B{row}*0.35", fmt='#,##0')
        c.fill = PatternFill("solid", start_color="EAF4FB")

        # 35% staff mix
        mix35 = f"{int(t35)}T"
        if r35 == 1.5: mix35 += " + 1.5R"
        elif r35 > 0:  mix35 += f" + {int(r35)}R"
        if s35 == 0.5: mix35 += " + 0.5RM"
        elif s35 == 0.33: mix35 += " + ⅓RM"
        elif s35 > 0:  mix35 += f" + {int(s35)}RM"
        c = ws.cell(row, 5)
        inp(c, mix35, align="center")
        c.fill = PatternFill("solid", start_color="EAF4FB")

        # 35% FTE
        fte35 = t35 + r35 + s35
        c = ws.cell(row, 6)
        inp(c, fte35, fmt='0.0')
        c.fill = PatternFill("solid", start_color="EAF4FB")

        # 35% cost
        c = ws.cell(row, 7)
        inp(c, cost35, fmt='#,##0')
        over35 = cost35 > rev * 0.35
        if over35:
            c.font = Font(color=RED_FONT, bold=True, size=10)
        c.fill = PatternFill("solid", start_color="EAF4FB")

        # 40% budget
        c = ws.cell(row, 8)
        calc(c, f"=B{row}*0.40", fmt='#,##0')
        c.fill = PatternFill("solid", start_color="E9F7EF")

        # 40% staff mix
        mix40 = f"{int(t40)}T"
        if r40 == 1.5: mix40 += " + 1.5R"
        elif r40 > 0:  mix40 += f" + {int(r40)}R"
        if s40 == 0.5: mix40 += " + 0.5RM"
        elif s40 == 0.33: mix40 += " + ⅓RM"
        elif s40 > 0:  mix40 += f" + {int(s40)}RM"
        c = ws.cell(row, 9)
        inp(c, mix40, align="center")
        c.fill = PatternFill("solid", start_color="E9F7EF")

        # 40% FTE
        fte40 = t40 + r40 + s40
        c = ws.cell(row, 10)
        inp(c, fte40, fmt='0.0')
        c.fill = PatternFill("solid", start_color="E9F7EF")

        # 40% cost
        c = ws.cell(row, 11)
        inp(c, cost40, fmt='#,##0')
        over40 = cost40 > rev * 0.40
        if over40:
            c.font = Font(color=RED_FONT, bold=True, size=10)
        c.fill = PatternFill("solid", start_color="E9F7EF")

        # Delta
        c = ws.cell(row, 12)
        calc(c, f"=K{row}-G{row}", fmt='+#,##0;-#,##0;"-"')
        if cost40 > cost35:
            c.font = Font(color=RED_FONT, size=10)
        else:
            c.font = Font(color="006400", size=10)
        c.fill = PatternFill("solid", start_color="F5EEF8")

        # Notes
        c = ws.cell(row, 13)
        note(c, note_txt, color="444444", italic=True, sz=9)
        c.fill = PatternFill("solid", start_color=tier_bg)

    # ── Totals row ──
    total_row = 6 + len(rows)
    ws.row_dimensions[total_row].height = 24
    # Write label to col A only — do NOT merge so we can still write formulas to cols B/C
    c = ws.cell(total_row, 1)
    hdr(c, "PORTFOLIO TOTAL", bg=NAVY, sz=10)
    # Manually style B and C as navy (no merge)
    for col_idx in [2, 3]:
        c2 = ws.cell(total_row, col_idx)
        c2.fill = PatternFill("solid", start_color=NAVY)

    # Rev total in col B
    c = ws.cell(total_row, 2)
    calc(c, f"=SUM(B6:B{total_row-1})", fmt='#,##0')
    c.fill = PatternFill("solid", start_color=NAVY)
    c.font = Font(color=WHITE, bold=True, size=10)

    for col, formula_col in [(4, "D"), (6, "F"), (7, "G"), (8, "H"), (10, "J"), (11, "K"), (12, "L")]:
        c = ws.cell(total_row, col)
        calc(c, f"=SUM({formula_col}6:{formula_col}{total_row-1})", fmt='#,##0')
        c.fill = PatternFill("solid", start_color=NAVY)
        c.font = Font(color=WHITE, bold=True, size=10)
        c.alignment = Alignment(horizontal="right", vertical="center")

    apply_border(ws, 4, total_row, 1, 13)

    # ── Salary assumptions box ──
    box_row = total_row + 2
    ws.merge_cells(f"A{box_row}:M{box_row}")
    sub_hdr(ws.cell(box_row, 1), "SALARY & COMMISSION ASSUMPTIONS  (Blue = hardcoded input — change to model scenarios)", bg=GOLD)

    labels = [
        "Therapist base salary / month (€)",
        "Therapist fully loaded / month (€)  [base + 11.8% employer SSC]",
        "Receptionist base salary / month (€)",
        "Receptionist fully loaded / month (€)  [base + 11.8% employer SSC]",
        "Regional Manager base / month (€)  [target fixed cost budget]",
        "Regional Manager fully loaded / month (€)  [RM base + 11.8% SSC]",
        "Employer SSC rate  (Malta Class 1 social security)",
        "Treatment commission rate  (blend %)",
        "Retail revenue as % of total spa revenue",
        "Retail commission rate  (% of retail sales)",
        "Regional Manager revenue commission  (% of cluster total revenue)",
        "Floating mobility allowance (€/month)",
        "Avg treatment revenue (€)",
        "Avg treatment duration (mins)",
        "Operating days / week",
        "Operating hours / day",
        "Peak utilisation target (%)",
    ]
    values = [
        1500, round(1500*1.118),
        1400, round(1400*1.118),
        2500, round(2500*1.118),
        "11.8%",
        "4.5%",
        "5.5%",
        "15%",
        "0.5%",
        75, 85, 75, 7, 11, "80%",
    ]
    notes = [
        "Base salary before employer taxes. FSS (income tax) is employee-side — not included here.",
        "Therapist base × 1.118. This is the employer's fully-loaded monthly cost per therapist.",
        "Base salary before employer taxes.",
        "Receptionist base × 1.118. Employer's fully-loaded monthly cost per receptionist.",
        "Base salary €2,500/mo. Each RM covers a cluster of 3 spas (⅓ FTE each).",
        "RM base × 1.118. Each RM covers 3 spas at ⅓ FTE each. ⅓ FTE = €932/mo fixed at that site.",
        "Employer pays 11.8% of gross salary as Class 1 SSC. FSS (income tax) is withheld from employee pay.",
        "Blend: 4.5% standard / 6% high-performer. Applied to treatment revenue (ex-retail) only.",
        "Industry benchmark: 5–6% of total spa revenue. Midpoint 5.5% used for modelling.",
        "Therapists earn 15% commission on all retail products sold at their location.",
        "RM earns 0.5% of total cluster revenue (sum of all 3 spas they cover). Charged ⅓ to each location.",
        "Cluster floater incentive to cover multi-site mobility.",
        "Mid-point €70–€100 range per treatment.",
        "Mid-point 60–90 min range.",
        "Mon–Sun, 7 days.",
        "09:00–20:00 operating window.",
        "Industry benchmark — mature spa at steady state.",
    ]

    for j, (label, val, n) in enumerate(zip(labels, values, notes)):
        r = box_row + 1 + j
        ws.row_dimensions[r].height = 18
        lbl(ws.cell(r, 1), label, bold=True)
        inp(ws.cell(r, 2), val)
        note(ws.cell(r, 3), n)
        ws.merge_cells(f"C{r}:M{r}")
        apply_border(ws, r, r, 1, 3)

    # ── Fully-loaded cost breakdown table ──
    bk_row = box_row + 1 + len(labels) + 1
    ws.merge_cells(f"A{bk_row}:M{bk_row}")
    sub_hdr(ws.cell(bk_row, 1), "FULLY LOADED MONTHLY EMPLOYER COST BREAKDOWN PER ROLE", bg="2C5F77")

    # Header row
    bk_hdr_row = bk_row + 1
    ws.row_dimensions[bk_hdr_row].height = 20
    for col_idx, txt in [(1, "Role"), (2, "Base Salary (€)"), (3, "Employer SSC 11.8% (€)"),
                         (4, "Fully Loaded Fixed (€)"), (5, "Treatment Comm"), (6, "Retail Comm"),
                         (7, "RM Rev Comm"), (8, "Notes")]:
        c = ws.cell(bk_hdr_row, col_idx)
        hdr(c, txt, bg="2C5F77", sz=9)
    ws.merge_cells(f"H{bk_hdr_row}:M{bk_hdr_row}")

    sample_rev = 30000   # illustrative monthly revenue for commission calculation display
    bk_data = [
        ("Therapist (1 FTE)",    1500, round(1500*0.118), round(1500*1.118),
         "4.5% × treatment rev",
         "15% × (5.5% of rev)",
         "N/A",
         f"At €{sample_rev:,}/mo rev: treatment ~€{round(sample_rev*0.945*0.045):,} + retail ~€{round(sample_rev*0.055*0.15):,}"),
        ("Receptionist (1 FTE)", 1400, round(1400*0.118), round(1400*1.118),
         "No commission",
         "No commission",
         "N/A",
         "Fixed cost only — no revenue-linked variable pay."),
        ("Regional Manager (1 FTE)", 3000, round(3000*0.118), round(3000*1.118),
         "None",
         "None",
         "0.5% × cluster total rev",
         f"At €{sample_rev:,}/mo rev: RM comm = €{round(sample_rev*0.005):,}. Allocated by FTE fraction."),
        ("RM @ 0.5 FTE (shared)", round(3000*1.118*0.5), round(3000*0.118*0.5), round(3000*1.118*0.5),
         "None",
         "None",
         "0.5% × location rev",
         f"0.5 FTE = €{round(3000*1.118*0.5):,} fixed + 0.5% rev commission charged to location."),
        ("RM @ ⅓ FTE (3-spa cluster)", round(3000*1.118/3), round(3000*0.118/3), round(3000*1.118/3),
         "None",
         "None",
         "0.5% × location rev",
         f"⅓ FTE = €{round(3000*1.118/3):,} fixed + 0.5% rev commission charged to location."),
    ]
    for k, (role, base, ni, loaded, t_comm, ret_comm, rm_comm, note_txt) in enumerate(bk_data):
        r = bk_hdr_row + 1 + k
        ws.row_dimensions[r].height = 18
        lbl(ws.cell(r, 1), role, bold=True)
        inp(ws.cell(r, 2), base, fmt='#,##0')
        inp(ws.cell(r, 3), ni, fmt='#,##0')
        inp(ws.cell(r, 4), loaded, fmt='#,##0')
        note(ws.cell(r, 5), t_comm)
        note(ws.cell(r, 6), ret_comm)
        note(ws.cell(r, 7), rm_comm)
        note(ws.cell(r, 8), note_txt)
        ws.merge_cells(f"H{r}:M{r}")
        apply_border(ws, r, r, 1, 8)

    # ── Legend ──
    # bk_row = box_row+1+len(labels)+1, bk_hdr_row = bk_row+1, data = len(bk_data) rows
    leg_row = bk_hdr_row + 1 + len(bk_data) + 1
    ws.merge_cells(f"A{leg_row}:M{leg_row}")
    sub_hdr(ws.cell(leg_row, 1), "COLOUR LEGEND", bg="666666")

    legends = [
        ("Blue text", "Hardcoded input — change to model scenarios", BLUE_FONT),
        ("Black text", "Formula / calculated value", BLACK),
        ("Red text", "Over-budget or deficit", RED_FONT),
        ("Green text", "Positive variance", "006400"),
    ]
    tier_legends = [
        ("Tier 1 (>€40k)", TIER_COLORS["Tier 1"]),
        ("Tier 2 (€25k–€40k)", TIER_COLORS["Tier 2"]),
        ("Tier 3 (€15k–€25k)", TIER_COLORS["Tier 3"]),
        ("Tier 4 (<€15k)", TIER_COLORS["Tier 4"]),
    ]
    for j, (label, desc, color) in enumerate(legends):
        r = leg_row + 1 + j
        ws.row_dimensions[r].height = 16
        c = ws.cell(r, 1); c.value = label
        c.font = Font(color=color, bold=True, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")
        note(ws.cell(r, 2), desc, sz=9, italic=False)
        ws.merge_cells(f"B{r}:F{r}")
    for j, (label, bg) in enumerate(tier_legends):
        r = leg_row + 1 + j
        c = ws.cell(r, 8); c.value = label
        c.font = Font(color=BLACK, size=10)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()


# =============================================================================
# SHEET 2 — WEEKLY ROSTER  (all 9 locations, Saturday-mandatory model)
# =============================================================================
def build_roster(wb):
    ws = wb.create_sheet("Weekly Roster")
    ws.sheet_view.showGridLines = False

    days = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]

    def section_title(start_row, title, subtitle):
        ws.merge_cells(f"A{start_row}:J{start_row}")
        c = ws.cell(start_row, 1)
        hdr(c, title, bg=NAVY, sz=11)
        ws.row_dimensions[start_row].height = 28
        ws.merge_cells(f"A{start_row+1}:J{start_row+1}")
        c = ws.cell(start_row + 1, 1)
        hdr(c, subtitle, bg=GOLD, sz=9, bold=False)
        ws.row_dimensions[start_row + 1].height = 16

    def roster_headers(row):
        ws.row_dimensions[row].height = 22
        hdr(ws.cell(row, 1), "STAFF", bg="374F6B", sz=10)
        hdr(ws.cell(row, 2), "ROLE", bg="374F6B", sz=10)
        for d, day in enumerate(days):
            c = ws.cell(row, 3 + d)
            bg = "7D3C98" if day in ("FRI","SAT","SUN") else "374F6B"
            hdr(c, day + (" ★" if day in ("FRI","SAT","SUN") else ""), bg=bg, sz=10)
        hdr(ws.cell(row, 10), "DAYS ON", bg="374F6B", sz=10)

    def roster_row(ws_row, name, role, schedule, even):
        ws.row_dimensions[ws_row].height = 22
        c = ws.cell(ws_row, 1)
        lbl(c, name, bold=True)
        c.fill = PatternFill("solid", start_color=LIGHT_BLUE if even else WHITE)
        c = ws.cell(ws_row, 2)
        lbl(c, role)
        c.fill = PatternFill("solid", start_color=LIGHT_BLUE if even else WHITE)
        days_on = sum(1 for s in schedule if s not in ("OFF", "—"))
        for d, code in enumerate(schedule):
            shift_cell(ws.cell(ws_row, 3 + d), code)
        c = ws.cell(ws_row, 10)
        inp(c, days_on, align="center")
        c.fill = PatternFill("solid", start_color="FFFDE7")

    def coverage_check_row(ws_row, schedule_list, min_per_day, label):
        ws.row_dimensions[ws_row].height = 20
        ws.merge_cells(f"A{ws_row}:B{ws_row}")
        c = ws.cell(ws_row, 1)
        c.value = label
        c.font = Font(bold=True, color=NAVY, size=9, italic=True)
        c.alignment = Alignment(horizontal="left", vertical="center")
        for d in range(7):
            count = sum(1 for sched in schedule_list if sched[d] != "OFF")
            c = ws.cell(ws_row, 3 + d)
            ok = count >= min_per_day
            c.value = f"✓ {count}" if ok else f"✗ {count}"
            c.font = Font(color="006400" if ok else RED_FONT, bold=True, size=9)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", start_color="E8F5E9" if ok else "FFEBEE")
            c.border = thin_border()

    def double_coverage_row(ws_row, all_staff):
        ws.row_dimensions[ws_row].height = 20
        ws.merge_cells(f"A{ws_row}:B{ws_row}")
        c = ws.cell(ws_row, 1)
        c.value = "Double shift coverage (A+B):"
        c.font = Font(bold=True, color=NAVY, size=9, italic=True)
        c.alignment = Alignment(horizontal="left", vertical="center")
        for d in range(7):
            a_count = sum(1 for (_, _r, s) in all_staff if s[d] == "A")
            b_count = sum(1 for (_, _r, s) in all_staff if s[d] == "B")
            both = a_count > 0 and b_count > 0
            c = ws.cell(ws_row, 3 + d)
            c.value = "A+B ✓" if both else ("A only" if a_count > 0 else "B only" if b_count > 0 else "—")
            c.font = Font(color="006400" if both else "FF8C00", bold=True, size=9)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", start_color="E8F5E9" if both else "FFF3E0")
            c.border = thin_border()

    def add_roster_section(r, title, subtitle, staff_list, min_t):
        section_title(r, title, subtitle)
        r += 2
        roster_headers(r); r += 1
        for i, (name, role, sched) in enumerate(staff_list):
            roster_row(r, name, role, sched, i % 2 == 0); r += 1
        ws.row_dimensions[r].height = 8; r += 1
        t_scheds = [s for (_, role, s) in staff_list if "Therapist" in role]
        r_scheds = [s for (_, role, s) in staff_list if role == "Receptionist"]
        coverage_check_row(r, t_scheds, min_t, f"Therapists on floor (min {min_t}):"); r += 1
        if r_scheds:
            coverage_check_row(r, r_scheds, 1, "Reception covered (min 1):"); r += 1
        double_coverage_row(r, staff_list); r += 2
        return r

    # ── Column widths ──
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    for d in range(7):
        ws.column_dimensions[get_column_letter(3 + d)].width = 9
    ws.column_dimensions["J"].width = 9

    # ── Main title ──
    ws.merge_cells("A1:J1")
    hdr(ws["A1"], "CARISMA SPA & WELLNESS — WEEKLY STAFF ROSTER  (ALL 9 LOCATIONS)", sz=13)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    hdr(ws["A2"],
        "Shift A = 09:00–18:00  |  Shift B = 11:00–20:00  |  OFF = Day Off  |  ★ = Peak Day (Fri/Sat/Sun)  |  "
        "SATURDAY IS MANDATORY — no Sat OFF for any staff member",
        bg=GOLD, sz=9, bold=False)
    ws.row_dimensions[2].height = 16

    # Legend for shifts
    ws.row_dimensions[3].height = 20
    for col, (code, bg, fg) in enumerate([("A", "2196A6", WHITE), ("B", "5C6BC0", WHITE), ("OFF", "E8E8E8", "888888")]):
        c = ws.cell(3, 1 + col * 3)
        c.value = f"  {code} — {'Early 09:00–18:00' if code=='A' else 'Late 11:00–20:00' if code=='B' else 'Day Off'}"
        c.font = Font(bold=True, color=fg, size=9)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=3, start_column=1 + col*3, end_row=3, end_column=2 + col*3)

    ws.row_dimensions[4].height = 10  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # Saturday-mandatory roster design rules:
    #   Valid off-day pairs (consecutive, exclude Sat): Mon/Tue | Tue/Wed | Wed/Thu
    #   All therapists work Fri+Sat+Sun. Off days fall within Mon–Thu only.
    # ══════════════════════════════════════════════════════════════════════════
    r = 5

    # ─── ROSTER 1 — INTER ───────────────────────────────────────────────────
    # 3 off-groups × 3 therapists each = 9T, all work Fri/Sat/Sun
    # Peak coverage: 6A + 3B every Fri/Sat/Sun
    inter_staff = [
        # Group X: Mon/Tue OFF — works Wed–Sun
        ("T1 — Isabelle",  "Therapist",    ["OFF","OFF","A","A","A","A","A"]),
        ("T2 — Margaux",   "Therapist",    ["OFF","OFF","A","A","A","A","B"]),
        ("T3 — Chiara",    "Therapist",    ["OFF","OFF","B","B","B","B","A"]),
        # Group Y: Tue/Wed OFF — works Mon + Thu–Sun
        ("T4 — Eleni",     "Therapist",    ["A","OFF","OFF","A","A","A","A"]),
        ("T5 — Priya",     "Therapist",    ["A","OFF","OFF","A","A","A","B"]),
        ("T6 — Sofia",     "Therapist",    ["B","OFF","OFF","B","B","B","A"]),
        # Group Z: Wed/Thu OFF — works Mon–Tue + Fri–Sun
        ("T7 — Anya",      "Therapist",    ["A","A","OFF","OFF","A","A","A"]),
        ("T8 — Leila",     "Therapist",    ["A","A","OFF","OFF","A","A","B"]),
        ("T9 — Nadia",     "Therapist",    ["B","B","OFF","OFF","B","B","A"]),
        ("R1 — Camille",   "Receptionist", ["OFF","OFF","A","A","A","A","A"]),  # Mon/Tue off
        ("R2 — Yasmin",    "Receptionist", ["A","A","OFF","OFF","A","A","A"]),  # Wed/Thu off
        ("RM1 — Jordan",   "Reg. Manager", ["OFF","—","—","A","A","A","OFF"]),  # Off Mon+Sun; Inter/Hugos Thu–Sat
    ]
    r = add_roster_section(r,
        "ROSTER 1 — INTER  (Tier 1 | Revenue-Max 40% HC | 9T + 2R + 1S)",
        "2025 Revenue: €49,475/mo (ex-VAT)  |  HC Budget (40%): €19,790  |  Base Cost: €18,000  |  "
        "9 therapists on Fri/Sat/Sun | 6A + 3B peak coverage",
        inter_staff, 2)

    # ─── ROSTER 2 — HUGOS ───────────────────────────────────────────────────
    # Identical structure to Inter (sister site, 5-min walk)
    hugos_staff = [
        # Group X: Mon/Tue OFF
        ("T1 — Lucia",     "Therapist",    ["OFF","OFF","A","A","A","A","A"]),
        ("T2 — Amara",     "Therapist",    ["OFF","OFF","A","A","A","A","B"]),
        ("T3 — Valentina", "Therapist",    ["OFF","OFF","B","B","B","B","A"]),
        # Group Y: Tue/Wed OFF
        ("T4 — Hana",      "Therapist",    ["A","OFF","OFF","A","A","A","A"]),
        ("T5 — Renata",    "Therapist",    ["A","OFF","OFF","A","A","A","B"]),
        ("T6 — Zara",      "Therapist",    ["B","OFF","OFF","B","B","B","A"]),
        # Group Z: Wed/Thu OFF
        ("T7 — Fatima",    "Therapist",    ["A","A","OFF","OFF","A","A","A"]),
        ("T8 — Ingrid",    "Therapist",    ["A","A","OFF","OFF","A","A","B"]),
        ("T9 — Celine",    "Therapist",    ["B","B","OFF","OFF","B","B","A"]),
        ("R1 — Miriam",    "Receptionist", ["OFF","OFF","A","A","A","A","A"]),
        ("R2 — Dina",      "Receptionist", ["A","A","OFF","OFF","A","A","A"]),
        ("RM1 — Jordan",   "Reg. Manager", ["OFF","—","—","A","A","A","OFF"]),  # Off Mon+Sun; Hugos/Inter Thu–Sat
    ]
    r = add_roster_section(r,
        "ROSTER 2 — HUGOS  (Tier 1 | 9T + 2R + 0.5S shared w/ Hyatt)",
        "2025 Revenue: €47,546/mo (ex-VAT)  |  HC Budget (40%): €19,018  |  Base Cost: €17,150 (0.5S shared w/ Hyatt)  |  "
        "Sister site to Inter (5-min walk) — identical schedule structure",
        hugos_staff, 2)

    # ─── ROSTER 3 — RAMLA ───────────────────────────────────────────────────
    # Off groups: 2×Mon/Tue + 1×Tue/Wed + 2×Wed/Thu
    # Day coverage: Mon=3, Tue=2, Wed=2, Thu=3, Fri/Sat/Sun=5
    ramla_staff = [
        ("T1 — Maria",         "Therapist",    ["OFF","OFF","A","A","A","A","A"]),  # Mon/Tue off
        ("T2 — Sofia",         "Therapist",    ["OFF","OFF","B","B","B","A","B"]),  # Mon/Tue off; Sat=A for A+B
        ("T3 — Elena",         "Therapist",    ["A","OFF","OFF","A","A","A","A"]),  # Tue/Wed off
        ("T4 — Isabelle",      "Therapist",    ["B","A","OFF","OFF","B","A","B"]),  # Wed/Thu off
        ("T5 — Natalia",       "Therapist",    ["A","B","OFF","OFF","A","B","A"]),  # Wed/Thu off
        ("R1 — Claire",        "Receptionist", ["OFF","OFF","A","A","A","A","A"]),  # Mon/Tue off
        ("R2 — Yasmine",       "Receptionist", ["A","A","OFF","OFF","A","A","A"]),  # Wed/Thu off
        ("RM2 — Lucas",        "Reg. Manager", ["—","—","—","—","OFF","A","OFF"]),  # Off Fri+Sun; Ramla Sat mgmt
    ]
    r = add_roster_section(r,
        "ROSTER 3 — RAMLA  (Tier 2 | Lean 35% HC | 5T + 2R + 0.5S)",
        "2025 Revenue: €29,886/mo (ex-VAT)  |  HC Budget (35%): €10,460  |  Base Cost: €11,150  |  "
        "All 5T work Fri/Sat/Sun | A+B on all days",
        ramla_staff, 1)

    # ─── ROSTER 4 — HYATT ───────────────────────────────────────────────────
    # Off groups: 2×Mon/Tue + 2×Wed/Thu (clean symmetry — exactly 2T every day)
    hyatt_staff = [
        ("T1 — Amara",          "Therapist",    ["OFF","OFF","A","A","A","A","A"]),  # Mon/Tue off
        ("T2 — Lucia",          "Therapist",    ["OFF","OFF","B","B","B","B","B"]),  # Mon/Tue off
        ("T3 — Priya",          "Therapist",    ["A","A","OFF","OFF","A","A","A"]),  # Wed/Thu off
        ("T4 — Lena",           "Therapist",    ["B","B","OFF","OFF","B","B","B"]),  # Wed/Thu off
        ("R1 — Nadia",          "Receptionist", ["A","OFF","OFF","A","A","A","A"]),  # Tue/Wed off; Mon+Thu–Sun
        ("RM1 — Jordan",        "Reg. Manager", ["OFF","A","A","—","—","A","OFF"]),  # Off Mon+Sun; Hyatt Tue+Wed reception; Sat mgmt
    ]
    r = add_roster_section(r,
        "ROSTER 4 — HYATT  (Tier 2 | 4T + 1R + 0.5S shared w/ Hugos)",
        "2025 Revenue: €25,358/mo (ex-VAT)  |  HC Budget (35%): €8,875  |  Base Cost: €8,250 (0.5S shared w/ Hugos)  |  "
        "Tier 2 upgrade | 2T every day | 4T peak days | Note: Tue/Wed no front desk",
        hyatt_staff, 1)

    # ─── ROSTER 5 — EXCELSIOR ───────────────────────────────────────────────
    # Same structure as Hyatt
    excelsior_staff = [
        ("T1 — Valentina",      "Therapist",    ["OFF","OFF","A","A","A","A","A"]),  # Mon/Tue off
        ("T2 — Miriam",         "Therapist",    ["OFF","OFF","B","B","B","B","B"]),  # Mon/Tue off
        ("T3 — Fatima",         "Therapist",    ["A","A","OFF","OFF","A","A","A"]),  # Wed/Thu off
        ("T4 — Chiara",         "Therapist",    ["B","B","OFF","OFF","B","B","B"]),  # Wed/Thu off
        ("R1 — Beatriz",        "Receptionist", ["OFF","OFF","A","A","A","A","A"]),  # Mon/Tue off; Wed–Sun
        ("RM3 — Morgan",        "Reg. Manager", ["A","A","—","—","OFF","A","OFF"]),  # Mon+Tue reception; Sat mgmt
    ]
    r = add_roster_section(r,
        "ROSTER 5 — EXCELSIOR  (Tier 2 | 4T + 1R + 0.33S shared w/ Novotel+Sliema)",
        "2025 Revenue: €24,263/mo (ex-VAT)  |  HC Budget (35%): €8,492  |  Base Cost: €7,961 (0.33S shared w/ Novotel+Sliema)  |  "
        "NEW location (6-month data) | Tier 2 upgrade | RM3 covers Mon+Tue (R1 off) | Note: Wed/Thu no front desk",
        excelsior_staff, 1)

    # ─── ROSTER 6 — RIVIERA ─────────────────────────────────────────────────
    # Off groups: 2×Mon/Tue + 1×Tue/Wed + 1×Wed/Thu
    # Note: Tuesday has only 1 therapist — structural constraint at 4T with Sat-mandatory
    riviera_staff = [
        ("T1 — Maya",    "Therapist",    ["OFF","OFF","A","B","A","A","B"]),  # Mon/Tue off
        ("T2 — Sofia",   "Therapist",    ["OFF","OFF","B","A","B","B","A"]),  # Mon/Tue off
        ("T3 — Elena",   "Therapist",    ["A","OFF","OFF","B","B","A","A"]),  # Tue/Wed off
        ("T4 — Lucia",   "Therapist",    ["B","A","OFF","OFF","A","B","A"]),  # Wed/Thu off
        ("R1 — Daniela", "Receptionist", ["OFF","OFF","A","A","A","A","B"]),  # Mon/Tue off; Wed–Sun
        ("RM2 — Lucas",  "Reg. Manager", ["A","A","—","—","OFF","A","OFF"]),  # Mon+Tue reception; Off Fri+Sun; Sat mgmt
    ]
    r = add_roster_section(r,
        "ROSTER 6 — RIVIERA  (Tier 3 | 4T + 1R + 0.33RM shared w/ Ramla+Sunny Coast)",
        "2025 Revenue: €19,936/mo (ex-VAT)  |  HC Budget (35%): €6,978  |  Base Cost: €8,250 (0.33RM shared)  |  "
        "All 4T on Fri/Sat/Sun | RM2 covers Mon+Tue reception | Note: Tue single-therapist day (structural — 4T + Sat-mandatory)",
        riviera_staff, 1)

    # ─── ROSTER 7 — SUNNY COAST (ODYCY) ─────────────────────────────────────
    # Off groups: 1×Mon/Tue + 1×Tue/Wed + 1×Wed/Thu
    # R1: Wed/Thu off → covers Mon/Tue + Fri/Sat/Sun (all peak days covered)
    # T2 Thu=A fixes Thu A+B coverage
    sunny_coast_staff = [
        ("T1 — Natasha",   "Therapist",    ["OFF","OFF","A","B","A","A","B"]),  # Mon/Tue off
        ("T2 — Priya",     "Therapist",    ["A","OFF","OFF","A","B","A","A"]),  # Tue/Wed off; Thu=A → A+B on Thu
        ("T3 — Irene",     "Therapist",    ["B","A","OFF","OFF","A","B","A"]),  # Wed/Thu off
        ("R1 — Valentina", "Receptionist", ["A","A","OFF","OFF","A","A","B"]),  # Wed/Thu off; covers all peak days
        ("RM2 — Lucas",    "Reg. Manager", ["—","—","A","A","OFF","A","OFF"]),  # Wed+Thu reception; Off Fri+Sun; Sat mgmt
    ]
    r = add_roster_section(r,
        "ROSTER 7 — SUNNY COAST (ODYCY)  (Tier 3 | 3T + 1R + 0.33RM shared w/ Ramla+Riviera)",
        "2025 Revenue: €18,204/mo (ex-VAT)  |  HC Budget (35%): €6,371  |  Base Cost: €6,750 (0.33RM shared)  |  "
        "Tier 3 upgrade (was Tier 4) | RM2 covers Wed+Thu reception | Note: Tue/Wed single-therapist days (structural — 3T)",
        sunny_coast_staff, 1)

    # ─── ROSTER 8 — NOVOTEL ─────────────────────────────────────────────────
    # 2 FT therapists + 1 PT (Fri/Sat/Sun only)
    # PT enables A+B on all peak days; removes single-point-of-failure risk
    novotel_staff = [
        ("T1 — Chiara",        "Therapist",      ["OFF","OFF","A","B","A","A","B"]),  # Mon/Tue off; Wed–Sun
        ("T2 — Bianca",        "Therapist",      ["B","A","OFF","OFF","B","B","A"]),  # Wed/Thu off; Mon/Tue+Fri–Sun
        ("PT1 — Rosaria (PT)", "Therapist (PT)", ["OFF","OFF","OFF","OFF","A","A","A"]),  # Fri/Sat/Sun only
        ("RM3 — Morgan",       "Reg. Manager",   ["—","—","—","—","OFF","A","OFF"]),  # Off Fri+Sun; Novotel Sat mgmt
    ]
    r = add_roster_section(r,
        "ROSTER 8 — NOVOTEL  (Tier 4 | 2 FT + 1 PT weekend + 0.33RM shared w/ Excelsior+Sliema)",
        "2026 Revenue: €14,000/mo (ex-VAT)  |  HC Budget (35%): €4,900  |  2 FT + 0.33RM: €3,561/mo  |  "
        "PT Fri+Sat+Sun: €1,092/mo  |  Total staffing: €4,653 (budget surplus: +€247)",
        novotel_staff, 1)

    # ─── ROSTER 9 — SLIEMA (NEW / INA ACQUISITION) ──────────────────────────
    sliema_staff = [
        ("T1 — Amara",   "Therapist",    ["A","OFF","OFF","B","A","A","B"]),  # Tue/Wed off; Mon+Thu–Sun
        ("T2 — Rania",   "Therapist",    ["OFF","A","A","OFF","B","B","A"]),  # Tue/Wed + Fri/Sat/Sun; opposite shift to Amara on peak days
        ("R1 — Chloe",   "Receptionist", ["A","A","OFF","OFF","A","A","B"]),  # Wed/Thu off; Mon/Tue+Fri–Sun
        ("RM3 — Morgan", "Reg. Manager", ["—","—","A","A","OFF","A","OFF"]),  # Wed+Thu reception; Off Fri+Sun; Sat mgmt
    ]
    r = add_roster_section(r,
        "ROSTER 9 — SLIEMA (NEW | INA Acquisition)  (Tier 3 | 2T + 1R + 0.33RM shared w/ Excelsior+Novotel)",
        "INA 2025: €426,018/yr ex-VAT → €41,667/mo avg  |  Projected 2026: €500,000/yr  |  "
        "HC Budget (35%): €14,583  |  Base Cost: €4,961 (0.33RM shared)  |  Surplus: +€9,622/mo  |  "
        "RM3 covers Wed+Thu reception | A+B double coverage on Fri/Sat/Sun: T1=A+T2=B (Fri/Sat), T1=B+T2=A (Sun)",
        sliema_staff, 1)

    # ══════════════════════════════════════════════════════════════════════════
    # SHIFT HOURS SUMMARY
    # ══════════════════════════════════════════════════════════════════════════
    section_title(r, "SHIFT HOURS SUMMARY", "Reference guide for all rosters")
    r += 2
    ws.row_dimensions[r].height = 20
    for col, hdr_txt in enumerate(["Shift", "Start", "Finish", "Break (unpaid)", "Productive hrs", "Weekly hrs"]):
        hdr(ws.cell(r, 1 + col), hdr_txt, bg=NAVY, sz=9)
    r += 1
    for code, start, end, brk, prod, weekly in [
        ("A — Early",    "09:00", "18:00",              "1 hr", "8 hrs", "40 hrs (5d)"),
        ("B — Late",     "11:00", "20:00",              "1 hr", "8 hrs", "40 hrs (5d)"),
        ("PT — Weekend", "09:00", "18:00 or 11:00–20:00","1 hr", "8 hrs", "24 hrs (3d)"),
    ]:
        ws.row_dimensions[r].height = 18
        for col, val in enumerate([code, start, end, brk, prod, weekly]):
            c = ws.cell(r, 1 + col)
            inp(c, val, align="center")
            c.border = thin_border()
        r += 1

    r += 1  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # PART-TIMER STRATEGY — WEEKEND-ONLY MODEL
    # ══════════════════════════════════════════════════════════════════════════
    section_title(r,
        "PART-TIMER (PT) STRATEGY — WEEKEND-ONLY STAFFING MODEL",
        "PT works Fri + Sat + Sun only (~24hrs/week, ~0.6 FTE)  |  Cost: ~€1,092/mo vs €1,500+/mo FT  |  "
        "No days-off scheduling constraint  |  Best for: Novotel, Sunny Coast, Sliema ramp-up")
    r += 2

    ws.row_dimensions[r].height = 22
    for col, txt in enumerate(["LOCATION", "PRIORITY", "PT SCHEDULE", "FT COST/MO",
                                "PT COST/MO", "SAVING/MO", "USE CASE / TRIGGER"]):
        hdr(ws.cell(r, 1 + col), txt, bg="4A235A", sz=9)
    r += 1

    pt_rows = [
        ("Novotel",       "HIGH ★",  "Fri+Sat+Sun (A)", "€1,500", "€1,092", "€408",
         "Enables A+B peak coverage. 2FT+1PT=3 therapists Fri–Sun. "
         "Self-funding: 13+ extra treatments/mo covers PT cost."),
        ("Sunny Coast",   "MEDIUM",  "Fri+Sat+Sun",     "€1,500", "€1,092", "€408",
         "Absence insurance + surge buffer for 3T location. "
         "Add when peak utilisation >80% consistently."),
        ("Sliema (ramp)",  "DEPLOYED","Tue+Wed+Fri–Sun",  "€1,500", "€1,092", "€408",
         "PT1 (Rania) covers T1 off days + B-shift on peak days. "
         "Enables A+B on all 3 peak days. Convert PT to FT at €25k+/mo revenue."),
        ("Riviera",       "LOW",     "Sat+Sun overflow", "€1,500", "€546",   "€954",
         "Sat/Sun only (16hrs, ~€546/mo) as overflow. "
         "Trigger: >85% peak utilisation on 2+ consecutive weekends."),
        ("Any location",  "TRIGGER", "Fri+Sat+Sun",      "n/a",    "€1,092", "n/a",
         "Deploy when peak bookings consistently exceed 80% of "
         "total therapist treatment capacity on Fri/Sat/Sun."),
    ]
    pt_colors = ["FDEDEC", "FEF9E7", "FEF9E7", "D6EAF8", "D5F5E3"]
    for i, row_data in enumerate(pt_rows):
        ws.row_dimensions[r].height = 36
        bg = pt_colors[i]
        for col, val in enumerate(row_data):
            c = ws.cell(r, 1 + col)
            c.value = val
            c.font = Font(color=BLACK, size=9, bold=(col == 0))
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="left" if col in (0, 6) else "center",
                                    vertical="center", wrap_text=True)
            c.border = thin_border()
        r += 1

    apply_border(ws, 1, r, 1, 10)


# =============================================================================
# SHEET 3 — REGIONAL ROTATION
# =============================================================================
def build_regional(wb):
    ws = wb.create_sheet("Regional Rotation")
    ws.sheet_view.showGridLines = False

    widths = {"A": 22, "B": 14, "C": 14, "D": 14, "E": 14,
              "F": 14, "G": 14, "H": 14, "I": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Title ──
    ws.merge_cells("A1:I1")
    hdr(ws["A1"], "CARISMA SPA — REGIONAL CLUSTER ROTATION MODEL", sz=13)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    hdr(ws["A2"],
        "Cluster A: Inter + Hugos + Hyatt (RM1)  |  "
        "Cluster B: Riviera + Sunny Coast + Ramla (RM2)  |  "
        "Cluster C: Excelsior + Novotel + Sliema (RM3)",
        bg=GOLD, sz=9, bold=False)
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 10  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION A: CLUSTER DEFINITIONS
    # ══════════════════════════════════════════════════════════════════════════
    r = 4
    ws.merge_cells(f"A{r}:I{r}")
    sub_hdr(ws.cell(r, 1), "SECTION 1 — CLUSTER DEFINITIONS & GEOGRAPHIC GROUPING", bg=NAVY)
    ws.row_dimensions[r].height = 22; r += 1

    cluster_hdr_cols = ["LOCATION", "CLUSTER", "REVENUE/MO", "DEMAND TIER", "RM", "DISTANCE TO HUB"]
    ws.row_dimensions[r].height = 22
    for col, txt in enumerate(cluster_hdr_cols):
        c = ws.cell(r, 1 + col)
        hdr(c, txt, bg="2C5F77", sz=9)
    r += 1

    cluster_data = [
        # Cluster A: Inter + Hugos + Hyatt (RM1)
        ("Inter",                "Cluster A",   49475, "High",        "RM1", "Hub — RM1 base"),
        ("Hugos",                "Cluster A",   47546, "High",        "RM1", "~2 min walk from Inter"),
        ("Hyatt",                "Cluster A",   25358, "Medium-High", "RM1", "~4 min walk"),
        # Cluster B: Riviera + Sunny Coast (Odycy) + Ramla (RM2)
        ("Riviera",              "Cluster B",   19936, "Medium",      "RM2", "Hub — RM2 base"),
        ("Sunny Coast (Odycy)",  "Cluster B",   18204, "Medium",      "RM2", "~5 min from Riviera"),
        ("Ramla",                "Cluster B",   29886, "Medium-High", "RM2", "RM2 management Sat"),
        # Cluster C: Excelsior + Novotel + Sliema (RM3)
        ("Excelsior",            "Cluster C",   24263, "Medium-High", "RM3", "Hub — RM3 base"),
        ("Novotel",              "Cluster C",   14000, "Low-Medium",  "RM3", "RM3 management Sat"),
        ("Sliema (NEW)",         "Cluster C",   41667, "Ramping",     "RM3", "~7 min walk"),
    ]
    for i, (loc, cluster, rev, demand, rm, dist) in enumerate(cluster_data):
        ws.row_dimensions[r].height = 20
        bg_map = {"Cluster A": "D4EFDF", "Cluster B": "D6EAF8", "Cluster C": "FEF9E7", "Standalone": "FDFEFE"}
        row_bg = bg_map.get(cluster, WHITE)
        for col, val in enumerate([loc, cluster, rev, demand, rm, dist]):
            c = ws.cell(r, 1 + col)
            if col == 2:
                inp(c, val, fmt='€#,##0')
            else:
                lbl(c, str(val), bold=(col == 0))
            c.fill = PatternFill("solid", start_color=row_bg)
            c.border = thin_border()
        r += 1

    r += 1  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION B: FLOATING STAFF MODEL
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells(f"A{r}:I{r}")
    sub_hdr(ws.cell(r, 1), "SECTION 2 — FLOATING STAFF MODEL: DEDICATED vs CLUSTER COMPARISON", bg=NAVY)
    ws.row_dimensions[r].height = 22; r += 1

    # Headers
    ws.row_dimensions[r].height = 22
    float_cols = ["CLUSTER", "METRIC", "DEDICATED MODEL", "CLUSTER MODEL", "SAVING/MO", "SAVING/YR", "NOTES"]
    for col, txt in enumerate(float_cols):
        hdr(ws.cell(r, 1 + col), txt, bg="1B7A3B", sz=9)
    r += 1

    float_data = [
        ("Cluster A", "Therapists",      "12 dedicated",  "8 dedicated + 3 floating", "€6,000", "€72,000",
         "Floaters cover Hyatt & Odycy demand spikes"),
        ("Cluster A", "Receptionists",   "4 dedicated",   "3 dedicated + 1 floating", "€1,400", "€16,800",
         "Floating recep. covers Odycy (no dedicated desk)"),
        ("Cluster A", "Supervisors",     "4 dedicated",   "2 dedicated (Inter+Hugos)", "€3,400", "€40,800",
         "RM1 absorbs oversight for Hyatt & Odycy"),
        ("Cluster A", "TOTAL BASE COST", "€30,400",       "€25,800",                  "€4,600", "€55,200",
         "15.1% cost reduction for Cluster A"),
        ("Cluster B", "Therapists",      "5 dedicated",   "3 dedicated + 2 floating", "€3,000", "€36,000",
         "Novotel fully served by floating pool"),
        ("Cluster B", "Receptionists",   "3 dedicated",   "2 dedicated + 1 floating", "€1,400", "€16,800",
         "Novotel reception via floater"),
        ("Cluster B", "Supervisors",     "3 dedicated",   "1 dedicated (Excelsior)",  "€3,400", "€40,800",
         "RM2 covers Novotel + Sliema oversight"),
        ("Cluster B", "TOTAL BASE COST", "€16,800",       "€13,625",                  "€3,175", "€38,100",
         "18.9% cost reduction for Cluster B"),
        ("TOTAL",     "ALL CLUSTERS",    "€71,000/mo",    "€55,325/mo",               "€15,675","€188,100",
         "22.1% total HC reduction — 43 FTE → 34 FTE"),
    ]
    for i, row_data in enumerate(float_data):
        ws.row_dimensions[r].height = 20
        even = i % 2 == 0
        is_total = row_data[0] in ("Cluster A", "Cluster B") and "TOTAL" in row_data[1]
        is_grand = row_data[0] == "TOTAL"
        for col, val in enumerate(row_data):
            c = ws.cell(r, 1 + col)
            c.value = val
            if is_grand:
                c.font = Font(bold=True, color=WHITE, size=10)
                c.fill = PatternFill("solid", start_color=NAVY)
            elif is_total:
                c.font = Font(bold=True, color=NAVY, size=10)
                c.fill = PatternFill("solid", start_color="D5F5E3")
            else:
                c.font = Font(color=BLACK, size=9)
                c.fill = PatternFill("solid", start_color=LIGHT_BLUE if even else WHITE)
            c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
            c.border = thin_border()
        r += 1

    r += 1  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION C: CLUSTER A — 7-DAY FLOATING DEPLOYMENT SCHEDULE
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells(f"A{r}:I{r}")
    sub_hdr(ws.cell(r, 1), "SECTION 3 — CLUSTER A: 7-DAY FLOATING THERAPIST DEPLOYMENT PLAN", bg=NAVY)
    ws.row_dimensions[r].height = 22; r += 1

    days = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
    ws.row_dimensions[r].height = 22
    hdr(ws.cell(r, 1), "FLOATER", bg="374F6B", sz=9)
    hdr(ws.cell(r, 2), "DEMAND LEVEL", bg="374F6B", sz=9)
    for d, day in enumerate(days):
        bg = "7D3C98" if day in ("FRI","SAT","SUN") else "374F6B"
        hdr(ws.cell(r, 3 + d), day, bg=bg, sz=9)
    r += 1

    # Demand row
    ws.row_dimensions[r].height = 18
    lbl(ws.cell(r, 1), "Cluster A Demand", bold=True)
    lbl(ws.cell(r, 2), "All locations")
    demand_levels = [("Medium","E8F4F8","1B6CA8"), ("Low","F8F9FA","6C757D"),
                     ("Medium","E8F4F8","1B6CA8"), ("High","E8F8E8","1B7A3B"),
                     ("HIGH ★","F8E8FF","7D3C98"), ("HIGH ★","F8E8FF","7D3C98"),
                     ("HIGH ★","F8E8FF","7D3C98")]
    for d, (level, bg, fg) in enumerate(demand_levels):
        c = ws.cell(r, 3 + d)
        c.value = level
        c.font = Font(bold=True, color=fg, size=9)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    r += 1

    floater_plan = [
        ("FT-1 (Cluster A)", "Therapist",  ["Stand-by: Inter", "OFF", "→ Hugos", "→ Inter", "→ Inter", "→ Inter", "→ Inter"]),
        ("FT-2 (Cluster A)", "Therapist",  ["→ Hugos", "Stand-by: Inter", "Stand-by: Inter", "→ Hugos", "→ Hugos", "→ Hugos", "OFF"]),
        ("FT-3 (Cluster A)", "Therapist",  ["OFF", "Stand-by pool", "→ Hyatt", "→ Hyatt", "→ Hyatt", "→ Odycy", "→ Hyatt"]),
        ("FR-1 (Cluster A)", "Float Recep.",["→ Odycy", "→ Odycy", "Hyatt/Odycy", "→ Hyatt", "→ Inter", "→ Inter", "→ Odycy"]),
    ]
    for i, (name, role, sched) in enumerate(floater_plan):
        ws.row_dimensions[r].height = 22
        row_even = i % 2 == 0
        lbl(ws.cell(r, 1), name, bold=True)
        ws.cell(r, 1).fill = PatternFill("solid", start_color=LIGHT_BLUE if row_even else WHITE)
        lbl(ws.cell(r, 2), role)
        ws.cell(r, 2).fill = PatternFill("solid", start_color=LIGHT_BLUE if row_even else WHITE)
        for d, action in enumerate(sched):
            c = ws.cell(r, 3 + d)
            c.value = action
            if action.startswith("→"):
                bg, fg = "D4EFDF", "1B7A3B"
            elif action.startswith("Stand"):
                bg, fg = "FEF9E7", "B7950B"
            else:
                bg, fg = "E8E8E8", "888888"
            c.font = Font(color=fg, size=9)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border()
        r += 1

    r += 1  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION D: SLIMA NEW HOTEL RAMP-UP PLAN
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells(f"A{r}:I{r}")
    sub_hdr(ws.cell(r, 1), "SECTION 4 — SLIMA NEW HOTEL: MONTH-BY-MONTH RAMP-UP STAFFING", bg=NAVY)
    ws.row_dimensions[r].height = 22; r += 1

    ws.row_dimensions[r].height = 22
    ramp_cols = ["PHASE", "MONTHS", "PROJ. REVENUE", "DEDICATED STAFF",
                 "FLOATING COVERAGE", "FIXED HC COST", "HC %", "TRIGGER TO NEXT PHASE"]
    for col, txt in enumerate(ramp_cols):
        hdr(ws.cell(r, 1 + col), txt, bg="2C5F77", sz=9)
    r += 1

    ramp_data = [
        ("Phase 1 — Launch", "Months 1–2", "€5k–€7k",
         "1T + 1R", "1–2 floaters from Cluster B",
         "€2,900", "~40%", "6+ treatments/day for 2 consecutive weeks"),
        ("Phase 2 — Growth", "Month 3", "€8k–€11k",
         "1T + 1R", "1–2 floaters (consistent pair)",
         "€2,900", "~30%", "Sustained €10k+/mo for 4 weeks"),
        ("Phase 3 — Stabilise", "Month 4+", "€12k–€15k",
         "2T + 1R", "On-demand from Cluster B",
         "€4,400", "~30%", "€15k sustained 2 months → full cluster member"),
        ("Phase 4 — Full ops", "Month 6+", "€15k+",
         "2T + 1R + shared S", "As needed",
         "€5,967", "~35–40%", "Standard Tier 3 staffing model applies"),
    ]
    phase_colors = ["EAF4FB", "D5F5E3", "FEF9E7", "F5EEF8"]
    for i, row_data in enumerate(ramp_data):
        ws.row_dimensions[r].height = 28
        bg = phase_colors[i]
        for col, val in enumerate(row_data):
            c = ws.cell(r, 1 + col)
            c.value = val
            c.font = Font(color=BLACK, size=9, bold=(col == 0))
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="left" if col in (0,1,3,4,7) else "center",
                                    vertical="center", wrap_text=True)
            c.border = thin_border()
        r += 1

    r += 1  # spacer

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION E: ROTATION RULES
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells(f"A{r}:I{r}")
    sub_hdr(ws.cell(r, 1), "SECTION 5 — FLOATING STAFF DEPLOYMENT RULES", bg=NAVY)
    ws.row_dimensions[r].height = 22; r += 1

    rules = [
        ("Rule 1", "Deployment Trigger",
         "Deploy a floating therapist when confirmed bookings in the next 4-hour window "
         "exceed 80% of dedicated therapist capacity at that location."),
        ("Rule 2", "Minimum Threshold",
         "Cluster A: 1 confirmed booking + 80% utilisation → deploy within 10 minutes. "
         "Cluster B: 2 confirmed bookings + 80% utilisation → deploy within 20 minutes."),
        ("Rule 3", "Surge Protocol",
         "Floating pool deployed FIRST. A dedicated therapist from a quiet sister location "
         "may cross-deploy only if the floating pool is fully committed. "
         "Cross-deployment bonus: €10/session."),
        ("Rule 4", "Shift Limits",
         "Max 6 treatment-hours per floater per shift. Min 30-min rest between locations. "
         "Max 2 different locations per floater per day."),
        ("Rule 5", "Priority Hierarchy",
         "1st: Highest-revenue location. 2nd: Earliest booking timestamp. 3rd: RM discretion."),
        ("Rule 6", "RM Schedule",
         "RM1 issues next-day floating assignments at 18:00 daily based on booking report. "
         "Same-day deployment is the fallback — pre-assignment is the standard."),
    ]
    ws.row_dimensions[r].height = 22
    hdr(ws.cell(r, 1), "RULE", bg="374F6B", sz=9)
    hdr(ws.cell(r, 2), "NAME", bg="374F6B", sz=9)
    ws.merge_cells(f"C{r}:I{r}")
    hdr(ws.cell(r, 3), "DESCRIPTION", bg="374F6B", sz=9)
    r += 1

    for i, (rule_id, name, desc) in enumerate(rules):
        ws.row_dimensions[r].height = 36
        even = i % 2 == 0
        row_bg = LIGHT_BLUE if even else WHITE
        c = ws.cell(r, 1)
        lbl(c, rule_id, bold=True, align="center")
        c.fill = PatternFill("solid", start_color=row_bg); c.border = thin_border()
        c = ws.cell(r, 2)
        lbl(c, name, bold=True)
        c.fill = PatternFill("solid", start_color=row_bg); c.border = thin_border()
        ws.merge_cells(f"C{r}:I{r}")
        c = ws.cell(r, 3)
        note(c, desc, italic=False, sz=9)
        c.fill = PatternFill("solid", start_color=row_bg); c.border = thin_border()
        r += 1

    apply_border(ws, 1, r, 1, 9)


# =============================================================================
# SHEETS 4–12 — PER-LOCATION MONTHLY DRILLDOWN
# =============================================================================
def add_drilldown_sheets(wb):
    """
    Add one sheet per spa (9 total). Each sheet contains:
      • ONE table — 40% HC Budget Model only (12-month revenue breakdown)
      • Weekly roster for the 40% model showing RM schedule per location
    Columns: Month | Est.Revenue | 40% Budget | Therapist Alloc | #Therapists |
             Receptionist Alloc | #Receptionists | RM Alloc | PT Alloc | TOTAL
    TOTAL == 40% Budget by design (no variance column).
    """
    import math

    # ── Cost model constants ──────────────────────────────────────────────────
    _T_LOADED   = round(1500 * 1.118)       # 1677
    _R_LOADED   = round(1400 * 1.118)       # 1565
    _RM_THIRD   = round(2500 * 1.118 / 3)  # 932
    _TREAT      = 0.045
    _RETAIL_PCT = 0.055
    _RETAIL_COM = 0.15
    _RM_REV     = 0.005

    # ── RM schedule per location (days: MON TUE WED THU FRI SAT SUN) ─────────
    # "A" = covering reception, "M" = management visit,
    # "—" = at another spa in cluster, "OFF" = RM day off
    RM_SCHED = {
        "Inter Monthly":        ["OFF", "—",  "—",  "M",   "M",   "M",   "OFF"],
        "Hugos Monthly":        ["OFF", "—",  "—",  "M",   "M",   "M",   "OFF"],
        "Hyatt Monthly":        ["OFF", "A",  "A",  "—",   "—",   "M",   "OFF"],
        "Ramla Monthly":        ["—",   "—",  "—",  "—",   "OFF", "M",   "OFF"],
        "Riviera Monthly":      ["A",   "A",  "—",  "—",   "OFF", "M",   "OFF"],
        "Sunny Coast Monthly":  ["—",   "—",  "A",  "A",   "OFF", "M",   "OFF"],
        "Excelsior Monthly":    ["A",   "A",  "—",  "—",   "OFF", "M",   "OFF"],
        "Novotel Monthly":      ["—",   "—",  "—",  "—",   "OFF", "M",   "OFF"],
        "Sliema Monthly":       ["—",   "—",  "A",  "A",   "OFF", "M",   "OFF"],
    }

    # ── Per-location data ─────────────────────────────────────────────────────
    # Format: (sheet_name, annual_rev, r_count, pt_monthly, staff_list)
    # staff_list: [(name, role, [MON,TUE,WED,THU,FRI,SAT,SUN]), ...]
    locations = [

        # ── INTER ─────────────────────────────────────────────────────────────
        ("Inter Monthly", 593700, 2, 0,
         [
             ("T1 — Isabelle",  "Therapist",    ["OFF","OFF","A","A","A","A","A"]),
             ("T2 — Margaux",   "Therapist",    ["OFF","OFF","A","A","A","A","B"]),
             ("T3 — Chiara",    "Therapist",    ["OFF","OFF","B","B","B","B","A"]),
             ("T4 — Eleni",     "Therapist",    ["A","OFF","OFF","A","A","A","A"]),
             ("T5 — Priya",     "Therapist",    ["A","OFF","OFF","A","A","A","B"]),
             ("T6 — Sofia",     "Therapist",    ["B","OFF","OFF","B","B","B","A"]),
             ("T7 — Anya",      "Therapist",    ["A","A","OFF","OFF","A","A","A"]),
             ("T8 — Leila",     "Therapist",    ["A","A","OFF","OFF","A","A","B"]),
             ("T9 — Nadia",     "Therapist",    ["B","B","OFF","OFF","B","B","A"]),
             ("R1 — Camille",   "Receptionist", ["OFF","OFF","A","A","A","A","A"]),
             ("R2 — Yasmin",    "Receptionist", ["A","A","OFF","OFF","A","A","A"]),
         ]),

        # ── HUGOS ─────────────────────────────────────────────────────────────
        ("Hugos Monthly", 570552, 2, 0,
         [
             ("T1 — Lucia",     "Therapist",    ["OFF","OFF","A","A","A","A","A"]),
             ("T2 — Amara",     "Therapist",    ["OFF","OFF","A","A","A","A","B"]),
             ("T3 — Valentina", "Therapist",    ["OFF","OFF","B","B","B","B","A"]),
             ("T4 — Hana",      "Therapist",    ["A","OFF","OFF","A","A","A","A"]),
             ("T5 — Renata",    "Therapist",    ["A","OFF","OFF","A","A","A","B"]),
             ("T6 — Zara",      "Therapist",    ["B","OFF","OFF","B","B","B","A"]),
             ("T7 — Fatima",    "Therapist",    ["A","A","OFF","OFF","A","A","A"]),
             ("T8 — Ingrid",    "Therapist",    ["A","A","OFF","OFF","A","A","B"]),
             ("T9 — Celine",    "Therapist",    ["B","B","OFF","OFF","B","B","A"]),
             ("R1 — Miriam",    "Receptionist", ["OFF","OFF","A","A","A","A","A"]),
             ("R2 — Dina",      "Receptionist", ["A","A","OFF","OFF","A","A","A"]),
         ]),

        # ── RAMLA ─────────────────────────────────────────────────────────────
        ("Ramla Monthly", 358632, 2, 0,
         [
             ("T1 — Maria",    "Therapist",    ["OFF","OFF","A","A","A","A","A"]),
             ("T2 — Sofia",    "Therapist",    ["OFF","OFF","B","B","B","A","B"]),
             ("T3 — Elena",    "Therapist",    ["A","OFF","OFF","A","A","A","A"]),
             ("T4 — Isabelle", "Therapist",    ["B","A","OFF","OFF","B","A","B"]),
             ("T5 — Natalia",  "Therapist",    ["A","B","OFF","OFF","A","B","A"]),
             ("R1 — Claire",   "Receptionist", ["OFF","OFF","A","A","A","A","A"]),
             ("R2 — Yasmine",  "Receptionist", ["A","A","OFF","OFF","A","A","A"]),
         ]),

        # ── HYATT ─────────────────────────────────────────────────────────────
        ("Hyatt Monthly", 304296, 1, 0,
         [
             ("T1 — Amara",  "Therapist",    ["OFF","OFF","A","A","A","A","A"]),
             ("T2 — Lucia",  "Therapist",    ["OFF","OFF","B","B","B","B","B"]),
             ("T3 — Priya",  "Therapist",    ["A","A","OFF","OFF","A","A","A"]),
             ("T4 — Lena",   "Therapist",    ["B","B","OFF","OFF","B","B","B"]),
             ("R1 — Nadia",  "Receptionist", ["A","OFF","OFF","A","A","A","A"]),
         ]),

        # ── EXCELSIOR ─────────────────────────────────────────────────────────
        ("Excelsior Monthly", 291156, 1, 0,
         [
             ("T1 — Valentina", "Therapist",    ["OFF","OFF","A","A","A","A","A"]),
             ("T2 — Miriam",    "Therapist",    ["OFF","OFF","B","B","B","B","B"]),
             ("T3 — Fatima",    "Therapist",    ["A","A","OFF","OFF","A","A","A"]),
             ("T4 — Chiara",    "Therapist",    ["B","B","OFF","OFF","B","B","B"]),
             ("R1 — Beatriz",   "Receptionist", ["OFF","OFF","A","A","A","A","A"]),
         ]),

        # ── RIVIERA ───────────────────────────────────────────────────────────
        ("Riviera Monthly", 239232, 1, 0,
         [
             ("T1 — Maya",    "Therapist",    ["OFF","OFF","A","B","A","A","B"]),
             ("T2 — Sofia",   "Therapist",    ["OFF","OFF","B","A","B","B","A"]),
             ("T3 — Elena",   "Therapist",    ["A","OFF","OFF","B","B","A","A"]),
             ("T4 — Lucia",   "Therapist",    ["B","A","OFF","OFF","A","B","A"]),
             ("R1 — Daniela", "Receptionist", ["OFF","OFF","A","A","A","A","B"]),
         ]),

        # ── SUNNY COAST ───────────────────────────────────────────────────────
        ("Sunny Coast Monthly", 218448, 1, 0,
         [
             ("T1 — Natasha",   "Therapist",    ["OFF","OFF","A","B","A","A","B"]),
             ("T2 — Priya",     "Therapist",    ["A","OFF","OFF","A","B","A","A"]),
             ("T3 — Irene",     "Therapist",    ["B","A","OFF","OFF","A","B","A"]),
             ("R1 — Valentina", "Receptionist", ["A","A","OFF","OFF","A","A","B"]),
         ]),

        # ── NOVOTEL ───────────────────────────────────────────────────────────
        ("Novotel Monthly", 168000, 0, 1092,
         [
             ("T1 — Chiara",        "Therapist",      ["OFF","OFF","A","B","A","A","B"]),
             ("T2 — Bianca",        "Therapist",      ["B","A","OFF","OFF","B","B","A"]),
             ("PT1 — Rosaria (PT)", "Therapist (PT)", ["OFF","OFF","OFF","OFF","A","A","A"]),
         ]),

        # ── SLIEMA ────────────────────────────────────────────────────────────
        ("Sliema Monthly", 500000, 1, 0,
         [
             ("T1 — Amara", "Therapist",    ["A","OFF","OFF","B","A","A","B"]),
             ("T2 — Rania", "Therapist",    ["OFF","A","A","OFF","B","B","A"]),
             ("R1 — Chloe", "Receptionist", ["A","A","OFF","OFF","A","A","B"]),
         ]),
    ]

    # ── Display names lookup ──────────────────────────────────────────────────
    DISPLAY_NAMES = {
        "Inter Monthly":       "Intercontinental",
        "Hugos Monthly":       "Hugos",
        "Ramla Monthly":       "Ramla Bay",
        "Hyatt Monthly":       "Hyatt Regency",
        "Excelsior Monthly":   "Excelsior",
        "Riviera Monthly":     "Riviera",
        "Sunny Coast Monthly": "Sunny Coast (Odycy)",
        "Novotel Monthly":     "Novotel",
        "Sliema Monthly":      "Sliema (INA)",
    }

    BG_SECTION  = "1B7A3B"   # dark green section headers
    BG_COL_HDR  = "2C5F77"   # column header blue
    BG_DATA_A   = "E9F7EF"   # alternating row A
    BG_DATA_B   = "FFFFFF"   # alternating row B (white)
    BG_TOTAL    = "1B7A3B"   # annual total row

    # ── Generic sheet builder ─────────────────────────────────────────────────
    def _write_drilldown(sheet_name, annual_rev, r_count, pt_monthly, staff_list):
        display_name = DISPLAY_NAMES.get(sheet_name, sheet_name)
        rm_sched     = RM_SCHED.get(sheet_name, ["—"]*7)

        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

        # Column widths: A-J (10 data cols) + K (extra)
        col_widths = {"A": 9, "B": 18, "C": 18, "D": 20, "E": 13,
                      "F": 20, "G": 14, "H": 20, "I": 18, "J": 14, "K": 10}
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

        # ── Title ─────────────────────────────────────────────────────────────
        ws.merge_cells("A1:J1")
        hdr(ws["A1"],
            f"{display_name.upper()} — Monthly HC Budget Drilldown  |  40% Model Only",
            sz=12)
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:J2")
        hdr(ws["A2"],
            f"Annual revenue (ex-VAT): \u20ac{annual_rev:,.0f}  |  Monthly avg: \u20ac{annual_rev/12:,.0f}  |  "
            f"Receptionists: {r_count}  |  PT monthly: \u20ac{pt_monthly:,}",
            bg=GOLD, sz=9, bold=False)
        ws.row_dimensions[2].height = 16
        ws.row_dimensions[3].height = 8

        r_idx = 4

        # ── Section header ────────────────────────────────────────────────────
        ws.merge_cells(f"A{r_idx}:J{r_idx}")
        sub_hdr(ws.cell(r_idx, 1),
                "40% HC BUDGET MODEL  —  RM covers receptionist off-days  |  "
                "Therapist allocation = budget remainder after fixed costs",
                bg=BG_SECTION)
        ws.row_dimensions[r_idx].height = 22
        r_idx += 1

        # ── Column headers ────────────────────────────────────────────────────
        ws.row_dimensions[r_idx].height = 30
        col_hdrs = [
            "Month",
            "Est. Revenue (\u20ac)",
            "40% HC Budget (\u20ac)",
            "Therapist Alloc (\u20ac)",
            "# Therapists",
            "Receptionist Alloc (\u20ac)",
            "# Receptionists",
            "Reg. Manager Alloc (\u20ac)",
            "PT Therapist Alloc (\u20ac)",
            "TOTAL (\u20ac)",
        ]
        for ci, txt in enumerate(col_hdrs):
            c = ws.cell(r_idx, 1 + ci)
            hdr(c, txt, bg=BG_COL_HDR, sz=9, wrap=True)
        r_idx += 1

        # ── Pre-compute fixed monthly allocations ─────────────────────────────
        # Receptionist allocation is fixed (does not vary with revenue)
        r_alloc_fixed = r_count * _R_LOADED   # e.g. 2 x 1565 = 3130

        # ── Data rows ─────────────────────────────────────────────────────────
        ann_rev = ann_budget = ann_t_alloc = ann_t_count = 0
        ann_r_alloc = ann_rm_alloc = ann_pt_alloc = ann_total = 0

        for m_i, (month, idx) in enumerate(zip(MONTHS, SEASONAL_IDX)):
            rev    = round(annual_rev * idx / 12)
            budget = round(rev * 0.40)

            # Commissions (deducted from therapist bucket)
            treat_comm  = round(rev * (1 - _RETAIL_PCT) * _TREAT)
            retail_comm = round(rev * _RETAIL_PCT * _RETAIL_COM)

            # RM allocation: fixed third share + 0.5% of this location's revenue
            rm_alloc = _RM_THIRD + round(rev * _RM_REV)

            # PT allocation: fixed monthly (0 for most)
            pt_alloc = pt_monthly

            # Therapist allocation = everything left after receptionists, RM, PT
            t_alloc = budget - r_alloc_fixed - rm_alloc - pt_alloc

            # Number of therapists = floor((t_alloc - commissions) / loaded rate)
            t_net   = t_alloc - treat_comm - retail_comm
            n_t     = math.floor(t_net / _T_LOADED) if t_net > 0 else 0

            # TOTAL = t_alloc + r_alloc + rm_alloc + pt_alloc (== budget by design)
            total = t_alloc + r_alloc_fixed + rm_alloc + pt_alloc

            row_bg = BG_DATA_A if m_i % 2 == 0 else BG_DATA_B
            ws.row_dimensions[r_idx].height = 18

            data = [month, rev, budget, t_alloc, n_t,
                    r_alloc_fixed, r_count, rm_alloc, pt_alloc, total]
            for ci, val in enumerate(data):
                c = ws.cell(r_idx, 1 + ci)
                c.fill = PatternFill("solid", start_color=row_bg)
                c.border = thin_border()
                if ci == 0:
                    lbl(c, val, bold=True)
                elif ci == 3:
                    # Therapist allocation — red if impossible to staff
                    c.value = val
                    c.number_format = '#,##0'
                    c.font = Font(
                        color=RED_FONT if val < 0 else BLUE_FONT,
                        bold=(val < 0), size=10)
                    c.alignment = Alignment(horizontal="right", vertical="center")
                elif ci in (4, 6):
                    # Counts — center aligned
                    inp(c, val, align="center")
                else:
                    inp(c, val, fmt='#,##0')

            ann_rev      += rev
            ann_budget   += budget
            ann_t_alloc  += t_alloc
            ann_t_count  += n_t
            ann_r_alloc  += r_alloc_fixed
            ann_rm_alloc += rm_alloc
            ann_pt_alloc += pt_alloc
            ann_total    += total
            r_idx += 1

        # ── Annual totals row ─────────────────────────────────────────────────
        ws.row_dimensions[r_idx].height = 22
        totals = [
            "ANNUAL TOTAL", ann_rev, ann_budget, ann_t_alloc, "",
            ann_r_alloc, r_count, ann_rm_alloc, ann_pt_alloc, ann_total
        ]
        for ci, val in enumerate(totals):
            c = ws.cell(r_idx, 1 + ci)
            c.fill = PatternFill("solid", start_color=BG_TOTAL)
            c.border = thin_border()
            c.font = Font(color=WHITE, bold=True, size=10)
            c.value = val
            if ci == 0:
                c.alignment = Alignment(horizontal="left", vertical="center")
            elif ci in (4, 6):
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal="right", vertical="center")
        r_idx += 1

        ws.row_dimensions[r_idx].height = 8; r_idx += 1   # gap

        # ── Mini Roster ───────────────────────────────────────────────────────
        ws.merge_cells(f"A{r_idx}:J{r_idx}")
        sub_hdr(ws.cell(r_idx, 1),
                "WEEKLY ROSTER \u2014 40% MODEL  |  RM covers receptionist off-days  \u2605",
                bg="374F6B")
        ws.row_dimensions[r_idx].height = 20; r_idx += 1

        ws.row_dimensions[r_idx].height = 20
        roster_hdr_vals = ["NAME", "ROLE", "MON", "TUE", "WED", "THU",
                           "FRI \u2605", "SAT \u2605", "SUN \u2605", "DAYS ON"]
        for ci, txt in enumerate(roster_hdr_vals):
            bg = "7D3C98" if "\u2605" in txt else "374F6B"
            hdr(ws.cell(r_idx, 1 + ci), txt, bg=bg, sz=9)
        r_idx += 1

        # Staff rows (therapists + receptionists)
        for si, (name, role, sched) in enumerate(staff_list):
            ws.row_dimensions[r_idx].height = 20
            even = si % 2 == 0
            c = ws.cell(r_idx, 1); lbl(c, name, bold=True)
            c.fill = PatternFill("solid", start_color=LIGHT_BLUE if even else WHITE)
            c = ws.cell(r_idx, 2); lbl(c, role)
            c.fill = PatternFill("solid", start_color=LIGHT_BLUE if even else WHITE)
            for d, code in enumerate(sched):
                shift_cell(ws.cell(r_idx, 3 + d), code)
            days_on = sum(1 for s in sched if s != "OFF")
            c = ws.cell(r_idx, 10); inp(c, days_on, align="center")
            c.fill = PatternFill("solid", start_color="FFFDE7")
            r_idx += 1

        # RM row — inline rendering for M / — / OFF / A codes
        ws.row_dimensions[r_idx].height = 20
        rm_row_bg = "FFF9E6"
        c = ws.cell(r_idx, 1)
        c.value = "RM (\u2153 FTE)"
        c.font = Font(bold=True, color=BLACK, size=10)
        c.fill = PatternFill("solid", start_color=rm_row_bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c = ws.cell(r_idx, 2)
        c.value = "Reg. Manager"
        c.font = Font(color=BLACK, size=10)
        c.fill = PatternFill("solid", start_color=rm_row_bg)
        c.alignment = Alignment(horizontal="left", vertical="center")

        for d, code in enumerate(rm_sched):
            cell = ws.cell(r_idx, 3 + d)
            if code == "M":
                cell.value = "M"
                cell.font = Font(bold=True, color=WHITE, size=10)
                cell.fill = PatternFill("solid", start_color="B8943E")  # gold
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border()
            elif code == "\u2014":
                cell.value = "\u2014"
                cell.font = Font(bold=False, color="888888", size=10)
                cell.fill = PatternFill("solid", start_color="D0D0D0")  # light grey
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border()
            elif code == "OFF":
                cell.value = "OFF"
                cell.font = Font(bold=True, color="888888", size=10)
                cell.fill = PatternFill("solid", start_color="E8E8E8")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border()
            else:
                # "A" — use normal shift_cell
                shift_cell(cell, code)

        days_on_rm = sum(1 for s in rm_sched if s not in ("OFF", "\u2014"))
        c = ws.cell(r_idx, 10)
        inp(c, days_on_rm, align="center")
        c.fill = PatternFill("solid", start_color="FFFDE7")
        r_idx += 1

        ws.row_dimensions[r_idx].height = 14; r_idx += 1   # spacer

    # ── Dispatch one sheet per location ──────────────────────────────────────
    for (sheet_name, annual_rev, r_count, pt_monthly, staff_list) in locations:
        _write_drilldown(sheet_name, annual_rev, r_count, pt_monthly, staff_list)


# =============================================================================
# MAIN
# =============================================================================
def main():
    wb = Workbook()
    build_staffing_model(wb)
    build_roster(wb)
    build_regional(wb)
    add_drilldown_sheets(wb)

    out_dir = "/Users/mertgulen/Library/CloudStorage/GoogleDrive-mertgulen98@gmail.com/My Drive/Carisma Wellness Group/Carisma AI /Carisma AI/HR"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "Carisma Spa Staffing Model.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")
    return out_path


if __name__ == "__main__":
    main()
