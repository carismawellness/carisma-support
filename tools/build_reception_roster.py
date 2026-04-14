"""
build_reception_roster.py
Adds "Reception Roster" sheet to hr/Carisma Spa Staffing Model.xlsx

7-day weekly roster for Inter, Hugos, Hyatt non-therapist staff.
Salary data sourced from image provided Apr 2026.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Palette (mirrors existing model) ──────────────────────────────────────────
NAVY       = "1B3A4B"
GOLD       = "B8943E"
TEAL       = "2C5F77"
WHITE      = "FFFFFF"
BLACK      = "000000"
CORE_BG    = "D6EAF8"
CORE_FNT   = "1B3A4B"
SEAS_BG    = "FAD7A0"
SEAS_FNT   = "A04000"
OK_BG      = "D4EFDF"
DARK_GREEN = "1B7A3B"
UNDER_BG   = "FADBD8"
UNDER_FNT  = "C0392B"
C_GREY     = "F2F3F4"

# Day-cell colours
WORK_BG    = "1B3A4B"   # navy   — full work day
WORK_FNT   = "FFFFFF"
RM_BG      = "2C5F77"   # teal   — RM site visit
RM_FNT     = "FFFFFF"
RT_BG      = "FAD7A0"   # amber  — relief / part-time cover
RT_FNT     = "7D6608"
OFF_BG     = "EAECEE"   # grey   — day off
OFF_FNT    = "BDC3C7"
WE_BG      = "D5D8DC"   # darker grey — weekend off

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
WEEKEND = {5, 6}  # Sat, Sun indices

# ── Roster definition ──────────────────────────────────────────────────────────
# pattern values: 1=WORK  0=OFF  "RM"=RM visit  "RT"=relief cover
# Salaries: monthly gross as loaded (incl. 11.8% SSC), per image Apr 2026
# RM 1/3 FTE: full RM rotates Mon(Inter)→Tue(Hugos)→Wed(3rd site)→Thu(Inter)→Fri(Hugos)
ROSTER = [
    {
        "hotel":    "INTERCONTINENTAL  (INTER)",
        "color":    NAVY,
        "staff": [
            {
                "role":    "RM  ⅓",
                "type":    "Rel. Manager — 3 days/wk here (Mon+Thu+Sat)",
                "salary":  1650,
                "contract":"Part-time allocation",
                "pattern": ["RM", 0, 0, "RM", 0, "RM", 0],  # Mon + Thu + Sat
            },
            {
                "role":    "Supervisor",
                "type":    "Spa Supervisor — Full-time",
                "salary":  2200,
                "contract":"Permanent",
                "pattern": [0, 0, 1, 1, 1, 1, 1],           # Wed–Sun (off Mon+Tue)
            },
            {
                "role":    "Advisor",
                "type":    "Spa Advisor — Full-time",
                "salary":  1750,
                "contract":"Permanent",
                "pattern": [1, 1, 1, 0, 0, 1, 1],           # Mon–Wed+Sat–Sun (off Thu+Fri)
            },
            {
                "role":    "Reception",
                "type":    "Receptionist — Full-time",
                "salary":  1750,
                "contract":"Permanent",
                "pattern": [1, 0, 0, 1, 1, 1, 1],           # Mon+Thu–Sun (off Tue+Wed)
            },
        ],
    },
    {
        "hotel":    "HUGOS",
        "color":    TEAL,
        "staff": [
            {
                "role":    "RM  ⅓",
                "type":    "Rel. Manager — 2 days/wk here (Tue+Fri)",
                "salary":  1100,
                "contract":"Part-time allocation",
                "pattern": [0, "RM", 0, 0, "RM", 0, 0],     # Tue + Fri
            },
            {
                "role":    "Supervisor",
                "type":    "Spa Supervisor — Full-time",
                "salary":  2200,
                "contract":"Permanent",
                "pattern": [0, 1, 0, 1, 1, 1, 1],           # Tue+Thu–Sun (off Mon+Wed)
            },
            {
                "role":    "Advisor",
                "type":    "Spa Advisor — Full-time",
                "salary":  1750,
                "contract":"Permanent",
                "pattern": [1, 0, 1, 0, 1, 1, 1],           # Mon+Wed+Fri–Sun (off Tue+Thu)
            },
            {
                "role":    "RT Reception",
                "type":    "Relief Receptionist — Part-time",
                "salary":  1000,
                "contract":"Part-time / relief",
                "pattern": [1, 0, 0, 0, 0, "RT", "RT"],     # Mon + Sat + Sun
            },
        ],
    },
    {
        "hotel":    "HYATT",
        "color":    "4A235A",
        "staff": [
            {
                "role":    "Supervisor",
                "type":    "Spa Supervisor — Full-time",
                "salary":  2200,
                "contract":"Permanent",
                "pattern": [0, 1, 1, 1, 0, 1, 1],           # Tue–Thu+Sat–Sun (off Mon+Fri)
            },
            {
                "role":    "RT Reception",
                "type":    "Relief Receptionist — Weekends only",
                "salary":  750,
                "contract":"Part-time / weekend",
                "pattern": [0, 0, 0, 0, 0, "RT", "RT"],     # Sat + Sun only
            },
        ],
    },
]


# ── Style helpers ──────────────────────────────────────────────────────────────
def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_top(color=NAVY):
    t = Side(style="thin",   color="CCCCCC")
    m = Side(style="medium", color=color)
    return Border(left=t, right=t, top=m, bottom=t)

def _hdr(cell, text, bg=NAVY, fg=WHITE, bold=True, sz=10, wrap=False, align="center"):
    cell.value = text
    cell.font = Font(bold=bold, color=fg, size=sz)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin()

def _lbl(cell, text, bold=False, align="left", color=BLACK, sz=10, bg=WHITE, wrap=False):
    cell.value = text
    cell.font = Font(bold=bold, color=color, size=sz)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin()

def _day(cell, val, di):
    is_we = di in WEEKEND
    if val == 1:
        bg, fg, txt, bold = WORK_BG, WORK_FNT, "WORK", True
    elif val == "RM":
        bg, fg, txt, bold = RM_BG, RM_FNT, "RM VISIT", True
    elif val == "RT":
        bg, fg, txt, bold = RT_BG, RT_FNT, "COVER", True
    else:
        bg = WE_BG if is_we else OFF_BG
        fg, txt, bold = OFF_FNT, "—", False
    cell.value = txt
    cell.font = Font(bold=bold, color=fg, size=9)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin()


# ── Main builder ───────────────────────────────────────────────────────────────
def build_roster(wb):
    if "Reception Roster" in wb.sheetnames:
        del wb["Reception Roster"]

    ws = wb.create_sheet("Reception Roster")
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 14    # Role
    ws.column_dimensions["B"].width = 34    # Type description
    ws.column_dimensions["C"].width = 12    # Contract
    ws.column_dimensions["D"].width = 12    # Monthly salary
    for i in range(7):
        ws.column_dimensions[get_column_letter(5 + i)].width = 9   # Mon–Sun
    ws.column_dimensions["L"].width = 9     # Days/wk
    ws.column_dimensions["M"].width = 12    # Monthly cost
    ws.column_dimensions["N"].width = 30    # Notes

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:N1")
    _hdr(ws["A1"],
         "CARISMA SPA  —  WEEKLY RECEPTION ROSTER  |  Inter  ·  Hugos  ·  Hyatt  "
         "|  Non-Therapist Staff  |  7-Day Coverage View",
         bg=NAVY, sz=13, align="left")
    ws.row_dimensions[1].height = 30

    # ── Row 2: Legend ─────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:C2")
    _hdr(ws["A2"], "LEGEND", bg=NAVY, sz=9)
    for ci, (bg, fg, lbl) in enumerate([
        (WORK_BG, WORK_FNT, "WORK — full day"),
        (RM_BG,   RM_FNT,   "RM VISIT — site day"),
        (RT_BG,   RT_FNT,   "COVER — relief/PT"),
        (OFF_BG,  OFF_FNT,  "— off (weekday)"),
        (WE_BG,   OFF_FNT,  "— off (weekend)"),
    ]):
        c = ws.cell(2, 4 + ci)
        c.value = lbl
        c.font = Font(bold=(bg not in (OFF_BG, WE_BG)), color=fg, size=9)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin()
    for ci in range(9, 14):
        c = ws.cell(2, ci)
        c.fill = PatternFill("solid", start_color="F4F6F7")
        c.border = _thin()

    # ── Row 3: Column headers ─────────────────────────────────────────────────
    ws.row_dimensions[3].height = 30
    col_hdrs = [
        ("Role",           NAVY),
        ("Staff / Contract Type",  NAVY),
        ("Contract",       NAVY),
        ("Monthly\nSalary (€)", GOLD),
        ("Mon",  WORK_BG),
        ("Tue",  WORK_BG),
        ("Wed",  WORK_BG),
        ("Thu",  WORK_BG),
        ("Fri",  WORK_BG),
        ("Sat",  "4A5568"),
        ("Sun",  "4A5568"),
        ("Days /\nWeek",   NAVY),
        ("Monthly\nCost (€)", GOLD),
        ("Notes",          NAVY),
    ]
    for ci, (hd, bg) in enumerate(col_hdrs):
        c = ws.cell(3, 1 + ci)
        c.value = hd
        c.font = Font(bold=True, color=WHITE, size=9)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin()

    r = 4

    for hotel in ROSTER:
        # Hotel header row
        ws.row_dimensions[r].height = 26
        ws.merge_cells(f"A{r}:N{r}")
        _hdr(ws.cell(r, 1), hotel["hotel"],
             bg=hotel["color"], sz=12, bold=True, align="left")
        r += 1

        coverage = [0] * 7   # staff on duty per day (excl. RM)
        hotel_monthly_cost = 0

        for s in hotel["staff"]:
            ws.row_dimensions[r].height = 22

            # Role cell background
            if "RM" in s["role"]:
                rbg, rfg = "EAF4FB", TEAL
            elif s["role"] == "Supervisor":
                rbg, rfg = CORE_BG, CORE_FNT
            elif s["role"] == "RT Reception":
                rbg, rfg = SEAS_BG, SEAS_FNT
            else:
                rbg, rfg = "F8F9FA", BLACK

            _lbl(ws.cell(r, 1), s["role"],     bold=True, bg=rbg, color=rfg)
            _lbl(ws.cell(r, 2), s["type"],      sz=9,     bg=rbg, color=rfg)
            _lbl(ws.cell(r, 3), s["contract"],  sz=9,     bg=rbg, color=rfg)

            # Salary (blue = hardcoded input)
            c = ws.cell(r, 4)
            c.value = s["salary"]
            c.font = Font(color="0000FF", size=10, bold=True)
            c.fill = PatternFill("solid", start_color=rbg)
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "€#,##0"
            c.border = _thin()

            # Day cells
            days_worked = 0
            for di, val in enumerate(s["pattern"]):
                _day(ws.cell(r, 5 + di), val, di)
                if val != 0:
                    days_worked += 1
                    if val != "RM":
                        coverage[di] += 1

            # Days/week
            c_d = ws.cell(r, 12)
            c_d.value = days_worked
            c_d.font = Font(bold=True, color=rfg, size=11)
            c_d.fill = PatternFill("solid", start_color=rbg)
            c_d.alignment = Alignment(horizontal="center", vertical="center")
            c_d.border = _thin()

            # Monthly cost = salary (shown blue as input)
            c_m = ws.cell(r, 13)
            c_m.value = f"=D{r}"
            c_m.font = Font(color=BLACK, size=10)
            c_m.fill = PatternFill("solid", start_color=rbg)
            c_m.alignment = Alignment(horizontal="right", vertical="center")
            c_m.number_format = "€#,##0"
            c_m.border = _thin()

            # Notes
            note = ""
            if "RM" in s["role"]:
                if hotel["hotel"].startswith("INTERCONTINENTAL"):
                    note = "Rotates: Mon+Thu+Sat here · Tue+Fri at Hugos · Wed at 3rd site  |  3/6 days → €1,650/mo"
                elif hotel["hotel"].startswith("HUGOS"):
                    note = "Rotates: Tue+Fri here · Mon+Thu+Sat at Inter · Wed at 3rd site  |  2/6 days → €1,100/mo"
            elif s["role"] == "Supervisor":
                days_on  = [DAYS[i] for i, v in enumerate(s["pattern"]) if v == 1]
                days_off = [DAYS[i] for i, v in enumerate(s["pattern"]) if v == 0]
                note = f"Works {'+'.join(days_on)} — off {'+'.join(days_off)}  (Sat+Sun mandatory)"
            elif s["role"] == "Advisor":
                days_on  = [DAYS[i] for i, v in enumerate(s["pattern"]) if v == 1]
                days_off = [DAYS[i] for i, v in enumerate(s["pattern"]) if v == 0]
                note = f"Works {'+'.join(days_on)} — off {'+'.join(days_off)}  (Sat+Sun mandatory)"
            elif s["role"] == "Reception":
                days_on  = [DAYS[i] for i, v in enumerate(s["pattern"]) if v == 1]
                days_off = [DAYS[i] for i, v in enumerate(s["pattern"]) if v == 0]
                note = f"Works {'+'.join(days_on)} — off {'+'.join(days_off)}  (Sat+Sun mandatory)"
            elif s["role"] == "RT Reception" and s["salary"] == 1000:
                note = "Mon + Sat + Sun  (3 days) — boosts peak weekend front desk coverage"
            elif s["role"] == "RT Reception" and s["salary"] == 750:
                note = "Sat+Sun only — Supervisor covers Tue–Thu front desk"

            c_n = ws.cell(r, 14)
            _lbl(c_n, note, sz=8, color="666666", bg="FAFAFA", wrap=True)

            hotel_monthly_cost += s["salary"]
            r += 1

        # Coverage row
        ws.row_dimensions[r].height = 22
        ws.cell(r, 1).value = "Front desk covered by:"
        ws.cell(r, 1).font = Font(bold=True, color=WHITE, size=9)
        ws.cell(r, 1).fill = PatternFill("solid", start_color=GOLD)
        ws.cell(r, 1).border = _thin()
        ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")

        ws.cell(r, 2).value = "No. of staff on duty each day (reception/relief — excl. RM)"
        ws.cell(r, 2).font = Font(italic=True, color=WHITE, size=9)
        ws.cell(r, 2).fill = PatternFill("solid", start_color=GOLD)
        ws.cell(r, 2).border = _thin()
        ws.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center")

        for ci in (3, 4):
            ws.cell(r, ci).fill = PatternFill("solid", start_color=GOLD)
            ws.cell(r, ci).border = _thin()

        for di, count in enumerate(coverage):
            c = ws.cell(r, 5 + di)
            if count == 0:
                bg_c, fg_c = UNDER_BG, UNDER_FNT
                txt = "⚠ 0"
            elif count == 1:
                bg_c, fg_c = OK_BG, DARK_GREEN
                txt = "✓ 1"
            else:
                bg_c, fg_c = CORE_BG, CORE_FNT
                txt = f"  {count}"
            c.value = txt
            c.font = Font(bold=True, color=fg_c, size=10)
            c.fill = PatternFill("solid", start_color=bg_c)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _thin()

        # Days/week and total cost cells in coverage row
        ws.cell(r, 12).fill = PatternFill("solid", start_color=GOLD)
        ws.cell(r, 12).border = _thin()

        # Total monthly non-therapist cost for this hotel
        first_staff_row = r - len(hotel["staff"])
        last_staff_row  = r - 1
        c_total = ws.cell(r, 13)
        c_total.value = f"=SUM(M{first_staff_row}:M{last_staff_row})"
        c_total.font = Font(bold=True, color=WHITE, size=10)
        c_total.fill = PatternFill("solid", start_color=GOLD)
        c_total.alignment = Alignment(horizontal="right", vertical="center")
        c_total.number_format = "€#,##0"
        c_total.border = _thin()

        ws.cell(r, 14).value = "← Total non-therapist monthly cost"
        ws.cell(r, 14).font = Font(italic=True, color=WHITE, size=9)
        ws.cell(r, 14).fill = PatternFill("solid", start_color=GOLD)
        ws.cell(r, 14).border = _thin()
        ws.cell(r, 14).alignment = Alignment(horizontal="left", vertical="center")

        r += 2   # gap between hotels

    # ── RM Rotation explainer table ───────────────────────────────────────────
    ws.row_dimensions[r].height = 8
    r += 1

    ws.merge_cells(f"A{r}:N{r}")
    _hdr(ws.cell(r, 1),
         "RM ROTATION SCHEDULE  —  1 full-time RM (6 days/wk, Sun off) shared across 3 sites  "
         "|  Cost split proportional to days: Inter 3/6 · Hugos 2/6 · 3rd 1/6  |  Total: €3,300/mo",
         bg=TEAL, sz=9, bold=False, align="left")
    ws.row_dimensions[r].height = 20
    r += 1

    # RM table headers
    ws.row_dimensions[r].height = 22
    _hdr(ws.cell(r, 1), "Site", bg=TEAL, sz=9)
    _hdr(ws.cell(r, 2), "Days at this site", bg=TEAL, sz=9)
    _hdr(ws.cell(r, 3), "Monthly cost", bg=TEAL, sz=9)
    for di, d in enumerate(DAYS):
        _hdr(ws.cell(r, 4 + di), d, bg=TEAL, sz=9)
    _hdr(ws.cell(r, 11), "Days/wk", bg=TEAL, sz=9)
    ws.merge_cells(f"L{r}:N{r}")
    _hdr(ws.cell(r, 12), "Notes", bg=TEAL, sz=9)
    r += 1

    rm_sites = [
        ("Intercontinental", NAVY,    ["RM", 0, 0, "RM", 0, "RM", 0], "Mon + Thu + Sat  (3 days)", 1650),
        ("Hugos",            TEAL,    [0, "RM", 0, 0, "RM", 0, 0],    "Tue + Fri  (2 days)",       1100),
        ("3rd site (TBC)",   "888888",[0, 0, "RM", 0, 0, 0, 0],       "Wed  (1 day)",               550),
    ]
    for site, sc, pat, days_txt, cost in rm_sites:
        ws.row_dimensions[r].height = 20
        _lbl(ws.cell(r, 1), site,     bold=True, bg="F8FAFC", color=sc)
        _lbl(ws.cell(r, 2), days_txt, sz=9,      bg="F8FAFC", color=sc)
        c = ws.cell(r, 3)
        c.value = cost
        c.font = Font(color="0000FF", size=10)
        c.fill = PatternFill("solid", start_color="F8FAFC")
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = "€#,##0"
        c.border = _thin()
        for di, val in enumerate(pat):
            _day(ws.cell(r, 4 + di), val, di)
        days_count = sum(1 for v in pat if v != 0)
        c_d = ws.cell(r, 11)
        c_d.value = days_count
        c_d.font = Font(bold=True, color=sc, size=10)
        c_d.fill = PatternFill("solid", start_color="F8FAFC")
        c_d.alignment = Alignment(horizontal="center", vertical="center")
        c_d.border = _thin()
        ws.merge_cells(f"L{r}:N{r}")
        note = ""
        if site == "3rd site (TBC)":
            note = "1 day (Wed) — assign to Ramla, Excelsior, or Riviera as needed  |  1/6 × €3,300 = €550"
        elif site == "Intercontinental":
            note = "Highest-revenue site gets 3 days (incl. Sat — busiest day)  |  3/6 × €3,300 = €1,650"
        elif site == "Hugos":
            note = "2/6 × €3,300 = €1,100"
        _lbl(ws.cell(r, 12), note, sz=8, color="888888", bg="F8FAFC")
        r += 1

    # RM Total row
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f"A{r}:B{r}")
    _hdr(ws.cell(r, 1), "TOTAL RM LOADED SALARY", bg=NAVY, sz=9, align="right")
    first_rm = r - 3
    c_rm = ws.cell(r, 3)
    c_rm.value = f"=SUM(C{first_rm}:C{r-1})"
    c_rm.font = Font(bold=True, color=WHITE, size=10)
    c_rm.fill = PatternFill("solid", start_color=NAVY)
    c_rm.alignment = Alignment(horizontal="right", vertical="center")
    c_rm.number_format = "€#,##0"
    c_rm.border = _thin()
    for ci in range(4, 15):
        ws.cell(r, ci).fill = PatternFill("solid", start_color=NAVY)
        ws.cell(r, ci).border = _thin()

    ws.freeze_panes = "E4"


# ── Entry point ───────────────────────────────────────────────────────────────
def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    repo_root  = os.path.dirname(script_dir)
    path = os.path.join(repo_root, "hr", "Carisma Spa Staffing Model.xlsx")

    wb = load_workbook(path)
    build_roster(wb)

    # Move "Reception Roster" right after "Weekly Roster" if it exists
    sheets = wb.sheetnames
    target = "Weekly Roster"
    if target in sheets:
        idx_after = sheets.index(target) + 1
        offset = idx_after - (len(sheets) - 1)
        wb.move_sheet("Reception Roster", offset=offset)

    wb.save(path)
    print(f"✓ Saved: {path}")
    print(f"  Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
