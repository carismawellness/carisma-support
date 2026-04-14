"""
recalc.py — Recalculate Excel formulas using LibreOffice and scan for errors.
Usage: python recalc.py <excel_file> [timeout_seconds]
Returns JSON with error details.
"""
import subprocess
import sys
import json
import os
import tempfile
import shutil

ERROR_TYPES = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}

MACRO_SCRIPT = """
import uno
import sys
import os

def recalc_and_save(filepath):
    localContext = uno.getComponentContext()
    resolver = localContext.ServiceManager.createInstanceWithContext(
        "com.sun.star.bridge.UnoUrlResolver", localContext)
    try:
        ctx = resolver.resolve(
            "uno:socket,host=localhost,port=2002;urp;StarOffice.ComponentContext")
    except Exception:
        sys.exit(1)
    smgr = ctx.ServiceManager
    desktop = smgr.createInstanceWithContext("com.sun.star.frame.Desktop", ctx)
    url = "file://" + filepath
    doc = desktop.loadComponentFromURL(url, "_blank", 0, ())
    doc.calculateAll()
    doc.store()
    doc.close(True)

recalc_and_save(sys.argv[1])
"""


def recalc_with_libreoffice(filepath: str, timeout: int = 60) -> dict:
    abs_path = os.path.abspath(filepath)

    # Check LibreOffice is available
    lo_bin = shutil.which("soffice") or shutil.which("libreoffice")
    if not lo_bin:
        return {"status": "error", "message": "LibreOffice not found. Install with: brew install libreoffice"}

    # Use LibreOffice headless to recalculate
    cmd = [
        lo_bin, "--headless", "--calc",
        "--infilter=Calc MS Excel 2007 XML",
        "--outdir", os.path.dirname(abs_path),
        abs_path
    ]

    # Better approach: use macro conversion
    cmd = [
        lo_bin, "--headless",
        "--convert-to", "xlsx",
        "--outdir", os.path.dirname(abs_path),
        abs_path
    ]

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    except subprocess.TimeoutExpired:
        return {"status": "error", "message": f"LibreOffice timed out after {timeout}s"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

    # Now scan for errors using openpyxl
    return scan_errors(abs_path)


def scan_errors(filepath: str) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"status": "error", "message": "openpyxl not installed. Run: pip install openpyxl"}

    wb = load_workbook(filepath, data_only=True)
    errors = {}
    total_formulas = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                val_str = str(val).strip().upper()
                for err in ERROR_TYPES:
                    if val_str == err or val_str.startswith(err):
                        location = f"{sheet_name}!{cell.coordinate}"
                        errors.setdefault(err, {"count": 0, "locations": []})
                        errors[err]["count"] += 1
                        errors[err]["locations"].append(location)
                # Count formula cells (when data_only=True, formulas are evaluated)
                if isinstance(val, (int, float)):
                    total_formulas += 1

    if errors:
        return {
            "status": "errors_found",
            "total_errors": sum(v["count"] for v in errors.values()),
            "total_formulas": total_formulas,
            "error_summary": errors
        }
    return {
        "status": "success",
        "total_errors": 0,
        "total_formulas": total_formulas,
        "error_summary": {}
    }


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"status": "error", "message": "Usage: python recalc.py <excel_file> [timeout]"}))
        sys.exit(1)

    filepath = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 60

    if not os.path.exists(filepath):
        print(json.dumps({"status": "error", "message": f"File not found: {filepath}"}))
        sys.exit(1)

    # Try LibreOffice recalc first, fallback to scan only
    lo_bin = shutil.which("soffice") or shutil.which("libreoffice")
    if lo_bin:
        result = recalc_with_libreoffice(filepath, timeout)
    else:
        # Just scan without recalc
        result = scan_errors(filepath)

    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
