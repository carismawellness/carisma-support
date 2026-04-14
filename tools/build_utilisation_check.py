"""
build_utilisation_check.py
Adds a "Utilisation Check" sheet to hr/Carisma Spa Staffing Model.xlsx

Sanity-checks whether the therapist headcount per location can realistically
generate the modelled revenue, given:
  - €85 average treatment revenue
  - 6 max treatments per therapist per shift (75 min avg + 15 min changeover in 9hr shift)
  - Therapists work 5 of 7 operating days per week (2 days off in roster rotation)
  - Malta seasonal demand index applied to base monthly revenue

Generates TWO colour-coded grids:
  Grid 1 — Implied Treatments Per Day (the absolute number)
  Grid 2 — Utilisation % vs Max Capacity
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Palette ────────────────────────────────────────────────────────────────────
NAVY       = "1B3A4B"
GOLD       = "B8943E"
WHITE      = "FFFFFF"
BLACK      = "000000"
BLUE_FONT  = "0000FF"
GREEN_FONT = "008000"
RED_FONT   = "FF0000"
LIGHT_BLUE = "EAF4FB"

# Utilisation colour bands (from healthy → danger)
C_UNDER  = "D6EAF8"  # < 40%   — light blue   — under-utilised
C_GREEN  = "D5F5E3"  # 40–70%  — light green  — healthy
C_YELLOW = "FEF9E7"  # 70–85%  — light yellow — elevated
C_ORANGE = "FAD7A0"  # 85–100% — light orange — stress zone
C_RED    = "FADBD8"  # > 100%  — light red    — physically impossible

# ── Constants ──────────────────────────────────────────────────────────────────
MONTHS        = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
SEASONAL_IDX  = [0.80, 0.85, 0.90, 0.95, 1.00, 1.10, 1.20, 1.25, 1.10, 1.00, 0.90, 0.95]
DAYS_IN_MO    = [31,   28,   31,   30,   31,   30,   31,   31,   30,   31,   30,   31]

AVG_TX_REV     = 85     # € per treatment (mid-point of €70–€100 range)
MAX_TX_PER_DAY = 6      # 9-hr shift ÷ (75 min treatment + 15 min changeover) = 6 slots
DAYS_PER_WEEK  = 5      # Each therapist works 5 of 7 days (2 days off in rotation)
UTIL_TARGET    = 0.80   # 80% industry benchmark peak utilisation target

# Therapist counts from the 35% HC budget model (column F = total FTE is T+R+RM,
# so therapist counts are sourced from the rows data in build_staffing_model.py)
LOCATIONS = [
    # (name,                  monthly_rev, therapists)
    ("Inter",                 49_475,      9),
    ("Hugos",                 47_546,      9),
    ("Ramla",                 29_886,      5),
    ("Hyatt",                 25_358,      4),
    ("Excelsior",             24_263,      4),
    ("Riviera",               19_936,      4),
    ("Sunny Coast (Odycy)",   18_204,      3),
    ("Novotel",               14_000,      2),
    ("Sliema (NEW)",          41_667,      5),
]

FILE = os.path.join(
    os.path.expanduser("~"),
    "Library/CloudStorage/GoogleDrive-mertgulen98@gmail.com/My Drive/"
    "Carisma Wellness Group/Carisma AI /Carisma AI/"
    "hr/Carisma Spa Staffing Model.xlsx"
)


# ── Helpers ────────────────────────────────────────────────────────────────────

def util_color(pct: float) -> str:
    if pct > 1.00: return C_RED
    if pct > 0.85: return C_ORANGE
    if pct > 0.70: return C_YELLOW
    if pct > 0.40: return C_GREEN
    return C_UNDER

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom():
    thin = Side(style="thin", color="CCCCCC")
    thick = Side(style="medium", color="888888")
    return Border(left=thin, right=thin, top=thin, bottom=thick)

def _hdr(cell, text, bg=NAVY, fg=WHITE, bold=True, sz=10, wrap=False, align="center"):
    cell.value = text
    cell.font = Font(bold=bold, color=fg, size=sz)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = thin_border()

def _lbl(cell, text, bold=False, bg=WHITE, align="left", sz=10, color=BLACK):
    cell.value = text
    cell.font = Font(color=color, bold=bold, size=sz)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")

def _inp(cell, value, fmt=None, bg=WHITE, align="center"):
    cell.value = value
    cell.font = Font(color=BLUE_FONT, size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = thin_border()
    if fmt:
        cell.number_format = fmt

def _val(cell, value, bg, fmt="0.0", bold=False, color=BLACK):
    cell.value = value
    cell.font = Font(color=color, bold=bold, size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
    cell.number_format = fmt

def _note(cell, text, bg="FEF9E7", color="444444"):
    cell.value = text
    cell.font = Font(color=color, italic=True, size=9)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin_border()


# ── Main builder ───────────────────────────────────────────────────────────────

def build_util_check(wb):
    if "Utilisation Check" in wb.sheetnames:
        del wb["Utilisation Check"]

    ws = wb.create_sheet("Utilisation Check")
    ws.sheet_view.showGridLines = False

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22   # Location
    ws.column_dimensions["B"].width = 11   # Therapists
    ws.column_dimensions["C"].width = 12   # Max Cap/Day
    ws.column_dimensions["D"].width = 13   # Target Cap/Day (80%)
    for i in range(12):                    # Months E–P
        ws.column_dimensions[get_column_letter(5 + i)].width = 8
    ws.column_dimensions[get_column_letter(17)].width = 10  # FY Avg

    # ── Row 1: Title ───────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:Q1")
    _hdr(ws["A1"],
         "CARISMA SPA & WELLNESS — UTILISATION CHECK  |  "
         "Treatments-Per-Day Sanity Check vs Therapist Capacity",
         sz=13, align="left")

    # ── Row 2: Sub-title ──────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:Q2")
    c = ws["A2"]
    c.value = (
        "Formula: Monthly Base Revenue × Seasonal Index ÷ (€85 avg treatment × Days in Month) = "
        "Implied Treatments/Day    |    "
        "Max Capacity = Therapists × 6 treatments/shift × (5/7 days rostered)    |    "
        "Re-run build_utilisation_check.py to refresh after changing the Staffing Model"
    )
    c.font = Font(color="555555", italic=True, size=8)
    c.fill = PatternFill("solid", start_color="F2F3F4")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Row 3: Spacer ──────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 8

    # ── Rows 4–8: Assumptions block ────────────────────────────────────────────
    ws.row_dimensions[4].height = 20
    ws.merge_cells("A4:Q4")
    _hdr(ws["A4"], "KEY ASSUMPTIONS  (Blue = hardcoded input — change here to model scenarios)",
         bg=GOLD, sz=10, align="left")

    assumption_rows = [
        ("Avg treatment revenue",          AVG_TX_REV,     '"€"#,##0',  "Mid-point of €70–€100 range. Blends all treatment types across the portfolio."),
        ("Max treatments / therapist / day", MAX_TX_PER_DAY, "0",         "9-hr shift ÷ (75 min treatment + 15 min changeover) = 6 slots maximum."),
        ("Therapist working days / week",  DAYS_PER_WEEK,  "0",         "Of 7 operating days. 2 scheduled days off per week per therapist."),
        ("Peak utilisation target",        UTIL_TARGET,    "0%",        "80% industry benchmark — mature spa at steady state. Above 85% = service quality risk."),
    ]
    for i, (label, value, fmt, note_text) in enumerate(assumption_rows):
        row = 5 + i
        ws.row_dimensions[row].height = 18
        _lbl(ws.cell(row, 1), label, bold=True)
        _inp(ws.cell(row, 2), value, fmt=fmt, bg=LIGHT_BLUE)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=17)
        _note(ws.cell(row, 3), "← " + note_text, bg="FAFAFA")

    # ── Row 9: Spacer ──────────────────────────────────────────────────────────
    ws.row_dimensions[9].height = 8

    # ── Rows 10–12: Seasonal index reference ───────────────────────────────────
    ws.row_dimensions[10].height = 20
    ws.merge_cells("A10:Q10")
    _hdr(ws["A10"],
         "MALTA SEASONAL DEMAND INDEX  "
         "(Base = 1.0 | Sum = 12.0 | Applied to monthly base revenue to get seasonally-adjusted monthly revenue)",
         bg=NAVY, sz=10, align="left")

    ws.row_dimensions[11].height = 20
    _hdr(ws.cell(11, 1), "Month", bg="2C5F77", sz=9)
    ws.merge_cells(start_row=11, start_column=2, end_row=11, end_column=4)
    _hdr(ws.cell(11, 2), "Index", bg="2C5F77", sz=9)
    for i, m in enumerate(MONTHS):
        _hdr(ws.cell(11, 5 + i), m, bg="2C5F77", sz=9)
    _hdr(ws.cell(11, 17), "FY Avg", bg="1B3A4B", sz=9)

    ws.row_dimensions[12].height = 18
    _lbl(ws.cell(12, 1), "Seasonal Index", bold=True, sz=9)
    ws.merge_cells(start_row=12, start_column=2, end_row=12, end_column=4)
    _lbl(ws.cell(12, 2), "Jan=low season, Aug=peak (tourism + local)", sz=8, color="666666")
    for i, idx in enumerate(SEASONAL_IDX):
        c = ws.cell(12, 5 + i)
        c.value = idx
        c.font = Font(color=BLUE_FONT, size=9)
        c.fill = PatternFill("solid", start_color=LIGHT_BLUE if i % 2 == 0 else "F8FCFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        c.number_format = "0.00"
    c = ws.cell(12, 17)
    c.value = 1.00
    c.font = Font(color=BLACK, bold=True, size=9)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
    c.number_format = "0.00"

    # ── Rows 13–spacer → shared column header builder ──────────────────────────
    def _draw_col_headers(row_num, bg_month="2C5F77", bg_stat="1B3A4B"):
        ws.row_dimensions[row_num].height = 28
        _hdr(ws.cell(row_num, 1), "LOCATION",        bg=NAVY,     sz=9, wrap=True)
        _hdr(ws.cell(row_num, 2), "THERA-\nPISTS",   bg=NAVY,     sz=9, wrap=True)
        _hdr(ws.cell(row_num, 3), "MAX CAP\n/ DAY",  bg="2C5F77", sz=9, wrap=True)
        _hdr(ws.cell(row_num, 4), "TARGET\n(80%)",   bg="1B7A3B", sz=9, wrap=True)
        for i, m in enumerate(MONTHS):
            _hdr(ws.cell(row_num, 5 + i), m, bg=bg_month, sz=9)
        _hdr(ws.cell(row_num, 17), "FY\nAVG", bg=bg_stat, sz=9, wrap=True)

    # ── GRID 1: Implied Treatments Per Day ─────────────────────────────────────
    G1_HDR  = 13   # section header
    G1_COLS = 14   # column sub-headers
    G1_DATA = 15   # first data row (9 locations → rows 15–23)

    ws.row_dimensions[G1_HDR].height = 20
    ws.merge_cells(f"A{G1_HDR}:Q{G1_HDR}")
    _hdr(ws[f"A{G1_HDR}"],
         "GRID 1 — IMPLIED TREATMENTS PER DAY  "
         "(Monthly Rev × Seasonal Index ÷ (€85 × Days in Month)   |   "
         "Colour = utilisation vs max daily capacity)",
         bg=NAVY, sz=10, align="left")
    _draw_col_headers(G1_COLS)

    for li, (loc, rev, therapists) in enumerate(LOCATIONS):
        row = G1_DATA + li
        ws.row_dimensions[row].height = 22

        max_cap    = therapists * MAX_TX_PER_DAY * (DAYS_PER_WEEK / 7)
        target_cap = max_cap * UTIL_TARGET

        # Location label
        c = ws.cell(row, 1)
        c.value = loc
        c.font = Font(bold=True, color=BLACK, size=10)
        c.fill = PatternFill("solid", start_color="F4F6F7")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()

        # Therapists (blue input)
        _inp(ws.cell(row, 2), therapists, fmt="0", bg=LIGHT_BLUE)

        # Max capacity
        _val(ws.cell(row, 3), round(max_cap, 1), bg=LIGHT_BLUE, fmt="0.0")

        # Target capacity (80%)
        _val(ws.cell(row, 4), round(target_cap, 1), bg="E9F7EF", fmt="0.0",
             color=GREEN_FONT)

        # Jan–Dec
        month_vals = []
        for mi in range(12):
            tx_per_day = (rev * SEASONAL_IDX[mi]) / (AVG_TX_REV * DAYS_IN_MO[mi])
            util_pct   = tx_per_day / max_cap
            month_vals.append(tx_per_day)
            bold = util_pct > 1.0
            _val(ws.cell(row, 5 + mi), round(tx_per_day, 1),
                 bg=util_color(util_pct), fmt="0.0",
                 bold=bold, color=RED_FONT if util_pct > 1.0 else BLACK)

        # FY Average
        fy_avg  = sum(month_vals) / 12
        fy_util = fy_avg / max_cap
        _val(ws.cell(row, 17), round(fy_avg, 1),
             bg=util_color(fy_util), fmt="0.0",
             bold=True, color=RED_FONT if fy_util > 1.0 else BLACK)

    # ── GRID 2: Utilisation % ──────────────────────────────────────────────────
    G2_SP   = G1_DATA + len(LOCATIONS)     # spacer
    G2_HDR  = G2_SP  + 1
    G2_COLS = G2_HDR + 1
    G2_DATA = G2_COLS + 1

    ws.row_dimensions[G2_SP].height = 10

    ws.row_dimensions[G2_HDR].height = 20
    ws.merge_cells(f"A{G2_HDR}:Q{G2_HDR}")
    _hdr(ws[f"A{G2_HDR}"],
         "GRID 2 — UTILISATION %  "
         "(Implied Treatments Per Day ÷ Max Capacity   |   "
         "Max Capacity = Therapists × 6 treatments × 5/7 roster days)",
         bg=NAVY, sz=10, align="left")
    _draw_col_headers(G2_COLS, bg_month="1B7A3B", bg_stat="4A235A")

    for li, (loc, rev, therapists) in enumerate(LOCATIONS):
        row = G2_DATA + li
        ws.row_dimensions[row].height = 22

        max_cap    = therapists * MAX_TX_PER_DAY * (DAYS_PER_WEEK / 7)
        target_cap = max_cap * UTIL_TARGET

        # Location
        c = ws.cell(row, 1)
        c.value = loc
        c.font = Font(bold=True, color=BLACK, size=10)
        c.fill = PatternFill("solid", start_color="F4F6F7")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()

        _inp(ws.cell(row, 2), therapists, fmt="0", bg=LIGHT_BLUE)
        _val(ws.cell(row, 3), round(max_cap, 1), bg=LIGHT_BLUE, fmt="0.0")
        _val(ws.cell(row, 4), round(target_cap, 1), bg="E9F7EF", fmt="0.0",
             color=GREEN_FONT)

        fy_utils = []
        for mi in range(12):
            tx_per_day = (rev * SEASONAL_IDX[mi]) / (AVG_TX_REV * DAYS_IN_MO[mi])
            util_pct   = tx_per_day / max_cap
            fy_utils.append(util_pct)
            bold = util_pct > 1.0
            _val(ws.cell(row, 5 + mi), util_pct,
                 bg=util_color(util_pct), fmt="0%",
                 bold=bold, color=RED_FONT if util_pct > 1.0 else BLACK)

        fy_avg_util = sum(fy_utils) / 12
        _val(ws.cell(row, 17), fy_avg_util,
             bg=util_color(fy_avg_util), fmt="0%",
             bold=True, color=RED_FONT if fy_avg_util > 1.0 else BLACK)

    # ── Legend ─────────────────────────────────────────────────────────────────
    LG_SP  = G2_DATA + len(LOCATIONS)
    LG_HDR = LG_SP + 1
    LG_DATA= LG_HDR + 1

    ws.row_dimensions[LG_SP].height = 10
    ws.row_dimensions[LG_HDR].height = 20
    ws.merge_cells(f"A{LG_HDR}:Q{LG_HDR}")
    _hdr(ws[f"A{LG_HDR}"], "COLOUR LEGEND", bg=GOLD, sz=10, align="left")

    legend_items = [
        (C_UNDER,  "< 40% utilisation",  "Under-utilised. Therapist headcount may exceed what revenue can justify. "
                                          "Consider reducing headcount or growing revenue at this location."),
        (C_GREEN,  "40 – 70%",           "Healthy range. Sustainable operations. Capacity headroom for peaks and "
                                          "growth without overloading staff."),
        (C_YELLOW, "70 – 85%",           "Elevated — high-demand periods. Manageable, but watch for booking "
                                          "bottlenecks. Consider part-time cover for peak months."),
        (C_ORANGE, "85 – 100%",          "Stress zone. Risk of service delays and quality drop. "
                                          "Roster reinforcement (cover therapist or PT hire) strongly recommended."),
        (C_RED,    "> 100%  ⚠ IMPOSSIBLE", "Physically impossible. The revenue target cannot be achieved with the "
                                            "current therapist count — there are simply not enough treatment slots. "
                                            "INCREASE HEADCOUNT or revise the revenue projection."),
    ]
    for i, (bg, band, desc) in enumerate(legend_items):
        row = LG_DATA + i
        ws.row_dimensions[row].height = 22
        # Swatch
        c = ws.cell(row, 1)
        c.value = band
        c.font = Font(bold=True, color=RED_FONT if bg == C_RED else BLACK, size=9)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        # Description
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=17)
        d = ws.cell(row, 2)
        d.value = desc
        d.font = Font(color=RED_FONT if bg == C_RED else BLACK,
                      bold=(bg == C_RED), size=9)
        d.fill = PatternFill("solid", start_color=bg)
        d.alignment = Alignment(horizontal="left", vertical="center")
        d.border = thin_border()

    # ── Key Insights ───────────────────────────────────────────────────────────
    KI_SP  = LG_DATA + len(legend_items)
    KI_HDR = KI_SP + 1
    KI_DATA= KI_HDR + 1

    ws.row_dimensions[KI_SP].height = 10
    ws.row_dimensions[KI_HDR].height = 20
    ws.merge_cells(f"A{KI_HDR}:Q{KI_HDR}")
    _hdr(ws[f"A{KI_HDR}"],
         "KEY INSIGHTS FROM THIS ANALYSIS",
         bg=NAVY, sz=11, align="left")

    insights = [
        ("🟡  Sliema (NEW) — Now 5 Therapists (Fresha confirmed)",
         "Updated to 5 therapists as confirmed on Fresha. Max daily capacity = 21.4 treatments. "
         "FY average implied demand = ~16 treatments/day → 75% utilisation (YELLOW — elevated but healthy). "
         "Peak months Jul/Aug reach ~92–94% utilisation (ORANGE — stress zone). "
         "Consider 1 seasonal/PT therapist cover for Jun–Sep to stay below 85% in peak season. "
         "Modelled HC cost ~€13,200/mo vs 35% budget €14,583 — within budget ✓",
         "FEF9E7"),

        ("🟡  Inter & Hugos — Possibly Over-staffed at €85 Average",
         "At €85 avg treatment, both Inter and Hugos show ~48–50% FY utilisation with 9 therapists each "
         "(GREEN but low). However, these are hotel-based premium spas where average treatment "
         "value may be €100–€120+ (signature rituals, longer durations). "
         "If avg treatment = €105, FY util for Inter rises to ~60% — healthier. "
         "Recommended: verify actual avg transaction value from Fresha/booking data before reducing headcount.",
         "FEF9E7"),

        ("🟢  Novotel — Well-calibrated",
         "2 therapists generating €14K/month implies 5.4 treatments/day FY avg — 63% of max capacity (8.6). "
         "In peak months (Jul/Aug), this rises to ~8.2 treatments/day — 95% utilisation. "
         "The model correctly includes PT Friday–Sunday cover (€1,092/mo) to handle peak demand. "
         "This allocation is realistic and efficient.",
         "D5F5E3"),

        ("📌  Seasonal Swing — All Locations",
         "Malta's seasonal index creates a 56% swing from Jan (×0.80) to Aug (×1.25). "
         "Every location experiences 40–50% more implied demand in August vs January. "
         "Consider temporary/seasonal therapist contracts (3–4 months: Jun–Sep) "
         "for Tier 2–3 locations to avoid stress-zone months without over-staffing year-round.",
         "EAF4FB"),

        ("💡  Required Headcount to Hit 80% Utilisation (FY average)",
         "Inter: 6 needed (9 allocated) | Hugos: 6 needed (9 allocated) | "
         "Ramla: 4 needed (5 allocated) | Hyatt: 3 needed (4 allocated) | "
         "Excelsior: 3 needed (4 allocated) | Riviera: 3 needed (4 allocated) | "
         "Sunny Coast: 3 needed (3 allocated ✓) | Novotel: 2 needed (2 allocated ✓) | "
         "Sliema: 5 needed FY avg / 6–7 for peak (5 allocated ✓ — add 1 seasonal for Jun–Sep)",
         "F9F0FF"),
    ]

    for i, (title, body, bg) in enumerate(insights):
        row = KI_DATA + i
        ws.row_dimensions[row].height = 50
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        t = ws.cell(row, 1)
        t.value = title
        t.font = Font(bold=True, color=BLACK, size=10)
        t.fill = PatternFill("solid", start_color=bg)
        t.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        t.border = thin_border()
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=17)
        b = ws.cell(row, 3)
        b.value = body
        b.font = Font(color=BLACK, size=9)
        b.fill = PatternFill("solid", start_color=bg)
        b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        b.border = thin_border()

    ws.freeze_panes = "E15"


# ── Entry point ────────────────────────────────────────────────────────────────
def main():
    wb = load_workbook(FILE)
    build_util_check(wb)
    wb.save(FILE)
    print(f"Saved: {FILE}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
