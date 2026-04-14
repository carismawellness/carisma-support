"""
rebuild_dynamic_staffing.py
Modifies: hr/Carisma Spa Staffing Model.xlsx  (in-place, keeps all existing sheets)

Adds:
  • "Dynamic Staffing Plan" sheet  — portfolio-level seasonal hiring calendar + savings table
  • Dynamic section appended to each of the 9 monthly location sheets, showing:
      - Demand-adjusted therapist count per month (optimal_t)
      - Core T (permanent) vs Seasonal T (fixed-term, Apr–Oct)
      - Avg Tx per day per therapist (NEW — revenue-based workload per staffed therapist)
      - Dynamic HC cost vs fixed model saving
      - Seasonal staff roster with colour-coded rows (teal = core, orange = seasonal)
      - T:R Ratio QC (NEW — reconciles therapist-to-reception cost as % of total HC spend)

Re-run safe: detects and clears existing dynamic sections in monthly sheets before re-writing.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math
import os

# =============================================================================
# PALETTE  (mirrors build_staffing_model.py exactly)
# =============================================================================
NAVY        = "1B3A4B"
GOLD        = "B8943E"
WHITE       = "FFFFFF"
BLACK       = "000000"
LIGHT_BLUE  = "D6E4F0"
TEAL        = "2C5F77"
DARK_GREEN  = "1B7A3B"

# Dynamic-layer specific colours
CORE_BG     = "D6EAF8"   # Soft blue  — core permanent staff rows
CORE_FNT    = "1B3A4B"   # Navy font
SEAS_BG     = "FAD7A0"   # Warm amber — seasonal contract staff rows
SEAS_FNT    = "A04000"   # Dark orange font
UNDER_BG    = "FADBD8"   # Soft red   — understaffing alert
UNDER_FNT   = "C0392B"   # Red font
OK_BG       = "D4EFDF"   # Green      — within budget
WARN_BG     = "FEF9E7"   # Pale yellow — slight over
SUMMARY_BG  = "EAF4FB"   # Light teal — summary rows
ACTION_BG   = "FCF3CF"   # Yellow     — action required

# QC section colours (matching add_hc_qc.py palette)
C_GREEN     = "D5F5E3"   # healthy range
C_YELLOW    = "FEF9E7"   # elevated
C_GREY      = "F2F3F4"   # N/A / not applicable

# =============================================================================
# MODEL ASSUMPTIONS  (matches existing model constants)
# =============================================================================
T_LOADED      = round(1500 * 1.118)    # 1677 €/mo per therapist (base + 11.8% SSC)
R_LOADED      = round(1400 * 1.118)    # 1565 €/mo per receptionist
RM_FULL       = round(2500 * 1.118)    # 2795 €/mo RM full FTE
RETAIL_PCT    = 0.055
TREAT_COMM    = 0.045
RETAIL_COMM   = 0.15
RM_REV_COMM   = 0.005

AVG_TX_REV    = 85
AVG_TX_MIN    = 75
SHIFT_HRS     = 8          # effective billable hours per therapist shift
UTIL_TARGET   = 0.80
WORK_DAYS_MO  = 21.7       # 5/7 × avg 30.4 calendar days
BUFFER        = 1.20       # 20% scheduling buffer above pure demand

TX_PER_T = (SHIFT_HRS * UTIL_TARGET * 60 / AVG_TX_MIN) * WORK_DAYS_MO  # ≈ 111.1

MONTHS       = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
SEASONAL_IDX = [0.80, 0.85, 0.90, 0.95, 1.00, 1.10, 1.20, 1.25, 1.10, 1.00, 0.90, 0.95]

# Months when seasonal contracts are active
SEASONAL_MONTHS = {"Apr","May","Jun","Jul","Aug","Sep","Oct"}


# =============================================================================
# LOCATION CONFIG
# =============================================================================
LOCATIONS = [
    {
        "sheet": "Inter Monthly", "name": "Intercontinental", "annual_rev": 593700,
        "floor_min": 5, "fixed_t": 9, "r": 2, "rm": 0.33,
        "core_names": ["T1 — Isabelle", "T2 — Margaux", "T3 — Chiara",
                       "T4 — Eleni", "T5 — Priya"],
        "seasonal_names": ["T6 — Sofia (seasonal)", "T7 — Anya (seasonal)",
                           "T8 — Leila (seasonal)"],
        "action": None,
    },
    {
        "sheet": "Hugos Monthly", "name": "Hugos", "annual_rev": 570552,
        "floor_min": 4, "fixed_t": 9, "r": 2, "rm": 0.33,
        "core_names": ["T1 — Lucia", "T2 — Amara", "T3 — Valentina", "T4 — Hana"],
        "seasonal_names": ["T5 — Renata (seasonal)", "T6 — Zara (seasonal)",
                           "T7 — Fatima (seasonal)", "T8 — Ingrid (seasonal)"],
        "action": None,
    },
    {
        "sheet": "Ramla Monthly", "name": "Ramla", "annual_rev": 358632,
        "floor_min": 3, "fixed_t": 5, "r": 2, "rm": 0.33,
        "core_names": ["T1 — Maria", "T2 — Sofia", "T3 — Elena"],
        "seasonal_names": ["T4 — Isabelle (seasonal)", "T5 — Natalia (seasonal)"],
        "action": None,
    },
    {
        "sheet": "Hyatt Monthly", "name": "Hyatt", "annual_rev": 304296,
        "floor_min": 2, "fixed_t": 4, "r": 1, "rm": 0.33,
        "core_names": ["T1 — Amara", "T2 — Lucia"],
        "seasonal_names": ["T3 — Priya (seasonal)", "T4 — Lena (seasonal)"],
        "action": None,
    },
    {
        "sheet": "Excelsior Monthly", "name": "Excelsior", "annual_rev": 291156,
        "floor_min": 2, "fixed_t": 4, "r": 1, "rm": 0.33,
        "core_names": ["T1 — Valentina", "T2 — Miriam"],
        "seasonal_names": ["T3 — Fatima (seasonal)", "T4 — Chiara (seasonal)"],
        "action": None,
    },
    {
        "sheet": "Riviera Monthly", "name": "Riviera", "annual_rev": 239232,
        "floor_min": 2, "fixed_t": 4, "r": 1, "rm": 0.33,
        "core_names": ["T1 — Maya", "T2 — Sofia"],
        "seasonal_names": ["T3 — Elena (seasonal)"],
        "action": (
            "ACTION: Reduce permanent headcount from 4T → 2T. "
            "Hire 1 seasonal therapist for Apr–Oct. "
            "Annual saving vs current model: €21,801."
        ),
    },
    {
        "sheet": "Sunny Coast Monthly", "name": "Sunny Coast", "annual_rev": 218448,
        "floor_min": 2, "fixed_t": 3, "r": 1, "rm": 0.33,
        "core_names": ["T1 — Natasha", "T2 — Priya"],
        "seasonal_names": ["T3 — Irene (seasonal)"],
        "action": None,
    },
    {
        "sheet": "Novotel Monthly", "name": "Novotel", "annual_rev": 168000,
        "floor_min": 1, "fixed_t": 2, "r": 0, "rm": 0.33,
        "core_names": ["T1 — Chiara"],
        "seasonal_names": ["T2 — Bianca (seasonal)", "T3 — (new hire, seasonal)"],
        "action": None,
    },
    {
        "sheet": "Sliema Monthly", "name": "Sliema (NEW)", "annual_rev": 500000,
        "floor_min": 5, "fixed_t": 2, "r": 1, "rm": 0.33,
        "core_names": ["T1 — Amara", "T2 — Rania",
                       "T3 — (HIRE NOW)", "T4 — (HIRE NOW)", "T5 — (HIRE NOW)"],
        "seasonal_names": ["T6 — (seasonal, from Apr)", "T7 — (seasonal, from Jun)"],
        "action": (
            "URGENT — Hire 3 permanent therapists immediately. "
            "Current 2T cannot service projected €500k/yr revenue. "
            "Estimated lost revenue from capacity shortfall: €50,000–€70,000/yr."
        ),
    },
]


# =============================================================================
# HELPERS
# =============================================================================
def _thin():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_bottom():
    thin = Side(style="thin", color="AAAAAA")
    thick = Side(style="medium", color=NAVY)
    return Border(left=thin, right=thin, top=thin, bottom=thick)


def _hdr(cell, text, bg=NAVY, fg=WHITE, bold=True, sz=10, wrap=False, align="center"):
    cell.value = text
    cell.font = Font(bold=bold, color=fg, size=sz)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin()


def _lbl(cell, text, bold=False, align="left", color=BLACK, sz=10, bg=WHITE, wrap=False):
    cell.value = text
    cell.font = Font(bold=bold, color=color, size=sz)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin()


def _num(cell, value, fmt="#,##0", bold=False, color=BLACK, bg=WHITE):
    cell.value = value
    cell.font = Font(bold=bold, color=color, size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = fmt
    cell.border = _thin()


def _blank(cell, bg=WHITE):
    cell.value = None
    cell.fill = PatternFill("solid", start_color=bg)
    cell.border = _thin()


def optimal_t(revenue, floor_min):
    """Demand-implied therapist count with 20% scheduling buffer, floored at minimum."""
    tx_needed = revenue * (1 - RETAIL_PCT) / AVG_TX_REV
    t_raw = tx_needed / TX_PER_T
    return max(floor_min, math.ceil(t_raw * BUFFER))


def hc_cost_breakdown(revenue, t, r, rm_frac=0.33):
    """Returns (fixed_base, variable_commissions, total)."""
    fixed = t * T_LOADED + r * R_LOADED + round(rm_frac * RM_FULL)
    var = (round(revenue * (1 - RETAIL_PCT) * TREAT_COMM) +
           round(revenue * RETAIL_PCT * RETAIL_COMM) +
           (round(revenue * RM_REV_COMM) if rm_frac > 0 else 0))
    return fixed, var, fixed + var


def monthly_revenues(annual_rev):
    return [round(annual_rev * idx / 12) for idx in SEASONAL_IDX]


# =============================================================================
# IDEMPOTENCY HELPERS FOR MONTHLY SHEETS
# =============================================================================
def find_dynamic_section_start(ws):
    """Return the row where the DYNAMIC STAFFING MODEL section starts, or None."""
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if val and isinstance(val, str) and "DYNAMIC STAFFING MODEL" in val:
            return row
    return None


def clear_from_row(ws, start_row):
    """Clear all content, formatting, and merged cells from start_row to max_row."""
    # Unmerge any merged cells in this range first
    ranges_to_unmerge = [
        str(mr) for mr in ws.merged_cells.ranges
        if mr.min_row >= start_row
    ]
    for r in ranges_to_unmerge:
        ws.unmerge_cells(r)
    # Clear content and formatting
    for row_num in range(start_row, ws.max_row + 1):
        for col in range(1, 20):
            cell = ws.cell(row_num, col)
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = "General"


# =============================================================================
# SHEET 1 (new): DYNAMIC STAFFING PLAN
# =============================================================================
def build_dynamic_summary(wb):
    """Insert 'Dynamic Staffing Plan' as the second sheet in the workbook."""
    ws = wb.create_sheet("Dynamic Staffing Plan")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, 14):
        ws.column_dimensions[get_column_letter(col_idx)].width = 9
    ws.column_dimensions["N"].width = 14
    ws.column_dimensions["O"].width = 30

    # Row 1: Main title
    ws.merge_cells("A1:O1")
    _hdr(ws["A1"], "CARISMA SPA — DYNAMIC STAFFING PLAN  |  Demand-Adjusted Seasonal Model",
         bg=NAVY, sz=13)
    ws.row_dimensions[1].height = 30

    # Row 2: Subtitle
    ws.merge_cells("A2:O2")
    total_saving = sum(
        sum(
            hc_cost_breakdown(rev, loc["fixed_t"], loc["r"], loc["rm"])[2] -
            hc_cost_breakdown(rev, optimal_t(rev, loc["floor_min"]), loc["r"], loc["rm"])[2]
            for rev in monthly_revenues(loc["annual_rev"])
        )
        for loc in LOCATIONS
    )
    _hdr(ws["A2"],
         f"Portfolio annual saving vs fixed model: €{total_saving:,.0f}  |  "
         f"Core = Permanent contract year-round  |  Seasonal = Fixed-term Apr–Oct  |  "
         f"Optimal T = demand × 1.20 buffer, floored at site minimum",
         bg=GOLD, sz=9, bold=False)
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8

    # Row 4: Legend
    ws.merge_cells("A4:C4")
    _hdr(ws["A4"], "LEGEND", bg=NAVY, sz=9)
    ws.merge_cells("D4:F4")
    _hdr(ws["D4"], "Core only",         bg=CORE_BG,  fg=CORE_FNT,  sz=9)
    ws.merge_cells("G4:I4")
    _hdr(ws["G4"], "Seasonal active",   bg=SEAS_BG,  fg=SEAS_FNT,  sz=9)
    ws.merge_cells("J4:L4")
    _hdr(ws["J4"], "Understaffed ⚠",    bg=UNDER_BG, fg=UNDER_FNT, sz=9)
    for c in range(13, 16):
        _blank(ws.cell(4, c), bg=WHITE)
    ws.row_dimensions[4].height = 18

    ws.row_dimensions[5].height = 8

    # Row 6: Heatmap header
    ws.row_dimensions[6].height = 26
    _hdr(ws.cell(6, 1), "Location",       bg=NAVY, sz=10)
    for m_i, m in enumerate(MONTHS):
        _hdr(ws.cell(6, 2 + m_i), m,      bg=NAVY, sz=9)
    _hdr(ws.cell(6, 14), "Annual Saving", bg=NAVY, sz=9)
    _hdr(ws.cell(6, 15), "Action Required", bg=NAVY, sz=9)

    # Rows 7–15: Heatmap data
    for loc_i, loc in enumerate(LOCATIONS):
        row = 7 + loc_i
        ws.row_dimensions[row].height = 22
        revs = monthly_revenues(loc["annual_rev"])

        _lbl(ws.cell(row, 1), loc["name"], bold=True, bg=CORE_BG if not loc["action"] else ACTION_BG,
             color=UNDER_FNT if loc["action"] and "URGENT" in loc["action"] else CORE_FNT)

        annual_saving = 0
        for m_i, (month, rev) in enumerate(zip(MONTHS, revs)):
            t_opt = optimal_t(rev, loc["floor_min"])
            core_t = loc["floor_min"]
            seas_t = max(0, t_opt - core_t)
            fixed_total = hc_cost_breakdown(rev, loc["fixed_t"], loc["r"], loc["rm"])[2]
            dyn_total   = hc_cost_breakdown(rev, t_opt,           loc["r"], loc["rm"])[2]
            saving_mo   = fixed_total - dyn_total
            annual_saving += saving_mo

            cell = ws.cell(row, 2 + m_i)

            if loc["floor_min"] > loc["fixed_t"]:
                bg  = UNDER_BG
                fg  = UNDER_FNT
                txt = f"⚠ {t_opt}T"
            elif seas_t > 0 and month in SEASONAL_MONTHS:
                bg  = SEAS_BG
                fg  = SEAS_FNT
                txt = f"{core_t}+{seas_t}S"
            else:
                bg  = CORE_BG
                fg  = CORE_FNT
                txt = f"{core_t}T"

            cell.value = txt
            cell.font  = Font(bold=True, color=fg, size=9)
            cell.fill  = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _thin()

        saving_cell = ws.cell(row, 14)
        is_negative = annual_saving < 0
        _num(saving_cell, annual_saving, fmt='€#,##0',
             bold=True,
             color=UNDER_FNT if is_negative else DARK_GREEN,
             bg=UNDER_BG if is_negative else OK_BG)

        action_cell = ws.cell(row, 15)
        if loc["action"]:
            _lbl(action_cell, loc["action"],
                 color=UNDER_FNT if "URGENT" in loc["action"] else SEAS_FNT,
                 bg=UNDER_BG if "URGENT" in loc["action"] else ACTION_BG,
                 sz=8, wrap=True)
            ws.row_dimensions[row].height = 36
        else:
            _blank(action_cell)

    # Totals row
    totals_row = 7 + len(LOCATIONS)
    ws.row_dimensions[totals_row].height = 22
    ws.merge_cells(f"A{totals_row}:M{totals_row}")
    _hdr(ws.cell(totals_row, 1), "PORTFOLIO TOTAL ANNUAL SAVING", bg=NAVY, sz=10, align="right")
    _num(ws.cell(totals_row, 14), total_saving,
         fmt='€#,##0', bold=True,
         color=DARK_GREEN if total_saving > 0 else UNDER_FNT,
         bg=OK_BG if total_saving > 0 else UNDER_BG)
    _blank(ws.cell(totals_row, 15))

    spacer = totals_row + 1
    ws.row_dimensions[spacer].height = 8

    # Seasonal contract calendar explainer
    explainer_row = spacer + 1
    ws.merge_cells(f"A{explainer_row}:O{explainer_row}")
    _hdr(ws.cell(explainer_row, 1),
         "SEASONAL CONTRACT CALENDAR  —  Fixed-term contracts typically run Apr 1 → Oct 31.  "
         "Recruitment to begin by Feb 15 each year.  Review and renew by Oct 1.",
         bg=GOLD, sz=9, bold=False, align="left")
    ws.row_dimensions[explainer_row].height = 20

    # Recruitment pipeline table
    pipe_row = explainer_row + 2
    ws.row_dimensions[pipe_row].height = 20
    for ci, hd in enumerate(["Location", "Core T (perm)", "Seasonal T needed",
                              "Active months", "Recruit by", "Release date", "Note"]):
        c = ws.cell(pipe_row, 1 + ci)
        if ci < 7:
            _hdr(c, hd, bg=TEAL, sz=9)
    ws.merge_cells(f"H{pipe_row}:O{pipe_row}")
    _blank(ws.cell(pipe_row, 8))

    for loc_i, loc in enumerate(LOCATIONS):
        pr = pipe_row + 1 + loc_i
        ws.row_dimensions[pr].height = 18
        revs = monthly_revenues(loc["annual_rev"])
        max_seasonal = max(max(0, optimal_t(rev, loc["floor_min"]) - loc["floor_min"]) for rev in revs)
        row_bg = UNDER_BG if loc["action"] and "URGENT" in loc["action"] else SUMMARY_BG

        _lbl(ws.cell(pr, 1), loc["name"],       bold=True, bg=row_bg, color=CORE_FNT)
        _num(ws.cell(pr, 2), loc["floor_min"],  bg=row_bg, color=CORE_FNT)
        _num(ws.cell(pr, 3), max_seasonal,       bg=SEAS_BG if max_seasonal > 0 else row_bg,
             color=SEAS_FNT if max_seasonal > 0 else CORE_FNT)
        _lbl(ws.cell(pr, 4), "Apr–Oct",          bg=row_bg, color=SEAS_FNT if max_seasonal > 0 else BLACK)
        _lbl(ws.cell(pr, 5), "Feb 15",           bg=row_bg, align="center")
        _lbl(ws.cell(pr, 6), "Oct 31",           bg=row_bg, align="center")
        _lbl(ws.cell(pr, 7),
             loc["action"] if loc["action"] else "—",
             bg=UNDER_BG if loc["action"] and "URGENT" in str(loc["action"]) else row_bg,
             color=UNDER_FNT if loc["action"] and "URGENT" in str(loc["action"]) else BLACK,
             sz=8, wrap=True)
        ws.merge_cells(f"H{pr}:O{pr}")
        _blank(ws.cell(pr, 8))


# =============================================================================
# MONTHLY SHEET APPEND
# =============================================================================
def append_dynamic_section(ws, loc):
    """
    Append the dynamic staffing section to a monthly location sheet.
    Idempotent: detects and clears any existing DYNAMIC STAFFING MODEL section first.

    Writes:
      1. Section header
      2. Column headers (12 cols, including new Avg Tx/Day/T)
      3. 12 monthly data rows
      4. Annual totals row
      5. Seasonal staff roster (teal = core, orange = seasonal)
      6. T:R Ratio QC — reconciles therapist-to-reception cost as % of total HC spend
    """
    # ── Idempotency: clear existing dynamic section if present ──
    dyn_start = find_dynamic_section_start(ws)
    if dyn_start is not None:
        clear_start = max(1, dyn_start - 4)
        clear_from_row(ws, clear_start)
        last_row = clear_start - 1
    else:
        last_row = ws.max_row
        while last_row > 1 and all(
            ws.cell(last_row, c).value is None
            for c in range(1, 13)
        ):
            last_row -= 1

    r = last_row + 4  # 3-row gap + 1

    # ── Column widths ──
    # 12 columns: Month | Revenue | 35% Budget | Optimal T | Core T | Seasonal T |
    #             Avg Tx/Day/T | Dynamic HC | Saving vs Fixed | Status | # Core | # Seas
    col_widths_needed = {
        1: 9, 2: 14, 3: 14, 4: 12, 5: 12, 6: 12,
        7: 12,   # Avg Tx/Day/T
        8: 16, 9: 16, 10: 14, 11: 9, 12: 9
    }
    for ci, w in col_widths_needed.items():
        col_letter = get_column_letter(ci)
        if ws.column_dimensions[col_letter].width < w:
            ws.column_dimensions[col_letter].width = w

    # ── Section header (merged A:L — 12 cols) ──
    ws.merge_cells(f"A{r}:L{r}")
    _hdr(ws.cell(r, 1),
         f"DYNAMIC STAFFING MODEL  —  Demand-Adjusted Headcount  |  "
         f"Core = Permanent (year-round)  |  Seasonal = Fixed-term (Apr–Oct)  |  "
         f"Buffer: {int(BUFFER*100-100)}% above demand floor",
         bg=GOLD, sz=9, bold=True, align="left")
    ws.row_dimensions[r].height = 22
    r += 1

    # ── Column headers ──
    ws.row_dimensions[r].height = 30
    col_hdrs = [
        "Month", "Revenue (€)", "35% Budget (€)", "Optimal T",
        "Core T", "Seasonal T", "Avg Tx /\nDay / T",
        "Dynamic HC (€)", "Saving vs\nFixed (€)",
        "Status", "# Core", "# Seas"
    ]
    for ci, txt in enumerate(col_hdrs):
        _hdr(ws.cell(r, 1 + ci), txt, bg=TEAL, sz=9, wrap=True)
    r += 1

    # ── Monthly data rows ──
    revs = monthly_revenues(loc["annual_rev"])
    ann = {"rev": 0, "budget": 0, "dyn_cost": 0, "saving": 0}

    for m_i, (month, rev) in enumerate(zip(MONTHS, revs)):
        ws.row_dimensions[r].height = 18
        t_opt  = optimal_t(rev, loc["floor_min"])
        core_t = loc["floor_min"]
        seas_t = max(0, t_opt - core_t)

        budget_35  = round(rev * 0.35)
        _, _, fixed_hc = hc_cost_breakdown(rev, loc["fixed_t"], loc["r"], loc["rm"])
        _, _, dyn_hc   = hc_cost_breakdown(rev, t_opt, loc["r"], loc["rm"])
        saving     = fixed_hc - dyn_hc

        # Avg treatments per day per therapist
        # Formula: (treatment revenue) / avg_tx_price / working_days / staffed_therapists
        avg_tx_per_t = round(
            (rev * (1 - RETAIL_PCT) / AVG_TX_REV) / WORK_DAYS_MO / t_opt, 1
        ) if t_opt > 0 else 0.0

        # Row background
        is_under = loc["floor_min"] > loc["fixed_t"]
        if is_under:
            row_bg     = UNDER_BG
            status_txt = "⚠ UNDERSTAFFED"
            status_bg  = UNDER_BG
            status_fg  = UNDER_FNT
        elif seas_t > 0 and month in SEASONAL_MONTHS:
            row_bg     = SEAS_BG
            status_txt = "▲ SEASONAL ACTIVE"
            status_bg  = SEAS_BG
            status_fg  = SEAS_FNT
        else:
            row_bg     = CORE_BG
            status_txt = "● ON CORE"
            status_bg  = CORE_BG
            status_fg  = CORE_FNT

        # Avg Tx/Day/T colour: orange if > max practical (6.0), red if impossible (> 8.0)
        tx_bg = UNDER_BG if avg_tx_per_t > 6.0 else OK_BG if avg_tx_per_t >= 4.0 else row_bg
        tx_fg = UNDER_FNT if avg_tx_per_t > 6.0 else DARK_GREEN if avg_tx_per_t >= 4.0 else CORE_FNT

        _lbl(ws.cell(r, 1),  month,       bold=True, bg=row_bg, color=CORE_FNT)
        _num(ws.cell(r, 2),  rev,          bg=row_bg)
        _num(ws.cell(r, 3),  budget_35,    bg=row_bg)
        _num(ws.cell(r, 4),  t_opt,        bg=row_bg, bold=True,
             color=UNDER_FNT if is_under else CORE_FNT)
        _num(ws.cell(r, 5),  core_t,       bg=CORE_BG, color=CORE_FNT, bold=True)
        _num(ws.cell(r, 6),  seas_t,
             bg=SEAS_BG if seas_t > 0 else CORE_BG,
             color=SEAS_FNT if seas_t > 0 else CORE_FNT,
             bold=seas_t > 0)
        # col 7: Avg Tx/Day/T  (NEW)
        _num(ws.cell(r, 7),  avg_tx_per_t, fmt="0.0", bg=tx_bg, color=tx_fg, bold=True)
        _num(ws.cell(r, 8),  dyn_hc,       bg=row_bg)
        _num(ws.cell(r, 9),  saving,
             bg=OK_BG if saving >= 0 else UNDER_BG,
             bold=True,
             color=DARK_GREEN if saving > 0 else UNDER_FNT if saving < 0 else BLACK)

        c_status = ws.cell(r, 10)
        _lbl(c_status, status_txt, bold=True, align="center",
             bg=status_bg, color=status_fg, sz=9)

        _num(ws.cell(r, 11), core_t,  bg=CORE_BG, color=CORE_FNT)
        _num(ws.cell(r, 12), seas_t,
             bg=SEAS_BG if seas_t > 0 else CORE_BG,
             color=SEAS_FNT if seas_t > 0 else CORE_FNT)

        ann["rev"]      += rev
        ann["budget"]   += budget_35
        ann["dyn_cost"] += dyn_hc
        ann["saving"]   += saving
        r += 1

    # ── Annual totals row ──
    ws.row_dimensions[r].height = 22
    _lbl(ws.cell(r, 1), "ANNUAL TOTAL", bold=True, bg=NAVY, color=WHITE, align="center")
    _num(ws.cell(r, 2), ann["rev"],    bg=NAVY, bold=True, color=WHITE)
    _num(ws.cell(r, 3), ann["budget"], bg=NAVY, bold=True, color=WHITE)
    for ci in (4, 5, 6, 7):   # blank: Optimal T, Core T, Seasonal T, Avg Tx/Day/T
        _blank(ws.cell(r, ci))
        ws.cell(r, ci).fill = PatternFill("solid", start_color=NAVY)
        ws.cell(r, ci).border = _thin()
    _num(ws.cell(r, 8),  ann["dyn_cost"], bg=NAVY, bold=True, color=WHITE)
    _num(ws.cell(r, 9),  ann["saving"],
         bg=OK_BG if ann["saving"] >= 0 else UNDER_BG,
         bold=True,
         color=DARK_GREEN if ann["saving"] > 0 else UNDER_FNT)
    for ci in (10, 11, 12):
        _blank(ws.cell(r, ci))
        ws.cell(r, ci).fill = PatternFill("solid", start_color=NAVY)
        ws.cell(r, ci).border = _thin()
    r += 2

    # ── Action note (if any) ──
    if loc["action"]:
        ws.merge_cells(f"A{r}:L{r}")
        _hdr(ws.cell(r, 1),
             f"⚠  ACTION REQUIRED: {loc['action']}",
             bg=UNDER_BG, fg=UNDER_FNT, sz=9, bold=True, align="left")
        ws.row_dimensions[r].height = 30
        r += 2

    # ── Seasonal staff roster ──
    ws.merge_cells(f"A{r}:L{r}")
    _hdr(ws.cell(r, 1),
         "SEASONAL STAFF ROSTER  —  Core = permanent contract  |  Seasonal = fixed-term Apr–Oct",
         bg=TEAL, sz=9, bold=True, align="left")
    ws.row_dimensions[r].height = 20
    r += 1

    ws.row_dimensions[r].height = 20
    for ci, hd in enumerate(["Name", "Contract Type", "Active", "Notes",
                              "", "", "", "", "", "", "", ""]):
        if ci < 4:
            _hdr(ws.cell(r, 1 + ci), hd, bg=TEAL, sz=9)
        else:
            _blank(ws.cell(r, 1 + ci))
    r += 1

    for name in loc["core_names"]:
        ws.row_dimensions[r].height = 18
        _lbl(ws.cell(r, 1), name,             bold=True, bg=CORE_BG, color=CORE_FNT)
        _lbl(ws.cell(r, 2), "Permanent",      bg=CORE_BG, color=CORE_FNT)
        _lbl(ws.cell(r, 3), "Year-round",     bg=CORE_BG, color=CORE_FNT)
        _lbl(ws.cell(r, 4), "Core team",      bg=CORE_BG, color=CORE_FNT, sz=9)
        for ci in range(5, 13):
            _blank(ws.cell(r, ci), bg=CORE_BG)
        r += 1

    for name in loc["seasonal_names"]:
        ws.row_dimensions[r].height = 18
        _lbl(ws.cell(r, 1), name,                  bold=True, bg=SEAS_BG, color=SEAS_FNT)
        _lbl(ws.cell(r, 2), "Fixed-term seasonal", bg=SEAS_BG, color=SEAS_FNT)
        _lbl(ws.cell(r, 3), "Apr 1 – Oct 31",      bg=SEAS_BG, color=SEAS_FNT)
        _lbl(ws.cell(r, 4), "Recruit by Feb 15",   bg=SEAS_BG, color=SEAS_FNT, sz=9)
        for ci in range(5, 13):
            _blank(ws.cell(r, ci), bg=SEAS_BG)
        r += 1

    # ==========================================================================
    # T:R RATIO QC — HC SPEND RECONCILIATION  (NEW)
    # ==========================================================================
    r += 1  # spacer

    ws.merge_cells(f"A{r}:L{r}")
    _hdr(ws.cell(r, 1),
         "T:R RATIO QC  —  THERAPIST · RECEPTION HC SPEND RECONCILIATION  "
         "|  Validates that cost allocation by role is within industry benchmarks",
         bg=NAVY, sz=10, bold=True, align="left")
    ws.row_dimensions[r].height = 22
    r += 1

    # QC column headers
    ws.row_dimensions[r].height = 30
    qc_col_hdrs = [
        "Role / Category", "Avg Annual\nFTEs", "Annual Cost (€)",
        "% of Total\nDynamic HC", "Industry\nBenchmark", "QC Status"
    ]
    for ci, hd in enumerate(qc_col_hdrs):
        _hdr(ws.cell(r, 1 + ci), hd, bg=TEAL, sz=9, wrap=True)
    for ci in range(7, 13):
        _blank(ws.cell(r, ci), bg=C_GREY)
    r += 1

    # ── Compute annual HC cost breakdown for QC ──
    annual_t_cost     = 0
    annual_t_fte_sum  = 0
    annual_r_cost     = loc["r"] * R_LOADED * 12
    annual_rm_cost    = round(loc["rm"] * RM_FULL) * 12
    annual_var        = 0

    for rev in revs:
        t_o = optimal_t(rev, loc["floor_min"])
        annual_t_cost    += t_o * T_LOADED
        annual_t_fte_sum += t_o
        _, var, _         = hc_cost_breakdown(rev, t_o, loc["r"], loc["rm"])
        annual_var        += var

    avg_t_fte      = round(annual_t_fte_sum / 12, 1)
    annual_total   = annual_t_cost + annual_r_cost + annual_rm_cost + annual_var

    t_pct   = annual_t_cost  / annual_total if annual_total else 0
    r_pct   = annual_r_cost  / annual_total if annual_total else 0
    rm_pct  = annual_rm_cost / annual_total if annual_total else 0
    var_pct = annual_var     / annual_total if annual_total else 0

    no_reception  = loc["r"] == 0
    t_r_count_val = avg_t_fte / loc["r"] if loc["r"] > 0 else None
    t_r_spend_val = annual_t_cost / annual_r_cost if annual_r_cost > 0 else None

    def _in_range(val, lo, hi):
        return val is not None and lo <= val <= hi

    def _qc_status_cell(cell, label, in_range_flag, is_na=False):
        if is_na:
            _lbl(cell, label, bg=C_GREY, color="888888", sz=9, align="center")
        elif in_range_flag:
            _lbl(cell, f"✓  {label}", bg=OK_BG, color=DARK_GREEN, bold=True, sz=9)
        else:
            _lbl(cell, f"⚠  {label}", bg=UNDER_BG, color=UNDER_FNT, bold=True, sz=9)

    def _fill_blanks(row_n, bg_col):
        for ci in range(7, 13):
            _blank(ws.cell(row_n, ci), bg=bg_col)

    # ── Therapist row ──
    ws.row_dimensions[r].height = 18
    t_in = _in_range(t_pct, 0.60, 0.85)
    _lbl(ws.cell(r, 1), "Therapists (T)", bold=True, bg=CORE_BG, color=CORE_FNT)
    _lbl(ws.cell(r, 2), f"{avg_t_fte:.1f}", bg=CORE_BG, color=CORE_FNT, align="center", bold=True)
    _num(ws.cell(r, 3), annual_t_cost, bg=CORE_BG)
    _num(ws.cell(r, 4), t_pct, fmt="0.0%",
         bg=OK_BG if t_in else UNDER_BG,
         bold=True, color=DARK_GREEN if t_in else UNDER_FNT)
    _lbl(ws.cell(r, 5), "60% – 85%", bg=CORE_BG, sz=9, align="center")
    _qc_status_cell(ws.cell(r, 6), f"T = {t_pct:.1%} of total HC", t_in)
    _fill_blanks(r, CORE_BG)
    r += 1

    # ── Reception row ──
    ws.row_dimensions[r].height = 18
    r_bg_qc = "E9F7EF"
    if no_reception:
        _lbl(ws.cell(r, 1), "Reception (R)", bold=True, bg=C_GREY, color="888888")
        _lbl(ws.cell(r, 2), "0", bg=C_GREY, color="888888", align="center")
        _num(ws.cell(r, 3), 0, bg=C_GREY, color="888888")
        _lbl(ws.cell(r, 4), "N/A", bg=C_GREY, color="888888", align="center", sz=9)
        _lbl(ws.cell(r, 5), "7% – 20%", bg=C_GREY, color="888888", sz=9, align="center")
        _qc_status_cell(ws.cell(r, 6), "Therapy-only model (by design)", True, is_na=True)
        _fill_blanks(r, C_GREY)
    else:
        r_in = _in_range(r_pct, 0.07, 0.20)
        _lbl(ws.cell(r, 1), "Reception (R)", bold=True, bg=r_bg_qc, color=DARK_GREEN)
        _lbl(ws.cell(r, 2), str(loc["r"]), bg=r_bg_qc, color=DARK_GREEN, align="center", bold=True)
        _num(ws.cell(r, 3), annual_r_cost, bg=r_bg_qc)
        _num(ws.cell(r, 4), r_pct, fmt="0.0%",
             bg=OK_BG if r_in else UNDER_BG,
             bold=True, color=DARK_GREEN if r_in else UNDER_FNT)
        _lbl(ws.cell(r, 5), "7% – 20%", bg=r_bg_qc, sz=9, align="center")
        _qc_status_cell(ws.cell(r, 6), f"R = {r_pct:.1%} of total HC", r_in)
        _fill_blanks(r, r_bg_qc)
    r += 1

    # ── RM row ──
    ws.row_dimensions[r].height = 18
    rm_bg_qc = "F5EEF8"
    rm_in = _in_range(rm_pct, 0.03, 0.10)
    _lbl(ws.cell(r, 1), "Relationship Manager (RM)", bold=True, bg=rm_bg_qc, color="4A235A")
    _lbl(ws.cell(r, 2), f"{loc['rm']:.0%} FTE", bg=rm_bg_qc, color="4A235A", align="center", sz=9)
    _num(ws.cell(r, 3), annual_rm_cost, bg=rm_bg_qc)
    _num(ws.cell(r, 4), rm_pct, fmt="0.0%",
         bg=OK_BG if rm_in else UNDER_BG,
         bold=True, color=DARK_GREEN if rm_in else UNDER_FNT)
    _lbl(ws.cell(r, 5), "3% – 10%", bg=rm_bg_qc, sz=9, align="center")
    _qc_status_cell(ws.cell(r, 6), f"RM = {rm_pct:.1%} of total HC", rm_in)
    _fill_blanks(r, rm_bg_qc)
    r += 1

    # ── Variable commissions row ──
    ws.row_dimensions[r].height = 18
    _lbl(ws.cell(r, 1), "Variable Commissions", bg=WARN_BG, color="888888", sz=9)
    _lbl(ws.cell(r, 2), "—", bg=WARN_BG, color="888888", align="center", sz=9)
    _num(ws.cell(r, 3), annual_var, bg=WARN_BG, color="888888")
    _num(ws.cell(r, 4), var_pct, fmt="0.0%", bg=WARN_BG, color="888888")
    _lbl(ws.cell(r, 5), "—", bg=WARN_BG, color="888888", align="center", sz=9)
    _lbl(ws.cell(r, 6), "Treatment + retail performance bonuses", bg=WARN_BG, color="888888", sz=9)
    _fill_blanks(r, WARN_BG)
    r += 1

    # ── TOTAL HC row ──
    ws.row_dimensions[r].height = 22
    _lbl(ws.cell(r, 1), "TOTAL DYNAMIC HC", bold=True, bg=NAVY, color=WHITE)
    _lbl(ws.cell(r, 2), "—", bg=NAVY, color=WHITE, align="center")
    _num(ws.cell(r, 3), annual_total, bg=NAVY, bold=True, color=WHITE)
    _lbl(ws.cell(r, 4), "100.0%", bg=NAVY, color=WHITE, bold=True, align="center")
    for ci in range(5, 13):
        _blank(ws.cell(r, ci))
        ws.cell(r, ci).fill = PatternFill("solid", start_color=NAVY)
        ws.cell(r, ci).border = _thin()
    r += 1

    # ── T:R Headcount Ratio row ──
    ws.row_dimensions[r].height = 20
    tr_count_ok = _in_range(t_r_count_val, 3.0, 7.0)
    _lbl(ws.cell(r, 1), "T:R Headcount Ratio", bold=True,
         bg=LIGHT_BLUE if not no_reception else C_GREY, color=CORE_FNT)
    if no_reception:
        _lbl(ws.cell(r, 2), "N/A", bg=C_GREY, color="888888", align="center", sz=9)
        _lbl(ws.cell(r, 3), "No reception staff at this location", bg=C_GREY, color="888888", sz=9)
        _lbl(ws.cell(r, 4), "—", bg=C_GREY, color="888888", align="center", sz=9)
        _lbl(ws.cell(r, 5), "3:1 – 7:1", bg=C_GREY, color="888888", sz=9, align="center")
        _qc_status_cell(ws.cell(r, 6), "N/A — therapy-only model", True, is_na=True)
    else:
        _lbl(ws.cell(r, 2), f"{t_r_count_val:.1f} : 1", bg=LIGHT_BLUE, color=CORE_FNT,
             align="center", bold=True)
        _lbl(ws.cell(r, 3), f"Avg {avg_t_fte:.1f} therapists : {loc['r']} receptionist(s)",
             bg=LIGHT_BLUE, color=CORE_FNT, sz=9)
        _num(ws.cell(r, 4), t_r_count_val, fmt="0.0",
             bg=OK_BG if tr_count_ok else UNDER_BG,
             bold=True, color=DARK_GREEN if tr_count_ok else UNDER_FNT)
        _lbl(ws.cell(r, 5), "3:1 – 7:1", bg=LIGHT_BLUE, sz=9, align="center")
        _qc_status_cell(ws.cell(r, 6),
                        f"Ratio {t_r_count_val:.1f}:1 — {'healthy' if tr_count_ok else 'review headcount split'}",
                        tr_count_ok)
    _fill_blanks(r, C_GREY)
    r += 1

    # ── T:R Spend Ratio row ──
    ws.row_dimensions[r].height = 20
    tr_spend_ok = _in_range(t_r_spend_val, 3.0, 8.0)
    _lbl(ws.cell(r, 1), "T:R Spend Ratio (€)", bold=True,
         bg=LIGHT_BLUE if not no_reception else C_GREY, color=CORE_FNT)
    if no_reception:
        _lbl(ws.cell(r, 2), "N/A", bg=C_GREY, color="888888", align="center", sz=9)
        _lbl(ws.cell(r, 3), "No reception spend — RM covers front desk", bg=C_GREY, color="888888", sz=9)
        _lbl(ws.cell(r, 4), "—", bg=C_GREY, color="888888", align="center", sz=9)
        _lbl(ws.cell(r, 5), "3× – 8×", bg=C_GREY, color="888888", sz=9, align="center")
        _qc_status_cell(ws.cell(r, 6), "N/A — therapy-only model", True, is_na=True)
    else:
        _lbl(ws.cell(r, 2), f"{t_r_spend_val:.1f}×", bg=LIGHT_BLUE, color=CORE_FNT,
             align="center", bold=True)
        _lbl(ws.cell(r, 3), f"For every €1 in reception cost, €{t_r_spend_val:.1f} spent on therapists",
             bg=LIGHT_BLUE, color=CORE_FNT, sz=9)
        _num(ws.cell(r, 4), t_r_spend_val, fmt="0.0",
             bg=OK_BG if tr_spend_ok else UNDER_BG,
             bold=True, color=DARK_GREEN if tr_spend_ok else UNDER_FNT)
        _lbl(ws.cell(r, 5), "3× – 8×", bg=LIGHT_BLUE, sz=9, align="center")
        _qc_status_cell(ws.cell(r, 6),
                        f"Spend ratio {t_r_spend_val:.1f}× — {'healthy' if tr_spend_ok else 'review cost balance'}",
                        tr_spend_ok)
    _fill_blanks(r, C_GREY)
    r += 1

    # ── Benchmark footnote ──
    ws.row_dimensions[r].height = 16
    ws.merge_cells(f"A{r}:L{r}")
    fn = ws.cell(r, 1)
    fn.value = (
        "BENCHMARKS  |  T% of HC: 60–85% ✓  |  R% of HC: 7–20% ✓  |  RM% of HC: 3–10% ✓  |  "
        "T:R Headcount: 3:1–7:1 ✓  |  T:R Spend: 3×–8× ✓  |  "
        "Source: ISPA Industry Benchmarks + Carisma operational data"
    )
    fn.font = Font(color="666666", italic=True, size=8)
    fn.fill = PatternFill("solid", start_color="F8F9FA")
    fn.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


# =============================================================================
# MAIN
# =============================================================================
def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    repo_root  = os.path.dirname(script_dir)
    path = os.path.join(repo_root, "hr", "Carisma Spa Staffing Model.xlsx")

    if not os.path.exists(path):
        print(f"ERROR: File not found: {path}")
        return

    print(f"Loading: {path}")
    wb = load_workbook(path)
    print(f"  Existing sheets: {wb.sheetnames}")

    # Remove old dynamic summary sheet if re-running
    if "Dynamic Staffing Plan" in wb.sheetnames:
        del wb["Dynamic Staffing Plan"]
        print("  Removed existing 'Dynamic Staffing Plan' sheet.")

    # Build new summary sheet
    build_dynamic_summary(wb)
    sheets = wb.sheetnames
    if "Staffing Model" in sheets:
        target_idx = sheets.index("Staffing Model") + 1
        offset = target_idx - (len(sheets) - 1)
        wb.move_sheet("Dynamic Staffing Plan", offset=offset)

    # Append dynamic section to each monthly sheet (idempotent)
    for loc in LOCATIONS:
        sheet_name = loc["sheet"]
        if sheet_name in wb.sheetnames:
            print(f"  Appending dynamic section → {sheet_name}")
            append_dynamic_section(wb[sheet_name], loc)
        else:
            print(f"  WARNING: Sheet '{sheet_name}' not found, skipping.")

    wb.save(path)
    print(f"\n✓ Saved: {path}")
    print(f"  Final sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

    # Sanity check
    total_saving = sum(
        sum(
            hc_cost_breakdown(rev, loc["fixed_t"], loc["r"], loc["rm"])[2] -
            hc_cost_breakdown(rev, optimal_t(rev, loc["floor_min"]), loc["r"], loc["rm"])[2]
            for rev in monthly_revenues(loc["annual_rev"])
        )
        for loc in LOCATIONS
    )
    print(f"\n  Portfolio annual saving: €{total_saving:,.0f}")
    print("  (Positive = savings vs fixed model | Negative = understaffed = need to hire)")


if __name__ == "__main__":
    main()
