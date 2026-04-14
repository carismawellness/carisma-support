"""Redesign Sheet1 in Carisma Spa Staffing Model.xlsx
Transforms raw brainstorm list into categorized, prioritized action plan.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

FILE = os.path.join(os.path.dirname(__file__), '../hr/Carisma Spa Staffing Model.xlsx')

# Colors
NAVY = "1B3A4B"
GOLD = "B8943E"
WHITE = "FFFFFF"
LIGHT_GOLD = "FDF3DC"
LIGHT_GREEN = "D4EFDF"
LIGHT_RED = "FADBD8"
LIGHT_AMBER = "FAD7A0"
LIGHT_GRAY = "F2F3F4"
DARK_GRAY = "5D6D7E"

def fill(color):
    return PatternFill("solid", fgColor=color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

thin_border = Border(
    bottom=Side(style="thin", color="D5D8DC"),
)

# ── DATA ──
# Each item: (category, action, priority, status, owner, notes)
# Priority: HIGH / MED / LOW
# Status: ✓ ANALYSED / → ACTIVE / ○ TODO

items = [
    # ── STAFFING & ROSTER ──
    ("STAFFING & ROSTER", "Receptionist + hourly who can do therapies (dual-competency T-S role)", "HIGH", "✓ ANALYSED", "HR / Ops", "Highest-ROI idea from strategic review. Saves €300-500/mo per location vs separate hire."),
    ("STAFFING & ROSTER", "Part-time therapists for weekends (Fri–Sun coverage)", "MED", "✓ ANALYSED", "HR", "Structurally correct but operationally complex. Best for Novotel & Excelsior."),
    ("STAFFING & ROSTER", "No more advisors — focus on therapists", "HIGH", "→ ACTIVE", "COO", "Inter has 3 advisors (€6K/mo). Ramla has 2 (€3.4K). Review if revenue justifies cost."),
    ("STAFFING & ROSTER", "Roster optimize — shift patterns A/B coverage", "HIGH", "✓ ANALYSED", "HR / Ops", "Full weekly roster built with actual names. 5-day week, Sat mandatory, A+B shift coverage."),
    ("STAFFING & ROSTER", "Reduce opening hours (9-19 vs 10-20) per location", "MED", "✓ ANALYSED", "COO", "Data supports location-specific approach. Riviera/Odycy/Novo could close earlier."),
    ("STAFFING & ROSTER", "HR check roster & Talexio weekly", "MED", "→ ACTIVE", "HR", "Roster review weekly with phone check-in. Ensure Talexio matches actual shifts."),
    ("STAFFING & ROSTER", "Put stop-leave on Talexio", "LOW", "○ TODO", "HR", "Block leave requests during peak periods. Requires Talexio config."),
    ("STAFFING & ROSTER", "Peak windows are painful — coverage gaps", "HIGH", "→ ACTIVE", "Ops", "Fri/Sat/Sun coverage needs all therapists. Current roster addresses this."),

    # ── COMPENSATION & INCENTIVES ──
    ("COMPENSATION", "Regional Manager promotions (Anna / Neli / Melanie)", "HIGH", "✓ ANALYSED", "CEO / HR", "CFO modelled: +€600/mo net cost. Anna €4,500, Neli €4,200, Melanie €3,800. Cost-neutral with T-S backfill."),
    ("COMPENSATION", "Managers getting paid on net revenue", "MED", "→ ACTIVE", "CFO / HR", "Align manager comp with location P&L. Requires defining 'net revenue' per location."),
    ("COMPENSATION", "HR to get bonus based on net revenue", "LOW", "○ TODO", "CFO", "HR incentive alignment. Define metric: retention rate + HC% target hit."),
    ("COMPENSATION", "Concierge 1.5% commission, 3% if spa hits target", "MED", "○ TODO", "CFO", "Tiered commission for advisors/concierge. Incentivises upsell + team performance."),
    ("COMPENSATION", "Show rate target % bonus", "MED", "○ TODO", "CFO / Ops", "Bonus for reducing no-show rate. Needs baseline measurement first."),
    ("COMPENSATION", "Hourly pay model?", "LOW", "○ TODO", "CFO / HR", "Evaluate hourly vs monthly for PT roles. Malta min wage €994/mo = ~€5.75/hr."),

    # ── OPERATIONS & UTILISATION ──
    ("OPERATIONS", "Utilisation optimise — Tx/Day/Therapist tracking", "HIGH", "✓ ANALYSED", "Ops", "Full analysis done. Ramla 3.07, Odycy 3.08, Excelsior 3.34 — all ⚠ UNDER target 3.5."),
    ("OPERATIONS", "Therapists sitting idle — address under-utilisation", "HIGH", "✓ ANALYSED", "Ops / HR", "Ramla (6T on €39K rev), Odycy (4T on €24K rev), Excelsior (4T on €29K rev) need reduction."),
    ("OPERATIONS", "Every department has EOD checklist", "MED", "○ TODO", "COO", "Standardise end-of-day process across all 9 locations. CarismaSoft will automate."),
    ("OPERATIONS", "Reception tasks reduction — too much admin", "MED", "○ TODO", "Ops", "Audit reception workflows. Remove non-essential admin. Free up for customer service."),
    ("OPERATIONS", "Number of treatments per month per spa", "HIGH", "✓ ANALYSED", "Ops", "Built into Treatment Analysis. Inter 556, Hugos 533, Ramla 333, Riviera 222, etc."),
    ("OPERATIONS", "HC % 40% target", "HIGH", "✓ ANALYSED", "CFO / COO", "Actuals built. Portfolio at 47.4% loaded. Ramla/Odycy/Excelsior at 49-55% — ❌ OVER."),

    # ── SALES & REVENUE ──
    ("SALES & REVENUE", "Smart pricing — location-specific + seasonal", "MED", "○ TODO", "COO / CFO", "CarismaSoft Phase 2 feature. Define peak/off-peak pricing per location."),
    ("SALES & REVENUE", "Fix speed-to-lead issue", "HIGH", "○ TODO", "CRM / Sales", "Photo attachment reference. Reduce response time to enquiries."),
    ("SALES & REVENUE", "CRM phone call sales training optimisation", "MED", "○ TODO", "CRM Manager", "Upsell script training. See Hormozi 'Menu Upsell' framework below."),
    ("SALES & REVENUE", "Alex Hormozi 'Menu Upsell' framework", "MED", "○ TODO", "CRM / Training", "1. Unsell  2. Prescribe  3. Choice (A or B?)  4. Card on file. Train all concierge staff."),
    ("SALES & REVENUE", "Arpit sales improvement in hotel", "MED", "○ TODO", "Sales", "Hotel partnership sales. Requires dedicated hotel liaison approach."),
    ("SALES & REVENUE", "Excelsior hotel packages: GHE Spa Package", "MED", "○ TODO", "Sales / Excl", "Package: Spa day + Lunch + Outdoor pool. Cross-sell with hotel guests."),
    ("SALES & REVENUE", "Spa day + Lunch + Outdoor pool package", "MED", "○ TODO", "Sales", "Premium day-pass concept. Requires F&B partnership with hotel."),
    ("SALES & REVENUE", "4 levers of growth: affiliates, referral, ads, cold outreach", "HIGH", "→ ACTIVE", "CMO", "Strategic growth framework. All 4 channels need activation."),

    # ── SYSTEMS & KPIs ──
    ("SYSTEMS & KPIs", "CRM team Roster, QC, KPIs, Bonuses, & Management systems", "HIGH", "→ ACTIVE", "CRM Manager", "Full CRM ops framework. Being built in carisma-support repo."),
    ("SYSTEMS & KPIs", "Full funnel KPIs buildout", "HIGH", "→ ACTIVE", "CMO / CRM", "Lead → Booking → Show → Rebook funnel. Needs per-brand dashboards."),

    # ── CULTURE & TRAINING ──
    ("CULTURE & TRAINING", "Trainer routine — standard operating procedures", "MED", "○ TODO", "HR / Training", "Create training SOP per role. Onboard new staff consistently."),
    ("CULTURE & TRAINING", "Smiling at all times: unreasonable hospitality", "MED", "→ ACTIVE", "All Managers", "Culture standard. Reinforce in daily briefings. Mystery shopper audits."),
    ("CULTURE & TRAINING", "Receptionist covered by therapists → phone quality drops", "HIGH", "→ ACTIVE", "Ops", "When therapist covers reception, calls aren't handled well. Fix with dedicated reception."),
    ("CULTURE & TRAINING", "Excelsior towels — need someone dedicated there", "LOW", "○ TODO", "Excl Manager", "Towel management issue. Assign to spa attendant or reception."),
    ("CULTURE & TRAINING", "Excelsior get reception to cover", "LOW", "○ TODO", "Excl Manager", "Reception coverage gap at Excelsior. Kemi RA should handle."),
]

CATEGORY_COLORS = {
    "STAFFING & ROSTER": "2E86C1",
    "COMPENSATION": "8E44AD",
    "OPERATIONS": "E67E22",
    "SALES & REVENUE": "27AE60",
    "SYSTEMS & KPIs": "2C3E50",
    "CULTURE & TRAINING": "C0392B",
}

PRIORITY_FILLS = {
    "HIGH": fill("FADBD8"),
    "MED": fill("FAD7A0"),
    "LOW": fill("D4EFDF"),
}

STATUS_FILLS = {
    "✓ ANALYSED": fill("D4EFDF"),
    "→ ACTIVE": fill("D6EAF8"),
    "○ TODO": fill(LIGHT_GRAY),
}


def build_sheet(wb):
    # Delete existing Sheet1 and create fresh
    if 'Sheet1' in wb.sheetnames:
        del wb['Sheet1']
    ws = wb.create_sheet('Action Plan', 0)

    # ── TITLE ──
    ws.merge_cells('A1:G1')
    c = ws.cell(row=1, column=1)
    c.value = "CARISMA SPA — STRATEGIC ACTION PLAN"
    c.font = font(bold=True, color=WHITE, size=14)
    c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells('A2:G2')
    c = ws.cell(row=2, column=1)
    c.value = "33 initiatives categorised from brainstorm session  |  Status as of Apr 2026  |  8 items analysed this session"
    c.font = font(italic=True, color=WHITE, size=9)
    c.fill = fill(GOLD)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # ── HEADER ROW ──
    headers = ["CATEGORY", "#", "ACTION ITEM", "PRIORITY", "STATUS", "OWNER", "NOTES / CONTEXT"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i)
        c.value = h
        c.font = font(bold=True, color=WHITE, size=9)
        c.fill = fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 28

    # ── COLUMN WIDTHS ──
    widths = [22, 4, 55, 10, 14, 16, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── DATA ROWS ──
    row = 5
    current_category = None
    cat_start_row = None
    num = 0

    for item in items:
        cat, action, priority, status, owner, notes = item
        num += 1

        # Category column
        if cat != current_category:
            if current_category and cat_start_row and row - 1 > cat_start_row:
                ws.merge_cells(start_row=cat_start_row, start_column=1,
                               end_row=row - 1, end_column=1)
            current_category = cat
            cat_start_row = row
            cat_color = CATEGORY_COLORS.get(cat, DARK_GRAY)
            c = ws.cell(row=row, column=1)
            c.value = cat
            c.font = font(bold=True, color=WHITE, size=8)
            c.fill = fill(cat_color)
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True, text_rotation=0)
        else:
            c = ws.cell(row=row, column=1)
            c.fill = fill(CATEGORY_COLORS.get(cat, DARK_GRAY))

        # Number
        c = ws.cell(row=row, column=2)
        c.value = num
        c.alignment = Alignment(horizontal="center")
        c.font = font(color=DARK_GRAY, size=9)

        # Action item
        c = ws.cell(row=row, column=3)
        c.value = action
        c.font = font(size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")

        # Priority
        c = ws.cell(row=row, column=4)
        c.value = priority
        c.font = font(bold=True, size=9)
        c.alignment = Alignment(horizontal="center")
        if priority in PRIORITY_FILLS:
            c.fill = PRIORITY_FILLS[priority]

        # Status
        c = ws.cell(row=row, column=5)
        c.value = status
        c.font = font(bold=True, size=8)
        c.alignment = Alignment(horizontal="center")
        if status in STATUS_FILLS:
            c.fill = STATUS_FILLS[status]

        # Owner
        c = ws.cell(row=row, column=6)
        c.value = owner
        c.font = font(size=9, color=DARK_GRAY)
        c.alignment = Alignment(horizontal="center", wrap_text=True)

        # Notes
        c = ws.cell(row=row, column=7)
        c.value = notes
        c.font = font(size=8, color=DARK_GRAY, italic=True)
        c.alignment = Alignment(wrap_text=True, vertical="top")

        # Row border
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border

        ws.row_dimensions[row].height = 32
        row += 1

    # Final category merge
    if current_category and cat_start_row and row - 1 > cat_start_row:
        ws.merge_cells(start_row=cat_start_row, start_column=1,
                       end_row=row - 1, end_column=1)

    # ── SUMMARY SECTION ──
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1)
    c.value = "SUMMARY"
    c.font = font(bold=True, color=WHITE, size=10)
    c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="left")

    stats = [
        ("Total initiatives", 33),
        ("✓ ANALYSED (this session)", sum(1 for i in items if i[3] == "✓ ANALYSED")),
        ("→ ACTIVE (in progress)", sum(1 for i in items if i[3] == "→ ACTIVE")),
        ("○ TODO (not started)", sum(1 for i in items if i[3] == "○ TODO")),
        ("HIGH priority", sum(1 for i in items if i[2] == "HIGH")),
        ("MED priority", sum(1 for i in items if i[2] == "MED")),
        ("LOW priority", sum(1 for i in items if i[2] == "LOW")),
    ]

    row += 1
    for label, val in stats:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = font(size=9)
        ws.cell(row=row, column=2).value = val
        ws.cell(row=row, column=2).font = font(bold=True, size=9)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1

    # ── LEGEND ──
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1)
    c.value = "LEGEND"
    c.font = font(bold=True, color=WHITE, size=9)
    c.fill = fill(DARK_GRAY)
    row += 1

    legend = [
        ("Priority:", "HIGH = urgent / high financial impact", PRIORITY_FILLS["HIGH"]),
        ("", "MED = important but not blocking", PRIORITY_FILLS["MED"]),
        ("", "LOW = nice-to-have / future", PRIORITY_FILLS["LOW"]),
        ("Status:", "✓ ANALYSED = data analysis completed this session", STATUS_FILLS["✓ ANALYSED"]),
        ("", "→ ACTIVE = work in progress", STATUS_FILLS["→ ACTIVE"]),
        ("", "○ TODO = not started", STATUS_FILLS["○ TODO"]),
    ]

    for label, desc, bg in legend:
        if label:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = font(bold=True, size=8)
        ws.cell(row=row, column=2).value = desc
        ws.cell(row=row, column=2).font = font(size=8)
        ws.cell(row=row, column=3).fill = bg
        row += 1

    return ws


if __name__ == '__main__':
    wb = load_workbook(FILE)
    build_sheet(wb)
    wb.save(FILE)
    print("✓ Sheet1 redesigned as 'Action Plan' and saved.")
