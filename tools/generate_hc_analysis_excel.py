#!/usr/bin/env python3
"""
Generate HC Cost Analysis Excel workbook with multiple sheets.
Covers 35% and 40% HC targets, revenue data, staffing, rosters, and pressure test results.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── Styling ──────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="2F5496")
NUM_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
PCT_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")
GREY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

EUR_FMT = '#,##0'
EUR_DEC_FMT = '#,##0.00'
PCT_FMT = '0.0%'


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_subheader_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, fmt=None, bold=False):
    cell = ws.cell(row=row, column=col)
    cell.font = BOLD_FONT if bold else NUM_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="right" if col > 1 else "left")
    if fmt:
        cell.number_format = fmt


def write_table(ws, start_row, headers, data, fmt=None, total_row=None):
    """Write a table with headers and data rows. Returns next empty row."""
    r = start_row
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, len(headers))
    r += 1

    for row_data in data:
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
            style_data_cell(ws, r, c, fmt=fmt if c > 1 else None)
        r += 1

    if total_row:
        for c, val in enumerate(total_row, 1):
            ws.cell(row=r, column=c, value=val)
            style_data_cell(ws, r, c, fmt=fmt if c > 1 else None, bold=True)
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).fill = SUBHEADER_FILL
        r += 1

    return r + 1


def auto_width(ws, min_width=10, max_width=18):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 3, max_width))


# ── Data ─────────────────────────────────────────────────────────────────────

MONTHS = ["Jan 25", "Feb 25", "Mar 25", "Apr 25", "May 25", "Jun 25",
          "Jul 25", "Aug 25", "Sep 25", "Oct 25", "Nov 25", "Dec 25", "Jan 26"]

HOTELS = ["Inter", "Hugos", "Hyatt", "Ramla", "Riviera", "Odycy", "Excelsior", "Novotel"]

# Services ex-VAT per hotel per month (from YTD tab)
REVENUE = {
    "Inter":     [36890, 43283, 50300, 45332, 48798, 40584, 47586, 57295, 51920, 57277, 49655, 36377, 50578],
    "Hugos":     [44411, 47851, 57034, 56849, 46791, 37041, 38302, 42479, 44564, 56213, 46415, 36026, 52052],
    "Hyatt":     [23492, 27685, 27747, 22679, 24002, 21047, 21844, 22055, 21490, 23536, 22371, 18794, 25734],
    "Ramla":     [20792, 24924, 26617, 26530, 30780, 23004, 25274, 31313, 32455, 46327, 32425, 28304, 34501],
    "Riviera":   [11381, 12873, 15403, 17442, 17999, 18094, 20859, 25032, 25774, 21434, 19079, 13928, 15468],
    "Odycy":     [9661, 11950, 12098, 14637, 18652, 16260, 19741, 22435, 20616, 23337, 18050, 11084, 17600],
    "Excelsior": [0, 0, 0, 0, 0, 0, 20990, 21625, 24634, 24708, 22223, 20356, 19553],
    "Novotel":   [0, 0, 0, 0, 0, 0, 0, 0, 0, 3104, 14143, 10301, 8019],
}

# Per-role loaded monthly costs
ROLE_COSTS = {
    "Regional Manager (RM)": {"base": 2350, "commission": 750, "ni": 310, "total": 3410, "annual": 40920},
    "Therapist-in-Charge (TIC)": {"base": 1600, "commission": 250, "ni": 185, "total": 2035, "annual": 24420},
    "Receptionist": {"base": 1350, "commission": 180, "ni": 153, "total": 1683, "annual": 20196},
    "FT Therapist": {"base": 1450, "commission": 200, "ni": 165, "total": 1815, "annual": 21780},
    "PT Therapist": {"base": 750, "commission": 100, "ni": 85, "total": 935, "annual": 11220},
    "Intern/Student": {"base": 650, "commission": 0, "ni": 65, "total": 715, "annual": 8580},
}

# RM allocation by hotel (revenue-proportional)
RM_ALLOC = {
    "Inter": 1432, "Hugos": 1330, "Hyatt": 648,
    "Ramla": 1603, "Riviera": 921, "Odycy": 886,
    "Excelsior": 2421, "Novotel": 989,
}

# Staffing at 35% target
STAFF_35 = {
    "Inter":     {"Recep": 1, "TIC": 1, "FT": 6, "PT": 1, "Intern": 0, "staff_cost": 15543, "headcount": 9},
    "Hugos":     {"Recep": 1, "TIC": 1, "FT": 5, "PT": 1, "Intern": 0, "staff_cost": 13728, "headcount": 8},
    "Ramla":     {"Recep": 1, "TIC": 0, "FT": 4, "PT": 0, "Intern": 1, "staff_cost": 9658, "headcount": 6},
    "Hyatt":     {"Recep": 0, "TIC": 1, "FT": 2, "PT": 1, "Intern": 0, "staff_cost": 6600, "headcount": 4},
    "Riviera":   {"Recep": 0, "TIC": 1, "FT": 2, "PT": 0, "Intern": 0, "staff_cost": 5665, "headcount": 3},
    "Odycy":     {"Recep": 0, "TIC": 1, "FT": 2, "PT": 0, "Intern": 0, "staff_cost": 5665, "headcount": 3},
    "Excelsior": {"Recep": 0, "TIC": 1, "FT": 2, "PT": 0, "Intern": 0, "staff_cost": 5665, "headcount": 3},
    "Novotel":   {"Recep": 0, "TIC": 1, "FT": 0, "PT": 1, "Intern": 0, "staff_cost": 2970, "headcount": 2},
}

# Staffing at 40% target
STAFF_40 = {
    "Inter":     {"Recep": 1, "TIC": 1, "FT": 7, "PT": 1, "Intern": 0, "staff_cost": 17358, "headcount": 10},
    "Hugos":     {"Recep": 1, "TIC": 1, "FT": 6, "PT": 1, "Intern": 0, "staff_cost": 15543, "headcount": 9},
    "Ramla":     {"Recep": 1, "TIC": 0, "FT": 5, "PT": 0, "Intern": 1, "staff_cost": 11473, "headcount": 7},
    "Hyatt":     {"Recep": 0, "TIC": 1, "FT": 3, "PT": 0, "Intern": 0, "staff_cost": 7480, "headcount": 4},
    "Riviera":   {"Recep": 0, "TIC": 1, "FT": 2, "PT": 1, "Intern": 0, "staff_cost": 6600, "headcount": 4},
    "Odycy":     {"Recep": 0, "TIC": 1, "FT": 2, "PT": 1, "Intern": 0, "staff_cost": 6600, "headcount": 4},
    "Excelsior": {"Recep": 0, "TIC": 1, "FT": 3, "PT": 0, "Intern": 0, "staff_cost": 7480, "headcount": 4},
    "Novotel":   {"Recep": 0, "TIC": 1, "FT": 1, "PT": 0, "Intern": 0, "staff_cost": 3850, "headcount": 2},
}

CURRENT_STAFF = {
    "Inter": 12, "Hugos": 9, "Hyatt": 5, "Ramla": 9,
    "Riviera": 7, "Odycy": 6, "Excelsior": 7, "Novotel": 6,
}

# Pressure test: peak utilization %
PEAK_UTIL = {
    "Inter":     {"40%": 0.69, "35%": 0.85},
    "Hugos":     {"40%": 0.76, "35%": 0.96},
    "Hyatt":     {"40%": 0.56, "35%": 0.78},
    "Ramla":     {"40%": 0.76, "35%": 1.00},
    "Riviera":   {"40%": 1.07, "35%": 1.07},
    "Odycy":     {"40%": 1.09, "35%": 1.09},
    "Excelsior": {"40%": 0.47, "35%": 0.65},
    "Novotel":   {"40%": 0.56, "35%": 2.40},
}

RISK_LEVELS = {
    "Inter":     {"40%": "LOW", "35%": "MEDIUM"},
    "Hugos":     {"40%": "LOW", "35%": "HIGH"},
    "Hyatt":     {"40%": "LOW", "35%": "MEDIUM"},
    "Ramla":     {"40%": "LOW-MEDIUM", "35%": "CRITICAL"},
    "Riviera":   {"40%": "MEDIUM", "35%": "HIGH"},
    "Odycy":     {"40%": "MEDIUM", "35%": "HIGH"},
    "Excelsior": {"40%": "LOW", "35%": "MEDIUM"},
    "Novotel":   {"40%": "HIGH", "35%": "CRITICAL"},
}

# Named personnel actions
PERSONNEL = [
    ("Neli", "Hugos", "Manager", "Promote to RM1", "Regional"),
    ("Kristina", "Ramla", "Supervisor", "Promote to RM2", "Regional"),
    ("Melanie", "Excelsior", "Manager", "Transition to RM3 + Corp Training", "Regional"),
    ("Anna", "Inter", "Manager", "Offer receptionist at Ramla OR redundancy", "Ramla / Exit"),
    ("Ebru", "Novotel", "Manager", "Redundancy", "Exit"),
    ("Flora", "Hyatt", "Supervisor", "Reclassify to TIC", "Hyatt"),
    ("Wanessa", "Riviera", "Supervisor", "Redundancy or redeploy as FT therapist", "Exit / Hub"),
    ("Katia", "Riviera", "Concierge", "Redundancy", "Exit"),
    ("Lanna", "Riviera", "Advisor", "Redundancy", "Exit"),
    ("Lais", "Ramla", "Advisor", "Redundancy", "Exit"),
    ("Benay", "Odycy", "Advisor", "Redundancy", "Exit"),
    ("Daniely", "Excelsior", "Advisor", "Redundancy", "Exit"),
    ("Linara", "Odycy", "Supervisor", "Redeploy as FT therapist at hub OR redundancy", "Hub / Exit"),
    ("Maila", "Inter", "Concierge", "Redundancy", "Exit"),
    ("Milena", "Inter", "Therapist", "Redeploy to Ramla", "Ramla"),
    ("Laura Camila", "Ramla", "Therapist", "Redeploy to Riviera (at 35%) / Stay (at 40%)", "Riviera / Ramla"),
    ("Dzanela", "Hyatt", "Therapist", "Redeploy to float pool", "Float"),
    ("Sebastian", "Excelsior", "Therapist", "Redeploy to hub (at 35%) / Stay (at 40%)", "Hub / Excelsior"),
    ("Rita", "Riviera", "Therapist", "Redeploy to Odycy or redundancy (at 35%) / Stay (at 40%)", "Odycy / Riviera"),
    ("Adriene", "Novotel", "Therapist", "Redeploy to hub", "Hub"),
    ("Lisa", "Novotel", "Therapist", "Redeploy or redundancy", "Hub / Exit"),
]

# Roster data — 35% target
ROSTER_35 = {
    "Inter": {
        "headers": ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"],
        "data": [
            ["Natasha", "Recep", "E", "E", "E", "E", "E", "E", "OFF", 40],
            ["Nathalia R", "TIC/Adv", "E", "E", "OFF", "E", "E", "L", "E", 40],
            ["Mini", "FT Th", "E", "L", "E", "OFF", "E", "E", "L", 40],
            ["Pakinee", "FT Th", "L", "E", "E", "E", "OFF", "L", "E", 40],
            ["Christopher", "FT Th", "E", "OFF", "L", "E", "L", "E", "E", 40],
            ["Valeri", "FT Th", "OFF", "E", "E", "L", "E", "E", "L", 40],
            ["Julie", "FT Th", "E", "L", "E", "E", "L", "OFF", "E", 40],
            ["Adriana", "FT Th", "L", "E", "L", "E", "E", "E", "OFF", 40],
            ["Matilde", "PT Th", "E4", "OFF", "E4", "E4", "E4", "E4", "OFF", 20],
        ],
    },
    "Hugos": {
        "headers": ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"],
        "data": [
            ["Alana", "Recep", "E", "E", "E", "E", "E", "E", "OFF", 40],
            ["Ajlin", "TIC/Adv", "E", "OFF", "E", "E", "E", "L", "E", 40],
            ["Lourdes", "FT Th", "E", "E", "L", "OFF", "E", "E", "L", 40],
            ["Lovely", "FT Th", "L", "E", "E", "E", "OFF", "L", "E", 40],
            ["Pacha", "FT Th", "E", "L", "OFF", "E", "L", "E", "E", 40],
            ["Tessa", "FT Th", "OFF", "E", "E", "L", "E", "E", "L", 40],
            ["Tamara", "FT Th", "E", "E", "E", "E", "L", "OFF", "E", 40],
            ["Tina", "PT Th", "E4", "OFF", "E4", "E4", "E4", "E4", "OFF", 20],
        ],
    },
    "Ramla": {
        "headers": ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"],
        "data": [
            ["Natalia B", "Recep", "E", "E", "E", "E", "E", "E", "OFF", 40],
            ["Marvick", "FT Th", "E", "L", "OFF", "E", "E", "L", "E", 40],
            ["Karla", "FT Th", "L", "E", "E", "OFF", "L", "E", "E", 40],
            ["Vitor", "FT Th", "E", "OFF", "L", "E", "E", "E", "L", 40],
            ["Irene", "FT Th", "OFF", "E", "E", "L", "E", "E", "E", 40],
            ["Intern", "Intern", "E4", "E4", "E4", "E4", "OFF", "E4", "OFF", 20],
        ],
    },
    "Hyatt": {
        "headers": ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"],
        "data": [
            ["Flora", "TIC", "E", "E", "L", "E", "OFF", "E", "E", 40],
            ["Ety", "FT Th", "E", "OFF", "E", "L", "E", "L", "E", 40],
            ["Claudia", "FT Th", "L", "E", "E", "E", "E", "OFF", "L", 40],
            ["Natasha H", "PT Th", "E6", "E6", "OFF", "E6", "E6", "E6", "OFF", 30],
        ],
    },
    "Riviera": {
        "headers": ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"],
        "data": [
            ["Silvia", "TIC", "E", "E", "OFF", "E", "L", "E", "E", 40],
            ["Blago", "FT Th", "L", "OFF", "E", "E", "E", "L", "E", 40],
            ["Laura C.", "FT Th", "E", "E", "L", "E", "OFF", "E", "L", 40],
        ],
    },
    "Odycy": {
        "headers": ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"],
        "data": [
            ["Jovanna", "TIC", "E", "E", "L", "OFF", "E", "E", "E", 40],
            ["Elizabetta", "FT Th", "OFF", "E", "E", "E", "L", "E", "L", 40],
            ["Jenny", "FT Th", "E", "L", "E", "E", "OFF", "L", "E", 40],
        ],
    },
    "Excelsior": {
        "headers": ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"],
        "data": [
            ["Sofia", "TIC", "E", "E", "OFF", "E", "E", "L", "E", 40],
            ["Carlos", "FT Th", "L", "OFF", "E", "E", "L", "E", "E", 40],
            ["Lorena", "FT Th", "E", "E", "L", "OFF", "E", "E", "L", 40],
        ],
    },
    "Novotel": {
        "headers": ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"],
        "data": [
            ["Angel", "TIC", "E", "E", "E", "L", "OFF", "E", "E", 40],
            ["Vanessa", "PT Th", "E6", "OFF", "E6", "E6", "E6", "E6", "OFF", 30],
        ],
    },
}

# Roster data — 40% target (only differences from 35%)
ROSTER_40 = {
    "Inter": {
        "headers": ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"],
        "data": [
            ["Natasha", "Recep", "E", "E", "E", "E", "E", "E", "OFF", 40],
            ["Nathalia R", "TIC/Adv", "E", "E", "OFF", "E", "E", "L", "E", 40],
            ["Mini", "FT Th", "E", "L", "E", "OFF", "E", "E", "L", 40],
            ["Pakinee", "FT Th", "L", "E", "E", "E", "OFF", "L", "E", 40],
            ["Christopher", "FT Th", "E", "OFF", "L", "E", "L", "E", "E", 40],
            ["Valeri", "FT Th", "OFF", "E", "E", "L", "E", "E", "L", 40],
            ["Julie", "FT Th", "E", "L", "E", "E", "L", "OFF", "E", 40],
            ["Adriana", "FT Th", "L", "E", "L", "E", "E", "E", "OFF", 40],
            ["Matilde", "FT Th", "E", "E", "OFF", "E", "E", "L", "E", 40],
            ["Milena", "PT Th", "OFF", "E4", "E4", "E4", "E4", "E4", "E4", 24],
        ],
    },
    "Hugos": {
        "headers": ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"],
        "data": [
            ["Alana", "Recep", "E", "E", "E", "E", "E", "E", "OFF", 40],
            ["Ajlin", "TIC/Adv", "E", "OFF", "E", "E", "E", "L", "E", 40],
            ["Lourdes", "FT Th", "E", "E", "L", "OFF", "E", "E", "L", 40],
            ["Lovely", "FT Th", "L", "E", "E", "E", "OFF", "L", "E", 40],
            ["Pacha", "FT Th", "E", "L", "OFF", "E", "L", "E", "E", 40],
            ["Tessa", "FT Th", "OFF", "E", "E", "L", "E", "E", "L", 40],
            ["Tamara", "FT Th", "E", "E", "E", "E", "L", "OFF", "E", 40],
            ["Tina", "FT Th", "E", "OFF", "E", "E", "E", "E", "L", 40],
            ["Linara", "PT Th", "OFF", "E4", "E4", "OFF", "E4", "E4", "E4", 20],
        ],
    },
    "Ramla": {
        "headers": ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"],
        "data": [
            ["Natalia B", "Recep", "E", "E", "E", "E", "E", "E", "OFF", 40],
            ["Marvick", "FT Th", "E", "L", "OFF", "E", "E", "L", "E", 40],
            ["Karla", "FT Th", "L", "E", "E", "OFF", "L", "E", "E", 40],
            ["Vitor", "FT Th", "E", "OFF", "L", "E", "E", "E", "L", 40],
            ["Irene", "FT Th", "OFF", "E", "E", "L", "E", "E", "E", 40],
            ["Laura Camila", "FT Th", "E", "E", "E", "E", "OFF", "E", "OFF", 40],
            ["Intern", "Intern", "E4", "E4", "OFF", "E4", "E4", "E4", "E4", 24],
        ],
    },
    "Hyatt": {
        "headers": ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"],
        "data": [
            ["Flora", "TIC", "E", "E", "L", "E", "OFF", "E", "E", 40],
            ["Ety", "FT Th", "E", "OFF", "E", "L", "E", "L", "E", 40],
            ["Claudia", "FT Th", "L", "E", "E", "E", "E", "OFF", "L", 40],
            ["Natasha H", "FT Th", "E", "E", "OFF", "E", "L", "E", "E", 40],
        ],
    },
    "Riviera": {
        "headers": ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"],
        "data": [
            ["Silvia", "TIC", "E", "E", "OFF", "E", "L", "E", "E", 40],
            ["Blago", "FT Th", "L", "OFF", "E", "E", "E", "L", "E", 40],
            ["Rita", "FT Th", "E", "E", "L", "OFF", "E", "E", "L", 40],
            ["PT (new)", "PT Th", "OFF", "E4", "E4", "E4", "OFF", "E4", "E4", 20],
        ],
    },
    "Odycy": {
        "headers": ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"],
        "data": [
            ["Jovanna", "TIC", "E", "E", "L", "OFF", "E", "E", "E", 40],
            ["Elizabetta", "FT Th", "OFF", "E", "E", "E", "L", "E", "L", 40],
            ["Jenny", "FT Th", "E", "L", "E", "E", "OFF", "L", "E", 40],
            ["PT (new)", "PT Th", "E4", "OFF", "E4", "E4", "E4", "OFF", "E4", 20],
        ],
    },
    "Excelsior": {
        "headers": ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"],
        "data": [
            ["Sofia", "TIC", "E", "E", "OFF", "E", "E", "L", "E", 40],
            ["Carlos", "FT Th", "L", "OFF", "E", "E", "L", "E", "E", 40],
            ["Lorena", "FT Th", "E", "E", "L", "OFF", "E", "E", "L", 40],
            ["Sebastian", "FT Th", "OFF", "E", "E", "E", "E", "OFF", "E", 40],
        ],
    },
    "Novotel": {
        "headers": ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"],
        "data": [
            ["Angel", "TIC", "E", "E", "E", "L", "OFF", "E", "E", 40],
            ["Vanessa", "FT Th", "E", "OFF", "E", "E", "E", "L", "OFF", 40],
        ],
    },
}

# Financial sensitivity per hotel
SENSITIVITY = {
    "Inter":     {"current_staff": 12, "current_cost": 20600, "40_staff": 9, "40_cost": 15408, "40_risk": 0, "40_net": 5192, "35_staff": 7, "35_cost": 12500, "35_risk": 1500, "35_net": 6600},
    "Hugos":     {"current_staff": 10, "current_cost": 18200, "40_staff": 8, "40_cost": 13593, "40_risk": 0, "40_net": 4607, "35_staff": 6, "35_cost": 10700, "35_risk": 5250, "35_net": 2250},
    "Hyatt":     {"current_staff": 6, "current_cost": 10600, "40_staff": 4, "40_cost": 7480, "40_risk": 400, "40_net": 2720, "35_staff": 3, "35_cost": 5665, "35_risk": 1200, "35_net": 3735},
    "Ramla":     {"current_staff": 9, "current_cost": 15900, "40_staff": 7, "40_cost": 11473, "40_risk": 0, "40_net": 4427, "35_staff": 5, "35_cost": 9100, "35_risk": 5000, "35_net": 1800},
    "Riviera":   {"current_staff": 7, "current_cost": 12200, "40_staff": 3, "40_cost": 5665, "40_risk": 1200, "40_net": 5335, "35_staff": 3, "35_cost": 5665, "35_risk": 2500, "35_net": 4035},
    "Odycy":     {"current_staff": 6, "current_cost": 10400, "40_staff": 3, "40_cost": 5665, "40_risk": 1000, "40_net": 3735, "35_staff": 3, "35_cost": 5665, "35_risk": 1500, "35_net": 3235},
    "Excelsior": {"current_staff": 7, "current_cost": 13100, "40_staff": 4, "40_cost": 7480, "40_risk": 800, "40_net": 4820, "35_staff": 3, "35_cost": 5665, "35_risk": 1200, "35_net": 6235},
    "Novotel":   {"current_staff": 7, "current_cost": 11500, "40_staff": 2, "40_cost": 3850, "40_risk": 0, "40_net": 7650, "35_staff": 1, "35_cost": 2035, "35_risk": 3000, "35_net": 6465},
}


def calc_6m_avg(rev_list):
    """Recent 6-month average (Aug 25 - Jan 26, indices 7-12)."""
    vals = [v for v in rev_list[7:13] if v > 0]
    return round(sum(vals) / len(vals)) if vals else 0


def calc_12m_avg(rev_list):
    """12-month average (Jan-Dec 2025, indices 0-11), only months with revenue."""
    vals = [v for v in rev_list[0:12] if v > 0]
    return round(sum(vals) / len(vals)) if vals else 0


# ── Sheet Builders ───────────────────────────────────────────────────────────

def build_executive_summary(wb):
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = "2F5496"

    ws.merge_cells("A1:I1")
    ws.cell(row=1, column=1, value="HC Cost Optimization: Executive Summary").font = TITLE_FONT

    ws.merge_cells("A3:I3")
    ws.cell(row=3, column=1, value="Key Finding: 40% is the only responsible Year 1 target — it saves EUR 47K MORE net than 35%.").font = SUBTITLE_FONT

    r = 5
    headers = ["Metric", "35% Target", "40% Target", "Current State"]
    data = [
        ["Total headcount", 46, 52, 62],
        ["Monthly loaded payroll (EUR)", 79354, 90244, 111000],
        ["Annual loaded payroll (EUR)", 952248, 1082928, 1332000],
        ["Gross annual savings (EUR)", 379752, 249072, "—"],
        ["Revenue loss (EUR/year)", -126000, -41000, "—"],
        ["Turnover/redundancy costs (EUR)", -64000, -10000, "—"],
        ["NET annual benefit (EUR)", 211750, 259182, "—"],
    ]
    r = write_table(ws, r, headers, data, fmt=EUR_FMT)

    r += 1
    ws.merge_cells(f"A{r}:I{r}")
    ws.cell(row=r, column=1, value="Recommended Glide Path").font = SUBTITLE_FONT
    r += 2

    headers = ["Phase", "Timeline", "Target HC%", "Headcount", "How"]
    data = [
        ["Phase 1", "Month 0-6", "40%", 52, "Restructure: 3 RMs, hub/spoke, 10 redundancies"],
        ["Phase 2", "Month 6-12", "37%", 49, "Revenue growth reduces HC% organically"],
        ["Phase 3", "Month 12-18", "35%", "45-47", "Natural attrition + continued growth"],
    ]
    r = write_table(ws, r, headers, data)

    r += 1
    ws.merge_cells(f"A{r}:I{r}")
    ws.cell(row=r, column=1, value="Cumulative 2-Year Savings").font = SUBTITLE_FONT
    r += 2
    headers = ["Approach", "Year 1", "Year 2", "Total"]
    data = [
        ["Glide path (40% → 35%)", 259000, 335000, 594000],
        ["Immediate 35%", 212000, 300000, 512000],
        ["Difference", 47000, 35000, 82000],
    ]
    r = write_table(ws, r, headers, data, fmt=EUR_FMT)

    auto_width(ws)


def build_revenue_sheet(wb):
    ws = wb.create_sheet("Monthly Revenue")
    ws.sheet_properties.tabColor = "548235"

    ws.merge_cells("A1:N1")
    ws.cell(row=1, column=1, value="Monthly Revenue by Hotel (EUR, Services ex-VAT)").font = TITLE_FONT

    r = 3
    headers = ["Hotel"] + MONTHS + ["6M Avg", "12M Avg"]
    data = []
    totals = [0] * 13
    for h in HOTELS:
        rev = REVENUE[h]
        avg_6m = calc_6m_avg(rev)
        avg_12m = calc_12m_avg(rev)
        data.append([h] + rev + [avg_6m, avg_12m])
        for i, v in enumerate(rev):
            totals[i] += v
    total_6m = sum(calc_6m_avg(REVENUE[h]) for h in HOTELS)
    total_12m = sum(calc_12m_avg(REVENUE[h]) for h in HOTELS)
    total_row = ["TOTAL"] + totals + [total_6m, total_12m]
    r = write_table(ws, r, headers, data, fmt=EUR_FMT, total_row=total_row)

    # Highlight seasonality
    r += 1
    ws.merge_cells(f"A{r}:N{r}")
    ws.cell(row=r, column=1, value="Seasonality: Peak = Jul-Oct (green), Low = Dec-Feb (red)").font = SUBTITLE_FONT

    auto_width(ws, min_width=9, max_width=14)


def build_hc_budget_sheet(wb, pct, tab_color):
    pct_label = f"{int(pct*100)}%"
    ws = wb.create_sheet(f"HC Budget {pct_label}")
    ws.sheet_properties.tabColor = tab_color

    ws.merge_cells("A1:N1")
    ws.cell(row=1, column=1, value=f"HC Budget at {pct_label} Target (EUR/month)").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Formula: Revenue x {pct}").font = Font(italic=True, color="666666")

    r = 4
    headers = ["Hotel"] + MONTHS + ["Avg Budget"]
    data = []
    totals = [0] * 13
    for h in HOTELS:
        rev = REVENUE[h]
        budget = [round(v * pct) for v in rev]
        avg_b = round(sum(budget[i] for i in range(13) if rev[i] > 0) / max(sum(1 for v in rev if v > 0), 1))
        data.append([h] + budget + [avg_b])
        for i, v in enumerate(budget):
            totals[i] += v
    total_avg = sum(totals) // 13
    total_row = ["TOTAL"] + totals + [total_avg]
    r = write_table(ws, r, headers, data, fmt=EUR_FMT, total_row=total_row)

    # What it can afford
    r += 1
    ws.merge_cells(f"A{r}:H{r}")
    ws.cell(row=r, column=1, value=f"What {pct_label} Can Afford (Recent 6M Avg Revenue)").font = SUBTITLE_FONT
    r += 2

    staff_dict = STAFF_35 if pct == 0.35 else STAFF_40
    headers2 = ["Hotel", "6M Avg Rev", f"{pct_label} Budget", "RM Alloc", "Net for Staff", "Staff Cost", "Total Cost", "HC%", "Headcount"]
    data2 = []
    for h in HOTELS:
        avg_rev = calc_6m_avg(REVENUE[h])
        budget = round(avg_rev * pct)
        rm = RM_ALLOC[h]
        net = budget - rm
        s = staff_dict[h]
        total_cost = s["staff_cost"] + rm
        hc = total_cost / avg_rev if avg_rev > 0 else 0
        data2.append([h, avg_rev, budget, rm, net, s["staff_cost"], total_cost, hc, s["headcount"]])
    r = write_table(ws, r, headers2, data2)

    # Format HC% column
    for row_idx in range(r - len(data2) - 1, r - 1):
        cell = ws.cell(row=row_idx, column=8)
        cell.number_format = PCT_FMT
        if isinstance(cell.value, (int, float)):
            if cell.value <= pct:
                cell.fill = GREEN_FILL
            elif cell.value <= pct + 0.05:
                cell.fill = YELLOW_FILL
            else:
                cell.fill = RED_FILL

    auto_width(ws, min_width=10, max_width=14)


def build_staffing_comparison(wb):
    ws = wb.create_sheet("Staffing Comparison")
    ws.sheet_properties.tabColor = "BF8F00"

    ws.merge_cells("A1:K1")
    ws.cell(row=1, column=1, value="Staffing Comparison: Current vs 35% vs 40%").font = TITLE_FONT

    r = 3
    headers = ["Hotel", "Current HC", "HC at 35%", "HC at 40%", "Cost 35% (EUR)", "Cost 40% (EUR)",
               "Current Cost (EUR)", "Savings 35%", "Savings 40%", "Extra 5% Buys"]
    data = []
    extra_desc = {
        "Inter": "+1 FT (Matilde), +1 PT (Milena)",
        "Hugos": "+1 FT (Tina), +1 PT (Linara)",
        "Ramla": "+1 FT (Laura Camila stays)",
        "Hyatt": "Natasha H stays FT (not PT)",
        "Riviera": "+1 PT for weekends",
        "Odycy": "+1 PT for weekends",
        "Excelsior": "+1 FT (Sebastian stays)",
        "Novotel": "Vanessa stays FT (not PT)",
    }
    t_current = t_35 = t_40 = t_c35 = t_c40 = t_cc = 0
    for h in HOTELS:
        cur = CURRENT_STAFF[h]
        s35 = STAFF_35[h]
        s40 = STAFF_40[h]
        cost_35 = s35["staff_cost"] + RM_ALLOC[h]
        cost_40 = s40["staff_cost"] + RM_ALLOC[h]
        # Estimate current cost
        cur_cost = SENSITIVITY[h]["current_cost"]
        data.append([h, cur, s35["headcount"], s40["headcount"],
                     cost_35, cost_40, cur_cost,
                     cur_cost - cost_35, cur_cost - cost_40,
                     extra_desc.get(h, "")])
        t_current += cur
        t_35 += s35["headcount"]
        t_40 += s40["headcount"]
        t_c35 += cost_35
        t_c40 += cost_40
        t_cc += cur_cost

    # Add float + RM rows
    data.append(["Float Pool", 2, 2, 2, 3630, 3630, 3630, 0, 0, "Same at both targets"])
    data.append(["RMs", 0, 3, 3, 10230, 10230, 0, -10230, -10230, "New regional management layer"])

    total_row = ["TOTAL", t_current + 2, t_35 + 5, t_40 + 5,
                 t_c35 + 3630 + 10230, t_c40 + 3630 + 10230, t_cc + 3630,
                 (t_cc + 3630) - (t_c35 + 3630 + 10230),
                 (t_cc + 3630) - (t_c40 + 3630 + 10230), ""]
    r = write_table(ws, r, headers, data, fmt=EUR_FMT, total_row=total_row)

    auto_width(ws, min_width=10, max_width=22)


def build_role_costs(wb):
    ws = wb.create_sheet("Role Costs")
    ws.sheet_properties.tabColor = "7030A0"

    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value="Per-Role Loaded Monthly Costs (EUR)").font = TITLE_FONT

    r = 3
    headers = ["Role", "Base", "Avg Commission", "Employer NI (10%)", "Total Monthly", "Annual"]
    data = []
    for role, costs in ROLE_COSTS.items():
        data.append([role, costs["base"], costs["commission"], costs["ni"], costs["total"], costs["annual"]])
    r = write_table(ws, r, headers, data, fmt=EUR_FMT)

    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(row=r, column=1, value="RM Cost Allocation by Hotel (Revenue-Proportional)").font = SUBTITLE_FONT
    r += 2

    headers2 = ["Hotel", "Region", "RM", "RM Share (EUR/mo)"]
    rm_data = [
        ["Inter", "Region 1", "Neli", RM_ALLOC["Inter"]],
        ["Hugos", "Region 1", "Neli", RM_ALLOC["Hugos"]],
        ["Hyatt", "Region 1", "Neli", RM_ALLOC["Hyatt"]],
        ["Ramla", "Region 2", "Kristina", RM_ALLOC["Ramla"]],
        ["Riviera", "Region 2", "Kristina", RM_ALLOC["Riviera"]],
        ["Odycy", "Region 2", "Kristina", RM_ALLOC["Odycy"]],
        ["Excelsior", "Region 3", "Melanie", RM_ALLOC["Excelsior"]],
        ["Novotel", "Region 3", "Melanie", RM_ALLOC["Novotel"]],
    ]
    r = write_table(ws, r, headers2, rm_data, fmt=EUR_FMT,
                    total_row=["TOTAL", "", "", sum(RM_ALLOC.values())])

    auto_width(ws)


def build_pressure_test(wb):
    ws = wb.create_sheet("Pressure Test")
    ws.sheet_properties.tabColor = "C00000"

    ws.merge_cells("A1:J1")
    ws.cell(row=1, column=1, value="Pressure Test: Revenue Feasibility").font = TITLE_FONT

    # Peak utilization
    r = 3
    ws.merge_cells(f"A{r}:J{r}")
    ws.cell(row=r, column=1, value="Peak Month Break-Even Utilization (must stay below 75%)").font = SUBTITLE_FONT
    r += 2

    headers = ["Hotel", "Peak Util at 40%", "Risk at 40%", "Peak Util at 35%", "Risk at 35%"]
    data = []
    for h in HOTELS:
        data.append([h, PEAK_UTIL[h]["40%"], RISK_LEVELS[h]["40%"],
                     PEAK_UTIL[h]["35%"], RISK_LEVELS[h]["35%"]])
    r = write_table(ws, r, headers, data)

    # Color-code utilization and risk
    for row_idx in range(r - len(data) - 1, r - 1):
        for col in [2, 4]:  # util columns
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = PCT_FMT
                if cell.value <= 0.65:
                    cell.fill = GREEN_FILL
                elif cell.value <= 0.80:
                    cell.fill = YELLOW_FILL
                else:
                    cell.fill = RED_FILL
        for col in [3, 5]:  # risk columns
            cell = ws.cell(row=row_idx, column=col)
            val = str(cell.value or "")
            if "CRITICAL" in val:
                cell.fill = RED_FILL
            elif "HIGH" in val:
                cell.fill = ORANGE_FILL
            elif "MEDIUM" in val:
                cell.fill = YELLOW_FILL
            else:
                cell.fill = GREEN_FILL

    # Financial sensitivity
    r += 1
    ws.merge_cells(f"A{r}:J{r}")
    ws.cell(row=r, column=1, value="Financial Sensitivity: Net Monthly Benefit by Hotel").font = SUBTITLE_FONT
    r += 2

    headers2 = ["Hotel", "Current Staff", "Current Cost", "Staff at 40%", "Cost at 40%",
                "Rev Risk 40%", "Net Benefit 40%", "Staff at 35%", "Cost at 35%",
                "Rev Risk 35%", "Net Benefit 35%"]
    data2 = []
    for h in HOTELS:
        s = SENSITIVITY[h]
        data2.append([h, s["current_staff"], s["current_cost"],
                      s["40_staff"], s["40_cost"], s["40_risk"], s["40_net"],
                      s["35_staff"], s["35_cost"], s["35_risk"], s["35_net"]])
    r = write_table(ws, r, headers2, data2, fmt=EUR_FMT)

    # Consolidated financial impact
    r += 1
    ws.merge_cells(f"A{r}:J{r}")
    ws.cell(row=r, column=1, value="Consolidated Annual Financial Impact").font = SUBTITLE_FONT
    r += 2

    headers3 = ["Item", "At 40% (EUR/year)", "At 35% (EUR/year)"]
    data3 = [
        ["Payroll savings (gross)", 309732, 402000],
        ["Revenue loss — walk-in/retail", -40800, -66000],
        ["Revenue loss — capacity constraints", 0, -36000],
        ["Revenue loss — quality degradation", 0, -24000],
        ["Turnover cost", -9750, -16250],
        ["Redundancy costs (amortized)", 0, -48000],
    ]
    total3 = ["NET ANNUAL BENEFIT", 259182, 211750]
    r = write_table(ws, r, headers3, data3, fmt=EUR_FMT, total_row=total3)

    auto_width(ws, min_width=12, max_width=22)


def build_roster_sheet(wb, roster_data, pct_label, tab_color):
    ws = wb.create_sheet(f"Rosters {pct_label}")
    ws.sheet_properties.tabColor = tab_color

    ws.merge_cells("A1:J1")
    ws.cell(row=1, column=1, value=f"7-Day Rosters — {pct_label} HC Target").font = TITLE_FONT

    ws.cell(row=2, column=1, value="E = Early (9:00-17:30) | L = Late (11:30-20:00) | E4 = Morning PT (9:00-13:00) | E6 = Extended (9:00-15:30) | OFF = Rest Day").font = Font(italic=True, size=10, color="666666")

    r = 4
    for hotel in HOTELS:
        if hotel not in roster_data:
            continue
        rd = roster_data[hotel]
        ws.merge_cells(f"A{r}:J{r}")
        ws.cell(row=r, column=1, value=hotel.upper()).font = SUBTITLE_FONT
        ws.cell(row=r, column=1).fill = SUBHEADER_FILL
        r += 1

        # Write headers
        for c, h in enumerate(rd["headers"], 1):
            ws.cell(row=r, column=c, value=h)
        style_header_row(ws, r, len(rd["headers"]))
        r += 1

        # Write roster data
        for row_data in rd["data"]:
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = THIN_BORDER
                cell.font = NUM_FONT
                if c == 1:
                    cell.font = BOLD_FONT
                    cell.alignment = Alignment(horizontal="left")
                elif isinstance(val, str):
                    cell.alignment = Alignment(horizontal="center")
                    if val == "OFF":
                        cell.fill = RED_FILL
                        cell.font = Font(bold=True, color="C00000", size=10)
                    elif val.startswith("L"):
                        cell.fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
                    elif val.startswith("E"):
                        cell.fill = GREEN_FILL
            r += 1
        r += 1  # blank row between hotels

    # Float pool
    ws.merge_cells(f"A{r}:J{r}")
    ws.cell(row=r, column=1, value="FLOAT POOL").font = SUBTITLE_FONT
    ws.cell(row=r, column=1).fill = SUBHEADER_FILL
    r += 1
    float_headers = ["Staff", "Role", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Hrs/wk"]
    for c, h in enumerate(float_headers, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, len(float_headers))
    r += 1
    float_data = [
        ["Romelia", "Float FT", "Deployed to highest-demand location each day", "", "", "", "", "", "", 40],
        ["Dzanela", "Float FT", "Deployed to highest-demand location each day", "", "", "", "", "", "", 40],
    ]
    for row_data in float_data:
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
            cell.font = BOLD_FONT if c == 1 else NUM_FONT
        r += 1

    auto_width(ws, min_width=8, max_width=14)


def build_personnel_actions(wb):
    ws = wb.create_sheet("Personnel Actions")
    ws.sheet_properties.tabColor = "FF6600"

    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value="Personnel Actions Required").font = TITLE_FONT

    r = 3
    headers = ["Name", "Current Location", "Current Role", "Action", "Destination"]
    data = [list(p) for p in PERSONNEL]
    r = write_table(ws, r, headers, data)

    # Color-code by action type
    for row_idx in range(4, 4 + len(data)):
        action = str(ws.cell(row=row_idx, column=4).value or "")
        if "Promote" in action or "Transition" in action:
            for c in range(1, 6):
                ws.cell(row=row_idx, column=c).fill = GREEN_FILL
        elif "Redundancy" in action and "OR" not in action:
            for c in range(1, 6):
                ws.cell(row=row_idx, column=c).fill = RED_FILL
        elif "Redeploy" in action or "Reclassify" in action:
            for c in range(1, 6):
                ws.cell(row=row_idx, column=c).fill = YELLOW_FILL
        else:
            for c in range(1, 6):
                ws.cell(row=row_idx, column=c).fill = ORANGE_FILL

    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    ws.cell(row=r, column=1, value="Legend: Green = Promotion | Yellow = Redeployment | Red = Redundancy | Orange = Decision needed").font = Font(italic=True, size=10)

    r += 2
    ws.merge_cells(f"A{r}:E{r}")
    ws.cell(row=r, column=1, value="Summary of Actions").font = SUBTITLE_FONT
    r += 2
    headers2 = ["Category", "Count"]
    data2 = [
        ["Promotions to RM", 3],
        ["Reclassifications (Supervisor → TIC, etc.)", 3],
        ["Redeployments to other locations", 6],
        ["PT conversions", "2-3"],
        ["Redundancies (definite)", "8-10"],
        ["Decisions pending (offer or redundancy)", 3],
    ]
    r = write_table(ws, r, headers2, data2)

    auto_width(ws, min_width=12, max_width=40)


def build_guardrails(wb):
    ws = wb.create_sheet("Guardrails")
    ws.sheet_properties.tabColor = "375623"

    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1, value="Automatic Guardrails & Triggers").font = TITLE_FONT

    r = 3
    headers = ["Trigger", "Action", "Owner"]
    data = [
        ["Hotel revenue drops >15% YoY (outside Dec-Feb)", "Freeze headcount cuts at that hotel", "RM + Mert"],
        ["Company weekly revenue <EUR 33K for 3 weeks", "Emergency review; consider re-hiring", "Mert"],
        ["Google review score <4.5 at any location", "Deploy float therapist for 2 weeks", "RM"],
        [">2 therapist departures in 60 days", "Halt restructuring; conduct stay interviews", "HR + Mert"],
        ["Hotel partner formal complaint", "Immediately add 1 FT therapist", "RM + Mert"],
        ["Utilization >80% at any location for 4+ weeks", "That location is understaffed — add 1 PT", "RM"],
        ["Sick leave >8 days/week for 2+ consecutive weeks", "Hire 1-2 temporary PT therapists", "HR"],
    ]
    r = write_table(ws, r, headers, data)

    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    ws.cell(row=r, column=1, value="Restructuring Sequence (Priority Order)").font = SUBTITLE_FONT
    r += 2
    headers2 = ["Priority", "Hotel(s)", "Action", "Timeline", "Savings/Month"]
    data2 = [
        [1, "Novotel", "Cut to 2 staff immediately", "Week 1-2", 7650],
        [2, "Excelsior", "Cut to 4, transition Melanie to RM3", "Week 2-4", 5620],
        [3, "Odycy", "Resolve dual supervisor, cut to 4", "Week 3-5", 4735],
        [4, "Riviera", "Cut to 4 (TIC model)", "Week 4-8", 6535],
        [5, "Inter", "Right-size reception, rebalance therapists", "Month 2-3", 5192],
        [6, "Hugos", "Transition Neli to RM1, adjust staffing", "Month 2-3", 4607],
        [7, "Hyatt", "Cut to 4 (TIC model)", "Month 2-4", 3120],
        [8, "Ramla", "Right-size to 7 (after RM2 in place)", "Month 3-4", 4427],
    ]
    r = write_table(ws, r, headers2, data2, fmt=EUR_FMT)

    auto_width(ws, min_width=12, max_width=40)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()

    build_executive_summary(wb)
    build_revenue_sheet(wb)
    build_hc_budget_sheet(wb, 0.35, "002060")
    build_hc_budget_sheet(wb, 0.40, "385723")
    build_staffing_comparison(wb)
    build_role_costs(wb)
    build_pressure_test(wb)
    build_roster_sheet(wb, ROSTER_35, "35%", "800000")
    build_roster_sheet(wb, ROSTER_40, "40%", "006100")
    build_personnel_actions(wb)
    build_guardrails(wb)

    out_dir = os.path.join(".tmp", "hr_analysis")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "HC_Cost_Analysis_35_vs_40.xlsx")
    wb.save(out_path)
    print(f"Workbook saved to: {out_path}")
    print(f"Sheets: {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    main()
